from __future__ import annotations

import argparse
from pathlib import Path


def main() -> int:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        print("ERROR: openpyxl not available:", e)
        return 2

    parser = argparse.ArgumentParser(description="Inspect an XLSX (headers, formulas, and table samples).")
    parser.add_argument(
        "xlsx",
        nargs="?",
        default=None,
        help="Path to XLSX file (default: ./Ведомость_по_тревогам_Новая_сентябрь.xlsx)",
    )
    args = parser.parse_args()

    path = Path(args.xlsx) if args.xlsx else (Path(__file__).resolve().parents[1] / "Ведомость_по_тревогам_Новая_сентябрь.xlsx")
    print("path=", path)
    if not path.exists():
        print("ERROR: file not found")
        return 1

    print("size=", path.stat().st_size)

    wb = load_workbook(path, read_only=True, data_only=False)
    print("sheets=", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    print("active=", ws.title)
    print("max_row=", ws.max_row, "max_col=", ws.max_column)

    max_c = min(ws.max_column or 80, 40)

    # Header-ish rows
    shown = 0
    for r in range(1, 80):
        row = [ws.cell(r, c).value for c in range(1, max_c + 1)]
        if any(v not in (None, "") for v in row):
            while row and row[-1] in (None, ""):
                row.pop()
            print(f"ROW {r}:", row)
            shown += 1
            if shown >= 25:
                break

    # Formula scan (bounded)
    formula_samples: list[tuple[int, int, str]] = []
    formula_count = 0
    for r in range(1, min(ws.max_row or 1, 500) + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                formula_count += 1
                if len(formula_samples) < 40:
                    formula_samples.append((r, c, v))

    print("formula_count=", formula_count)
    print("formula_samples=", formula_samples)

    # Summarize the main data table (heuristic based on observed layout)
    # Columns: B=Date, C=Dispatchers, D=Shift, E=Name, F=Alarms, G=Percent, H=Sum, I=Bonus, J=Total
    def v(row: int, col: int):
        return ws.cell(row, col).value

    data_rows: list[dict] = []
    for r in range(1, (ws.max_row or 1) + 1):
        shift = v(r, 4)
        name = v(r, 5)
        alarms = v(r, 6)
        dt = v(r, 2)
        if shift in ("день", "ночь") and isinstance(name, str) and name.strip():
            data_rows.append(
                {
                    "r": r,
                    "date": dt,
                    "shift": shift,
                    "name": name,
                    "alarms": alarms,
                    "bonus": v(r, 9),
                }
            )

    print("data_rows=", len(data_rows))
    if data_rows:
        bonus_values = {}
        for x in data_rows:
            b = x.get("bonus")
            bonus_values[str(b)] = bonus_values.get(str(b), 0) + 1
        print("bonus_values=", dict(sorted(bonus_values.items(), key=lambda kv: (-kv[1], kv[0]))))
        print("data_samples=", data_rows[:10])

        # Try to read cached computed values (if the xlsx was saved with them)
        wb_vals = load_workbook(path, read_only=True, data_only=True)
        ws_vals = wb_vals[wb_vals.sheetnames[0]]
        samples2 = []
        for x in data_rows[:15]:
            r = x["r"]
            samples2.append(
                {
                    "r": r,
                    "date": ws_vals.cell(r, 2).value,
                    "shift": ws_vals.cell(r, 4).value,
                    "name": ws_vals.cell(r, 5).value,
                    "alarms": ws_vals.cell(r, 6).value,
                    "dispatchers": ws_vals.cell(r, 3).value,
                    "percent": ws_vals.cell(r, 7).value,
                    "sum": ws_vals.cell(r, 8).value,
                    "bonus": ws_vals.cell(r, 9).value,
                    "total": ws_vals.cell(r, 10).value,
                }
            )
        print("computed_samples=", samples2)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
