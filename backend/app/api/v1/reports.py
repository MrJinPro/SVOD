from __future__ import annotations

import asyncio
import json
import logging
from pathlib import Path
from time import monotonic
from typing import Any, Awaitable, Callable, Iterator, cast
from uuid import uuid4

from datetime import datetime, timezone
from datetime import date as date_type
from datetime import time as time_type
from datetime import timedelta
import re

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from starlette.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import Integer, and_, case, cast as sql_cast, func, or_, select

from app.api.v1.deps import get_current_user
from app.core.config import settings
from app.db.session import SessionLocal, get_session
from app.services.report_service import export_daily_report_csv, today_str
from app.models.event_action import EventAction
from app.models.event import Event
from app.models.object import Object
from app.models.report import Report

router = APIRouter(prefix="/reports")
logger = logging.getLogger(__name__)


def _build_report_worker_logger() -> logging.Logger:
    report_logger = logging.getLogger("app.reports.worker")
    if report_logger.handlers:
        return report_logger

    log_path = Path(__file__).resolve().parents[3] / "report_worker.log"
    handler = logging.FileHandler(log_path, encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    report_logger.addHandler(handler)
    report_logger.setLevel(logging.INFO)
    report_logger.propagate = False
    return report_logger


report_worker_logger = _build_report_worker_logger()


# If report generation is interrupted (server restart/crash), records can stay in
# 'pending' forever. In operator workflow, waiting many minutes is already a
# broken state, so fail stale reports much earlier and let users regenerate.
_REPORT_PENDING_STALE_SECONDS = 10 * 60  # 10 minutes


_PCN_EXCLUDED_ALARM_EXACT_VALUES = (
    "",
    "-",
    "—",
    "отмена тревоги",
)


_PCN_EXCLUDED_ALARM_PATTERNS = (
    "тревога при открытии",
    "тревога при закрытии",
    "на объекте работает инженер",
    "тревожная кнопка - проверка ответственных",
    "тревожная кнопка проверка ответственных",
    "отмена группы",
    "групповая обработка",
    "х/о оповещен",
    "х/о оповещ",
    "х/о на связи",
    "х/о на связи н/нет",
    "оповещен начальник караула",
    "снятие не по расписанию отзвонились",
)


_GBR_ARCHIVE_CANCEL_PATTERNS = (
    "отмен",
    "ложн",
    "свобод",
)


_GBR_EXCLUDED_RESULT_PATTERNS = (
    "сопровожд",
    "на азс",
    "на сто",
    "на обеде",
    "2-е сутки",
    "2 е сутки",
    "2е сутки",
    "развод",
)


_GBR_REAL_NAME_PREFIXES = (
    "булат",
    "гром",
    "накат",
)


def _report_log_value(value: Any, *, limit: int = 1200) -> str:
    try:
        text = json.dumps(value, ensure_ascii=False, default=str, separators=(",", ":"))
    except Exception:
        text = str(value)
    if len(text) > limit:
        return f"{text[:limit]}...<truncated>"
    return text


def _format_report_error(error: Exception | str) -> str:
    if isinstance(error, Exception):
        message = str(error).strip()
        if message:
            return f"{type(error).__name__}: {message}"
        return type(error).__name__
    return str(error)


def _public_alarm_id(*candidates: object) -> str | None:
    for candidate in candidates:
        text = str(candidate or "").strip()
        if not text:
            continue
        if ":" in text:
            tail = text.rsplit(":", 1)[-1].strip()
            if tail.isdigit():
                return tail
            continue
        return text
    return None


def _agency_event_id(event_id: str | None) -> str | None:
    """Extract agency numeric event id from stored local id.

    For MSSQL imports we store ids like: 'mssql:{date_key}:{event_id}'.
    """

    if not event_id:
        return None
    parts = str(event_id).split(":")
    if len(parts) >= 3:
        return parts[-1] or None
    return None


def _numeric_event_id_predicate(dialect_name: str | None) -> Any:
    if (dialect_name or "").lower() == "postgresql":
        return Event.id.op("~")(r"^\d+$")

    return and_(Event.id.op("GLOB")("[0-9]*"), ~Event.id.op("GLOB")("*[^0-9]*"))


def _event_date_key_expr(dialect_name: str | None) -> Any:
    if (dialect_name or "").lower() == "sqlite":
        return sql_cast(func.strftime("%Y%m%d", Event.timestamp), Integer)
    return sql_cast(func.to_char(Event.timestamp, "YYYYMMDD"), Integer)


def _event_action_to_event_join_condition(dialect_name: str | None) -> Any:
    raw_id_expr = case(
        (_numeric_event_id_predicate(dialect_name), sql_cast(Event.id, Integer)),
        else_=None,
    )
    return or_(
        EventAction.event_id == Event.id,
        and_(
            raw_id_expr.is_not(None),
            EventAction.raw_event_id == raw_id_expr,
            EventAction.date_key == _event_date_key_expr(dialect_name),
        ),
    )


def _ensure_reports_manage_perm(current: dict) -> None:
    role = str(current.get("role") or "")
    if role in {"admin", "analyst"}:
        return
    raise HTTPException(status_code=403, detail={"code": "FORBIDDEN", "message": "Missing permissions"})


def _parse_dt(value: str) -> datetime | None:
    try:
        v = (value or "").strip()
        # Frontend often sends UTC ISO with trailing 'Z'.
        if v.endswith("Z"):
            v = v[:-1] + "+00:00"
        dt = datetime.fromisoformat(v)
        # DB stores timestamps as naive datetimes (no timezone). If frontend sends
        # tz-aware ISO strings (e.g. trailing 'Z'), normalize to server local time
        # and drop tzinfo so comparisons match stored values.
        if dt.tzinfo is not None:
            return dt.astimezone().replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _parse_date(value: str) -> date_type | None:
    try:
        return date_type.fromisoformat(value)
    except Exception:
        return None


def _parse_hhmm(value: str, *, default: time_type) -> time_type:
    v = (value or "").strip()
    if not v:
        return default
    try:
        parts = v.split(":")
        hh = int(parts[0])
        mm = int(parts[1]) if len(parts) > 1 else 0
        return time_type(hour=hh, minute=mm)
    except Exception:
        return default


def _time_to_minutes(value: time_type) -> int:
    return int(value.hour) * 60 + int(value.minute)


def _minutes_to_time(value: int) -> time_type:
    total = int(value) % (24 * 60)
    return time_type(hour=total // 60, minute=total % 60)


def _shift_bucket(
    ts: datetime,
    *,
    day_start: time_type,
    day_end: time_type,
    night_start: time_type,
    night_end: time_type,
) -> tuple[date_type, str]:
    """Returns (shift_date, shift_name).

    day shift: [day_start, day_end)
    night shift: [night_start, next_day night_end)

    Gaps between day/night windows are attributed to the upcoming shift, not to the
    previous one. This prevents wide handover windows from dragging operators into
    both shifts at once.
    """

    t = ts.time()
    d = ts.date()
    if day_start <= t < day_end:
        return (d, "день")
    if t >= night_start:
        return (d, "ночь")
    if t < night_end:
        return (d - timedelta(days=1), "ночь")
    if day_end <= t < night_start:
        return (d, "ночь")
    return (d - timedelta(days=1), "ночь")


def _pcn_ledger_payout(
    dispatchers: int,
    percent: float,
    *,
    payouts: tuple[int, int, int, int],
    thresholds: dict[int, tuple[int, int, int]],
) -> int:
    """Payout based on percent and dispatcher count.

    Logic matches the Excel formula shape:
    - < t1 -> pay0
    - < t2 -> pay1
    - <= t3 -> pay2
    - else -> pay3
    """

    dispatchers_i = int(dispatchers)

    # Business matrix is configured only for 3/4/5 dispatchers.
    # If the staffing source undercounts (for example presence sees only 1 user
    # while alarm actions show 3 operators), we should not collapse payout to 0.
    # Clamp to the nearest configured bracket instead of returning pay0.
    if dispatchers_i <= 3:
        t = thresholds.get(3)
    elif dispatchers_i == 4:
        t = thresholds.get(4)
    else:
        t = thresholds.get(5)

    if not t:
        return int(payouts[0])
    t1, t2, t3 = t
    pay0, pay1, pay2, pay3 = payouts
    if percent < t1:
        return int(pay0)
    if percent < t2:
        return int(pay1)
    if percent <= t3:
        return int(pay2)
    return int(pay3)


def _pcn_excluded_alarm_predicate() -> Any:
    normalized_result = func.lower(func.trim(func.coalesce(Event.result_text, "")))
    checks: list[Any] = [normalized_result.in_(_PCN_EXCLUDED_ALARM_EXACT_VALUES)]
    for pattern in _PCN_EXCLUDED_ALARM_PATTERNS:
        like = f"%{pattern}%"
        checks.extend(
            [
                func.lower(func.coalesce(Event.result_text, "")).like(like),
                func.lower(func.coalesce(Event.description, "")).like(like),
                func.lower(func.coalesce(Event.state_name, "")).like(like),
            ]
        )
    return or_(*checks)


def _pcn_accept_action_predicate(action_name: str | None) -> Any:
    checks: list[Any] = [
        EventAction.action_name == "Прием на обработку",
        EventAction.action_name.ilike("%прин%в обработ%"),
        EventAction.action_name.ilike("%прием%обработ%"),
    ]
    act = str(action_name or "").strip()
    if act:
        checks.extend(
            [
                EventAction.action_name == act,
                EventAction.action_name.ilike(f"%{act}%"),
            ]
        )
    return or_(*checks)


def _is_real_gbr_name(value: object) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    return any(text.startswith(prefix) for prefix in _GBR_REAL_NAME_PREFIXES)


def _resolve_manual_operator_names(
    requested_names: set[str],
    candidate_names: set[str],
) -> tuple[set[str], set[str]]:
    resolved: set[str] = set()
    unresolved: set[str] = set()

    def _norm_name(value: object) -> str:
        text = str(value or "").strip().lower()
        if not text:
            return ""
        text = text.replace(".", " ")
        text = re.sub(r"\s+", " ", text).strip()
        return text

    def _norm_tokens(value: object) -> list[str]:
        text = _norm_name(value)
        if not text:
            return []
        return [token for token in text.split(" ") if token]

    def _tokens_match(wanted_tokens: list[str], candidate_tokens: list[str]) -> bool:
        if not wanted_tokens or not candidate_tokens:
            return False
        matched = 0
        for wanted in wanted_tokens:
            found = False
            for candidate in candidate_tokens:
                if candidate.startswith(wanted) or wanted.startswith(candidate):
                    found = True
                    matched += 1
                    break
            if not found:
                return False
        return matched == len(wanted_tokens)

    normalized_candidates = {
        candidate: _norm_name(candidate)
        for candidate in candidate_names
        if _norm_name(candidate)
    }

    for raw_name in requested_names:
        wanted = str(raw_name or "").strip()
        if not wanted:
            continue
        wanted_norm = _norm_name(wanted)

        # 1) Exact match (safe)
        exact = {candidate for candidate, candidate_norm in normalized_candidates.items() if wanted_norm == candidate_norm}
        if len(exact) == 1:
            resolved.update(exact)
            continue
        if len(exact) > 1:
            # Unlikely, but treat as ambiguous.
            unresolved.add(wanted)
            continue

        # 2) Substring match only if it resolves to exactly one operator.
        partial = {
            candidate
            for candidate, candidate_norm in normalized_candidates.items()
            if wanted_norm and (wanted_norm in candidate_norm or candidate_norm in wanted_norm)
        }
        if len(partial) == 1:
            resolved.update(partial)
            continue

        # 3) Token/prefix match for cases like "Иванов И.И." vs "Иванов Иван Иванович".
        wanted_tokens = _norm_tokens(wanted)
        token_matches = {
            candidate
            for candidate, candidate_norm in normalized_candidates.items()
            if _tokens_match(wanted_tokens, _norm_tokens(candidate_norm))
        }
        if len(token_matches) == 1:
            resolved.update(token_matches)
        else:
            # 0 matches or >1 matches (ambiguous surname-only, etc.)
            unresolved.add(wanted)

    return resolved, unresolved


def _pcn_ledger_title(
    period_start_date: date_type,
    period_end_date: date_type,
    operator_query: str | None,
) -> str:
    operator_label = (operator_query or "").strip()
    if (
        period_start_date.year == period_end_date.year
        and period_start_date.month == period_end_date.month
        and period_start_date.day == 1
    ):
        period_label = f"за {period_start_date.strftime('%m.%Y')}г."
    else:
        period_label = f"за период {period_start_date.isoformat()}–{period_end_date.isoformat()}"

    if operator_label:
        return f"Ведомость учета работы оператора ПЦН ({operator_label}) {period_label}"
    return f"Ведомость учета работы операторов ПЦН {period_label}"


def _gbr_archive_is_cancelled(status_reason: object) -> bool:
    text = str(status_reason or "").strip().lower()
    if not text:
        return False
    return any(pattern in text for pattern in _GBR_ARCHIVE_CANCEL_PATTERNS)


def _gbr_trip_dedupe_key(row: dict[str, Any]) -> tuple[str, ...] | None:
    gbr_name = str(row.get("gbrName") or "").strip().lower()
    alarm_id = str(row.get("alarmId") or "").strip()
    if gbr_name and alarm_id:
        return ("alarm", gbr_name, alarm_id)

    agency_event_id = str(row.get("agencyEventId") or "").strip()
    if gbr_name and agency_event_id:
        return ("agency", gbr_name, agency_event_id)

    object_id = str(row.get("objectId") or "").strip()
    ts = str(row.get("calledAt") or row.get("lastActionAt") or row.get("arrivedAt") or row.get("cancelledAt") or "").strip()
    if gbr_name and object_id and ts:
        return ("object-time", gbr_name, object_id, ts[:16])

    event_id = str(row.get("eventId") or "").strip()
    if gbr_name and event_id:
        return ("event", gbr_name, event_id)

    return None


def _gbr_archive_row_to_trip(row: dict[str, object]) -> dict[str, object]:
    archive_id = str(row.get("id") or "").strip()
    called_at = row.get("StartTime")
    raw_end = row.get("EndTime")
    status_reason = str(row.get("StatusReason") or "").strip() or None
    cancelled = _gbr_archive_is_cancelled(status_reason)
    arrived_at = raw_end if raw_end is not None and not cancelled else None
    cancelled_at = raw_end if raw_end is not None and cancelled else None

    if cancelled:
        trip_status = "Свободна"
    elif arrived_at is not None:
        trip_status = "На объекте"
    else:
        trip_status = "На выезде"

    object_name = str(row.get("ObjectName") or "").strip() or None
    object_address = str(row.get("ObjectAddress") or "").strip() or None

    return {
        "eventId": f"archive:{archive_id}" if archive_id else "archive:unknown",
        "agencyEventId": archive_id or None,
        "alarmId": archive_id or None,
        "gbrName": str(row.get("GroupName") or "").strip() or "Не указан",
        "calledAt": called_at.isoformat() if isinstance(called_at, datetime) else None,
        "arrivedAt": arrived_at.isoformat() if isinstance(arrived_at, datetime) else None,
        "cancelledAt": cancelled_at.isoformat() if isinstance(cancelled_at, datetime) else None,
        "lastActionAt": raw_end.isoformat() if isinstance(raw_end, datetime) else (called_at.isoformat() if isinstance(called_at, datetime) else None),
        "objectId": str(row.get("Panel_id") or "").strip() or None,
        "objectName": object_name,
        "address": object_address,
        "clientName": object_name,
        "responsibleName": None,
        "calledOperator": None,
        "travelSeconds": row.get("DurationSeconds"),
        "meterCount": None,
        "timeMeterCount": None,
        "resultText": status_reason,
        "tripStatus": trip_status,
    }


def _gbr_report_row_is_excluded(row: dict[str, Any]) -> bool:
    text = " ".join(
        [
            str(row.get("tripStatus") or ""),
            str(row.get("resultText") or ""),
            str(row.get("resultInspection") or ""),
        ]
    ).strip().lower()
    if not text:
        return False
    return any(pattern in text for pattern in _GBR_EXCLUDED_RESULT_PATTERNS)


def _backend_root_dir() -> Path:
    # backend/app/api/v1/reports.py -> parents[3] == backend/
    return Path(__file__).resolve().parents[3]


def _reports_store_dir() -> Path:
    return _backend_root_dir() / "reports_store"


def _safe_report_filename(filename: str) -> str:
    name = re.sub(r'[<>:"/\\|?*\x00-\x1F]+', '_', str(filename or '').strip())
    name = re.sub(r'\s+', ' ', name).strip(' .')
    return name or 'report.xlsx'


def _iter_chunks[T](values: list[T], chunk_size: int = 1000) -> Iterator[list[T]]:
    for i in range(0, len(values), chunk_size):
        yield values[i : i + chunk_size]


def _write_report_file(report_id: str, filename: str, content: bytes) -> Path:
    store = _reports_store_dir()
    store.mkdir(parents=True, exist_ok=True)
    safe_name = _safe_report_filename(filename)
    path = store / f"{report_id}-{safe_name}"
    path.write_bytes(content)
    return path


def _utcnow_iso() -> str:
    return datetime.utcnow().isoformat(timespec="seconds")


def _parse_iso_dt_utc_naive(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        s = str(value).strip()
        if not s:
            return None
        # Support both naive and Z-suffixed values.
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _pending_is_stale(r: "Report", *, now_utc_naive: datetime, stale_seconds: int) -> bool:
    if str(getattr(r, "status", "")) != "pending":
        return False
    created = _parse_iso_dt_utc_naive(getattr(r, "generated_at", None))
    if created is None:
        return False
    age = (now_utc_naive - created).total_seconds()
    return age >= float(max(0, int(stale_seconds)))


async def _create_pending_report(
    session: AsyncSession,
    *,
    report_id: str,
    report_type: str,
    period_start: str,
    period_end: str,
    params: dict[str, Any],
) -> Report:
    r = Report(
        id=report_id,
        type=report_type,
        period_start=period_start,
        period_end=period_end,
        generated_at=_utcnow_iso(),
        status="pending",
        events_count=0,
        critical_count=0,
        file_name=None,
        mime_type=None,
        storage_path=None,
        params_json=json.dumps(params, ensure_ascii=False),
        error_message=None,
    )
    session.add(r)
    await session.commit()
    await session.refresh(r)
    return r


async def _mark_report_failed(report_id: str, error: Exception | str, *, session: AsyncSession) -> None:
    r = await session.get(Report, report_id)
    if r is None:
        return
    r.status = "failed"
    r.error_message = _format_report_error(error)
    r.generated_at = _utcnow_iso()
    await session.commit()


def _start_report_worker(
    *,
    report_id: str,
    worker: Callable[[AsyncSession], Awaitable[Any]],
) -> None:
    async def _runner() -> None:
        async with SessionLocal() as session:
            report_type = "unknown"
            period_start = ""
            period_end = ""
            params_text = "{}"
            try:
                report = await session.get(Report, report_id)
                if report is not None:
                    report_type = str(report.type or "unknown")
                    period_start = str(report.period_start or "")
                    period_end = str(report.period_end or "")
                    try:
                        params_text = _report_log_value(json.loads(report.params_json or "{}"))
                    except Exception:
                        params_text = _report_log_value(report.params_json or "")

                report_worker_logger.info(
                    "Report worker started report_id=%s type=%s period_start=%s period_end=%s params=%s",
                    report_id,
                    report_type,
                    period_start,
                    period_end,
                    params_text,
                )
                await worker(session)
                report_worker_logger.info(
                    "Report worker finished report_id=%s type=%s",
                    report_id,
                    report_type,
                )
            except Exception as e:  # noqa: BLE001
                report_worker_logger.exception(
                    "Report worker failed report_id=%s type=%s period_start=%s period_end=%s params=%s",
                    report_id,
                    report_type,
                    period_start,
                    period_end,
                    params_text,
                )
                logger.exception(
                    "Report worker failed report_id=%s type=%s period_start=%s period_end=%s params=%s",
                    report_id,
                    report_type,
                    period_start,
                    period_end,
                    params_text,
                )
                await _mark_report_failed(report_id, e, session=session)

    asyncio.create_task(_runner())


async def _store_generated_report(
    session: AsyncSession,
    *,
    report_id: str,
    report_type: str,
    period_start: str,
    period_end: str,
    filename: str,
    mime_type: str,
    content: bytes,
    events_count: int,
    critical_count: int,
    params: dict[str, Any],
) -> dict:
    path = _write_report_file(report_id, filename, content)

    r = await session.get(Report, report_id)
    if r is None:
        r = Report(
            id=report_id,
            type=report_type,
            period_start=period_start,
            period_end=period_end,
            generated_at=_utcnow_iso(),
            status="generated",
            events_count=int(events_count or 0),
            critical_count=int(critical_count or 0),
            file_name=filename,
            mime_type=mime_type,
            storage_path=str(path),
            params_json=json.dumps(params, ensure_ascii=False),
            error_message=None,
        )
        session.add(r)
    else:
        r.type = report_type
        r.period_start = period_start
        r.period_end = period_end
        r.generated_at = _utcnow_iso()
        r.status = "generated"
        r.events_count = int(events_count or 0)
        r.critical_count = int(critical_count or 0)
        r.file_name = filename
        r.mime_type = mime_type
        r.storage_path = str(path)
        r.params_json = json.dumps(params, ensure_ascii=False)
        r.error_message = None

    await session.commit()
    await session.refresh(r)
    return _as_report_out_dict(r)


def _as_report_out_dict(r: Report) -> dict:
    params: dict = {}
    try:
        if r.params_json:
            params = json.loads(r.params_json)
    except Exception:
        params = {}

    title: str | None = None
    rt = str(r.type)
    if rt == "gbrRaportXlsx":
        gbr = str(params.get("gbrName") or "").strip()
        title = f"Рапорт ГБР по {gbr}" if gbr else "Рапорт ГБР"
    elif rt == "objectsByCode":
        code = str(params.get("eventCode") or "").strip()
        title = f"Объекты по коду {code}" if code else "Объекты по коду"
    elif rt == "daily":
        title = "Суточный отчёт"
    elif rt == "pcnLedger":
        ps = str(r.period_start or "").strip()
        pe = str(r.period_end or "").strip()
        operator = str(params.get("operatorQuery") or "").strip()
        if operator and ps and pe:
            title = f"Ведомость ПЦН по оператору {operator} {ps}–{pe}"
        elif operator:
            title = f"Ведомость ПЦН по оператору {operator}"
        else:
            title = f"Ведомость по тревогам (ПЦН) {ps}–{pe}" if ps and pe else "Ведомость по тревогам (ПЦН)"
    elif rt == "eventsRaportXlsx":
        ps = str(r.period_start or "").strip()
        pe = str(r.period_end or "").strip()
        title = f"Рапорт по событиям {ps}–{pe}" if ps and pe else "Рапорт по событиям"
    elif rt == "alarmMessages":
        ps = str(r.period_start or "").strip()
        pe = str(r.period_end or "").strip()
        title = f"Тревожные сообщения {ps}–{pe}" if ps and pe else "Тревожные сообщения"

    d = {
        "id": str(r.id),
        "type": str(r.type),
        "title": title,
        "periodStart": str(r.period_start),
        "periodEnd": str(r.period_end),
        "generatedAt": str(r.generated_at or ""),
        "status": str(r.status),
        "eventsCount": int(r.events_count or 0),
        "criticalCount": int(r.critical_count or 0),
        "errorMessage": str(r.error_message or "") or None,
        "downloadUrl": None,
        "fileName": r.file_name,
        "mimeType": r.mime_type,
    }
    if r.storage_path:
        d["downloadUrl"] = f"/reports/{r.id}/download"
    return d


def _resolve_and_validate_store_path(storage_path: str) -> Path:
    path = Path(str(storage_path))
    if not path.exists():
        raise HTTPException(status_code=410, detail={"code": "GONE", "message": "Stored file not found"})

    store = _reports_store_dir().resolve()
    try:
        resolved = path.resolve()
    except Exception:
        raise HTTPException(status_code=400, detail={"code": "BAD_PATH", "message": "Invalid file path"})
    if store not in resolved.parents and resolved != store:
        raise HTTPException(status_code=400, detail={"code": "BAD_PATH", "message": "Invalid file path"})
    return resolved


def _file_ext(filename: str | None) -> str:
    if not filename:
        return ""
    f = filename.lower().strip()
    if f.endswith(".csv"):
        return "csv"
    if f.endswith(".xlsx"):
        return "xlsx"
    return ""


def _preview_table_from_csv_bytes(content: bytes, max_rows: int = 200) -> dict:
    import csv
    import io

    text = content.decode("utf-8-sig", errors="replace")
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        return {"kind": "table", "columns": [], "rows": []}

    reader = csv.reader(lines, delimiter=";")
    rows = list(reader)
    columns = [str(x or "") for x in (rows[0] if rows else [])]
    out_rows: list[list[str]] = []
    for r in rows[1 : 1 + max_rows]:
        out_rows.append([str(x or "") for x in r])
    return {"kind": "table", "columns": columns, "rows": out_rows}


def _preview_table_from_xlsx_bytes(
    content: bytes,
    max_rows: int = 200,
    max_cols: int = 50,
    sheet_name: str | None = None,
) -> dict:
    from io import BytesIO

    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(status_code=500, detail={"code": "MISSING_DEP", "message": "openpyxl not installed"})

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet_names = [str(ws.title or "") for ws in wb.worksheets]

    ws = None
    requested_sheet = str(sheet_name or "").strip()
    if requested_sheet:
        for candidate in wb.worksheets:
            if str(candidate.title or "").strip().lower() == requested_sheet.lower():
                ws = candidate
                break
    if ws is None:
        ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        return {"kind": "table", "columns": [], "rows": [], "sheets": [], "sheetName": ""}

    collected: list[list[str]] = []
    for row in ws.iter_rows(values_only=True, max_row=max_rows, max_col=max_cols):
        collected.append(["" if v is None else str(v) for v in row])

    if not collected:
        return {
            "kind": "table",
            "columns": [],
            "rows": [],
            "titleLines": [],
            "sheets": sheet_names,
            "sheetName": str(ws.title or ""),
        }

    def _norm(v: str) -> str:
        return (v or "").strip().lower()

    def _non_empty_count(row: list[str]) -> int:
        return sum(1 for x in row if (x or "").strip())

    # Heuristic: find a header row (for Raport templates it's not row 1).
    header_idx: int = 0
    for i, r in enumerate(collected[: min(len(collected), 50)]):
        low = [_norm(x) for x in r]
        if any("№ объекта" in x or "номер объекта" in x for x in low):
            header_idx = i
            break
        if (
            any(x == "дата" for x in low)
            and any("количество диспетчеров" in x for x in low)
            and any(x == "смена" for x in low)
            and any(x == "фио" for x in low)
            and any(x == "тревоги" for x in low)
        ):
            header_idx = i
            break
        if _non_empty_count(r) >= 3 and any(x in {"адрес", "гбр", "вызов", "прибыл", "оператор"} for x in low):
            header_idx = i
            break

    title_lines: list[str] = []
    if header_idx > 0:
        for r in collected[:header_idx]:
            parts = [x.strip() for x in r if (x or "").strip()]
            if parts:
                title_lines.append(" ".join(parts))

    columns = collected[header_idx]
    rows = collected[header_idx + 1 :]

    # Trim trailing completely empty rows for nicer preview.
    while rows and _non_empty_count(rows[-1]) == 0:
        rows.pop()

    return {
        "kind": "table",
        "columns": columns,
        "rows": rows,
        "titleLines": title_lines,
        "sheets": sheet_names,
        "sheetName": str(ws.title or ""),
    }


def _csv_bytes_to_xlsx_bytes(content: bytes) -> bytes:
    import csv
    import io
    from io import BytesIO

    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(status_code=500, detail={"code": "MISSING_DEP", "message": "openpyxl not installed"})

    text = content.decode("utf-8-sig", errors="replace")
    lines = [l for l in text.splitlines() if l.strip()]
    reader = csv.reader(lines, delimiter=";")

    wb = Workbook()
    ws = cast(Any, wb.active)
    ws.title = "Report"
    for r_idx, row in enumerate(reader, start=1):
        for c_idx, v in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _xlsx_bytes_to_csv_bytes(content: bytes) -> bytes:
    import csv
    from io import BytesIO, StringIO

    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(status_code=500, detail={"code": "MISSING_DEP", "message": "openpyxl not installed"})

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0] if wb.worksheets else None
    buf = StringIO()
    writer = csv.writer(buf, delimiter=";")
    if ws is not None:
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else str(v) for v in row])
    return buf.getvalue().encode("utf-8-sig")


@router.get("")
async def list_reports(
    session: AsyncSession = Depends(get_session),
    _current: dict = Depends(get_current_user),
) -> list[dict]:
    # 1) Stored reports (history)
    stored = (
        await session.execute(
            select(Report)
            .where(Report.type != "daily")
            .order_by(Report.generated_at.desc())
            .limit(200)
        )
    ).scalars().all()

    # Auto-fail stale pending reports (e.g. after server restart).
    now = datetime.utcnow()
    changed = False
    for r in stored:
        if _pending_is_stale(r, now_utc_naive=now, stale_seconds=_REPORT_PENDING_STALE_SECONDS):
            r.status = "failed"
            r.error_message = (
                "Отчёт завис в статусе 'Ожидает' (возможно, сервер перезапускался). "
                "Перегенерируйте отчёт из меню."
            )
            r.generated_at = _utcnow_iso()
            changed = True
    if changed:
        await session.commit()

    out: list[dict] = [_as_report_out_dict(r) for r in stored]

    return out


@router.post("/generate/daily", include_in_schema=False)
async def generate_daily_report(
    date: str = Query(default_factory=today_str, description="YYYY-MM-DD"),
    session: AsyncSession = Depends(get_session),
    _current: dict = Depends(get_current_user),
) -> dict:
    raise HTTPException(
        status_code=410,
        detail={
            "code": "DISABLED",
            "message": "Суточный отчёт отключён. Формируйте отчёты только по согласованным событиям/объектам.",
        },
    )

    # Generate now, but return a record and keep it in history.
    day = _parse_date(date)
    if not day:
        raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Invalid date"})

    report_id = str(uuid4())
    r = Report(
        id=report_id,
        type="daily",
        period_start=day.isoformat(),
        period_end=day.isoformat(),
        generated_at=datetime.utcnow().isoformat(timespec="seconds"),
        status="pending",
        events_count=0,
        critical_count=0,
        file_name=None,
        mime_type=None,
        storage_path=None,
        params_json=json.dumps({"date": day.isoformat()}, ensure_ascii=False),
        error_message=None,
    )
    session.add(r)
    await session.commit()

    try:
        content = await export_daily_report_csv(session=session, date=day.isoformat())
        xlsx = _csv_bytes_to_xlsx_bytes(content)
        filename = f"daily-report-{day.isoformat()}.xlsx"
        path = _write_report_file(report_id, filename, xlsx)

        # fill counts
        dt_from = datetime.combine(day, datetime.min.time())
        dt_to = datetime.combine(day, datetime.max.time())
        counts = (
            await session.execute(
                select(
                    func.count().label("events_count"),
                    func.sum(case((Event.severity == "critical", 1), else_=0)).label("critical_count"),
                ).where(Event.timestamp >= dt_from, Event.timestamp <= dt_to)
            )
        ).first()
        events_count = int((counts[0] if counts else 0) or 0)
        critical_count = int((counts[1] if counts else 0) or 0)

        r.status = "generated"
        r.file_name = filename
        r.mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        r.storage_path = str(path)
        r.events_count = events_count
        r.critical_count = critical_count
        r.generated_at = datetime.utcnow().isoformat(timespec="seconds")
        await session.commit()
    except Exception as e:
        r.status = "failed"
        r.error_message = str(e)
        await session.commit()

    return _as_report_out_dict(r)


@router.post("/generate/objects-by-code")
async def generate_objects_by_code_report(
    eventCode: str = Query(min_length=1, max_length=16, description="Код события, например E1001"),
    dateFrom: str | None = Query(default=None, description="ISO datetime или YYYY-MM-DD"),
    dateTo: str | None = Query(default=None, description="ISO datetime или YYYY-MM-DD"),
    year: int | None = Query(default=None, ge=1970, le=2100, description="Год (если указан — задаёт период)"),
    clientName: str | None = Query(default=None, description="Контрагент/клиент (поиск по подстроке)"),
    objectQuery: str | None = Query(default=None, description="Поиск по объекту/адресу/ID"),
    reportId: str | None = None,
    session: AsyncSession = Depends(get_session),
    _current: dict = Depends(get_current_user),
) -> dict:
    # Generate XLSX now and store (CSV is not generated as an artifact).
    # Build XLSX directly so headers are human-friendly and we can include more columns.
    dt_from: datetime | None = None
    dt_to: datetime | None = None

    if year is not None:
        dt_from = datetime(year, 1, 1, 0, 0, 0)
        dt_to = datetime(year, 12, 31, 23, 59, 59, 999999)

    if dateFrom:
        parsed_dt = _parse_dt(dateFrom)
        parsed_d = _parse_date(dateFrom)
        if parsed_dt:
            dt_from = parsed_dt
        elif parsed_d:
            dt_from = datetime.combine(parsed_d, datetime.min.time())
    if dateTo:
        parsed_dt = _parse_dt(dateTo)
        parsed_d = _parse_date(dateTo)
        if parsed_dt:
            dt_to = parsed_dt
        elif parsed_d:
            dt_to = datetime.combine(parsed_d, datetime.max.time())

    ps = (dt_from.date().isoformat() if dt_from else (str(year) if year else ""))
    pe = (dt_to.date().isoformat() if dt_to else (str(year) if year else ""))
    if not ps:
        ps = date_type.today().isoformat()
    if not pe:
        pe = ps

    if reportId is None:
        report_id = str(uuid4())
        pending = await _create_pending_report(
            session,
            report_id=report_id,
            report_type="objectsByCode",
            period_start=ps,
            period_end=pe,
            params={
                "eventCode": eventCode.strip(),
                "dateFrom": dateFrom,
                "dateTo": dateTo,
                "year": year,
                "clientName": clientName,
                "objectQuery": objectQuery,
            },
        )
        _start_report_worker(
            report_id=report_id,
            worker=lambda bg_session: generate_objects_by_code_report(
                eventCode=eventCode,
                dateFrom=dateFrom,
                dateTo=dateTo,
                year=year,
                clientName=clientName,
                objectQuery=objectQuery,
                reportId=report_id,
                session=bg_session,
                _current=_current,
            ),
        )
        return _as_report_out_dict(pending)

    filters: list[Any] = [Event.code == eventCode.strip()]
    if dt_from is not None:
        filters.append(Event.timestamp >= dt_from)
    if dt_to is not None:
        filters.append(Event.timestamp <= dt_to)

    client = (clientName or "").strip()
    if client:
        needle = f"%{client}%"
        filters.append(or_(Object.client_name.ilike(needle), Event.client_name.ilike(needle)))

    obj_q = (objectQuery or "").strip()
    if obj_q:
        needle = f"%{obj_q}%"
        filters.append(
            or_(
                Event.object_id.ilike(needle),
                Object.name.ilike(needle),
                Object.address.ilike(needle),
                Event.object_name.ilike(needle),
                Event.location.ilike(needle),
            )
        )

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    # best-effort: take one code_text for code
    code_text = (
        await session.execute(select(func.max(Event.code_text)).where(Event.code == eventCode.strip()))
    ).scalar_one_or_none()

    obj_name = func.coalesce(Object.name, Event.object_name)
    obj_addr = func.coalesce(Object.address, Event.location)

    where = and_(*filters)

    base_stmt = (
        select(
            Event.id.label("event_id"),
            func.coalesce(Event.parent_event_id, Event.id).label("alarm_id"),
            Event.object_id.label("object_id"),
            obj_name.label("object_name"),
            obj_addr.label("address"),
            Event.timestamp.label("timestamp"),
            Event.result_text.label("result_text"),
            Event.meter_count.label("meter_count"),
        )
        .select_from(Event)
        .outerjoin(Object, Object.id == Event.object_id)
        .where(where)
    )
    base = base_stmt.subquery("base")

    agg_stmt = (
        select(
            base.c.object_id,
            base.c.object_name,
            base.c.address,
            func.count(func.distinct(base.c.alarm_id)).label("events_count"),
            func.min(base.c.timestamp).label("first_time"),
            func.max(base.c.timestamp).label("last_time"),
        )
        .group_by(base.c.object_id, base.c.object_name, base.c.address)
        .order_by(func.min(base.c.timestamp).asc(), base.c.object_name.asc())
        .limit(200000)
    )
    agg = agg_stmt.subquery("agg")

    rn_stmt = select(
        base.c.object_id.label("object_id"),
        base.c.alarm_id.label("alarm_id"),
        base.c.result_text.label("result_text"),
        base.c.meter_count.label("meter_count"),
        base.c.timestamp.label("timestamp"),
        func.row_number().over(partition_by=base.c.object_id, order_by=base.c.timestamp.desc()).label("rn"),
    )
    rn = rn_stmt.subquery("rn")
    last_note = (
        select(
            rn.c.object_id.label("object_id"),
            rn.c.alarm_id.label("last_event_id"),
            rn.c.result_text.label("last_result_text"),
            rn.c.meter_count.label("last_meter_count"),
        )
        .where(rn.c.rn == 1)
        .subquery("last_note")
    )

    stmt = (
        select(
            agg.c.object_id,
            agg.c.object_name,
            agg.c.address,
            agg.c.events_count,
            agg.c.first_time,
            agg.c.last_time,
            last_note.c.last_event_id,
            last_note.c.last_meter_count,
            last_note.c.last_result_text,
        )
        .select_from(agg)
        .outerjoin(last_note, last_note.c.object_id == agg.c.object_id)
        .order_by(agg.c.events_count.desc())
    )

    rows = (await session.execute(stmt)).all()

    wb = Workbook()
    ws = cast(Any, wb.active)
    ws.title = "Объекты"

    headers = [
        "Номер объекта",
        "Название объекта",
        "Адрес",
        "Количество событий",
        "Код события",
        "Событие",
        "Первое срабатывание",
        "Последнее срабатывание",
        "ID события (аг.)",
        "Параметр (MeterCount)",
        "Комментарий оператора (Result_Text)",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def fmt_ts(dt: datetime | None) -> str:
        if not dt:
            return ""
        return dt.isoformat(sep=" ", timespec="seconds")

    for object_id, object_name, address, events_count, first_time, last_time, last_event_id, last_meter_count, last_result_text in rows:
        ws.append(
            [
                object_id or "",
                object_name or "",
                address or "",
                int(events_count or 0),
                eventCode.strip(),
                code_text or "",
                fmt_ts(first_time),
                fmt_ts(last_time),
                _agency_event_id(last_event_id) or "",
                last_meter_count or "",
                last_result_text or "",
            ]
        )

    total_events = sum(int(events_count or 0) for _object_id, _object_name, _address, events_count, _first_time, _last_time, _last_event_id, _last_meter_count, _last_result_text in rows)
    ws.append(
        [
            "",
            "ИТОГО",
            "",
            total_events,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    total_row_idx = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_idx, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 32
    ws.column_dimensions["K"].width = 35

    out = BytesIO()
    wb.save(out)
    xlsx = out.getvalue()

    report_id = reportId or str(uuid4())
    filename = f"objects-by-code-{eventCode.strip()}-{ps}-{pe}.xlsx"
    return await _store_generated_report(
        session,
        report_id=report_id,
        report_type="objectsByCode",
        period_start=ps,
        period_end=pe,
        filename=filename,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=xlsx,
        events_count=sum(int(x[3] or 0) for x in rows) if rows else 0,
        critical_count=0,
        params={
            "eventCode": eventCode.strip(),
            "dateFrom": dateFrom,
            "dateTo": dateTo,
            "year": year,
            "clientName": clientName,
            "objectQuery": objectQuery,
        },
    )


@router.post("/generate/gbr-raport-xlsx")
async def generate_gbr_raport_xlsx(
    dateFrom: str = Query(description="ISO datetime"),
    dateTo: str = Query(description="ISO datetime"),
    gbrName: str | None = Query(default=None),
    objectId: str | None = Query(default=None),
    reportId: str | None = None,
    session: AsyncSession = Depends(get_session),
    current: dict = Depends(get_current_user),
) -> dict:
    # Permission gate
    have = set(map(str, current.get("permissions") or []))
    if "analytics:read" not in have and current.get("role") != "admin":
        raise HTTPException(status_code=403, detail={"code": "FORBIDDEN", "message": "Missing permissions"})

    from_dt = _parse_dt(dateFrom)
    to_dt = _parse_dt(dateTo)
    if not from_dt or not to_dt:
        raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Invalid date range"})

    ps = from_dt.date().isoformat()
    pe = to_dt.date().isoformat()
    if reportId is None:
        report_id = str(uuid4())
        pending = await _create_pending_report(
            session,
            report_id=report_id,
            report_type="gbrRaportXlsx",
            period_start=ps,
            period_end=pe,
            params={"dateFrom": dateFrom, "dateTo": dateTo, "gbrName": gbrName, "objectId": objectId},
        )
        current_snapshot = dict(current)
        _start_report_worker(
            report_id=report_id,
            worker=lambda bg_session: generate_gbr_raport_xlsx(
                dateFrom=dateFrom,
                dateTo=dateTo,
                gbrName=gbrName,
                objectId=objectId,
                reportId=report_id,
                session=bg_session,
                current=current_snapshot,
            ),
        )
        return _as_report_out_dict(pending)

    # Prefer ArchiveGroupResponse for the main GBR report. It is more reliable than
    # eventservice-derived actions for actual trip history. Fallback to gbr_trips if
    # archive data is unavailable in the current environment.
    from app.api.v1.analytics import gbr_archive_trips, gbr_trips  # local import to avoid circular deps

    page_size = 2000
    max_rows = 50000
    trips_source = "eventservice"
    rows_all: list[dict[str, Any]] = []
    total = 0
    trips: dict[str, Any] = {"data": rows_all, "total": total}

    url = (settings.agency_database_url or "").strip()
    scheme = (url.split(":", 1)[0] or "").lower()
    if url and (scheme.startswith("sqlite") or scheme.startswith("mssql")):
        try:
            archive_payload = await gbr_archive_trips(
                date_from=dateFrom,
                date_to=dateTo,
                group_id=None,
                panel_id=(objectId or None),
                limit=max_rows,
                _perm=current,
            )
            archive_rows = list(archive_payload.get("rows") or [])
            archive_rows = [r for r in archive_rows if _is_real_gbr_name(r.get("GroupName"))]
            if (gbrName or "").strip():
                wanted = (gbrName or "").strip().lower()
                archive_rows = [r for r in archive_rows if str(r.get("GroupName") or "").strip().lower() == wanted]

            rows_all = [
                _gbr_archive_row_to_trip(r)
                for r in archive_rows
                if not _gbr_archive_is_cancelled(r.get("StatusReason"))
            ]
            rows_all.sort(key=lambda x: str(x.get("calledAt") or ""), reverse=True)
            total = len(rows_all)
            trips_source = "archive"
        except Exception:
            rows_all = []
            total = 0

    if trips_source != "archive":
        trips = await gbr_trips(
            date_from=dateFrom,
            date_to=dateTo,
            gbr_name=(gbrName or None),
            object_id=(objectId or None),
            limit=page_size,
            offset=0,
            session=session,
            _perm=current,
        )

        total = int(trips.get("total") or 0)
        if max_rows and total > max_rows:
            raise HTTPException(
                status_code=400,
                detail={
                    "code": "TOO_MANY_ROWS",
                    "message": f"Слишком много строк для рапорта: {total}. Сузьте период/фильтры.",
                },
            )

        rows_all = list(trips.get("data") or [])
        if total > len(rows_all) and len(rows_all) >= page_size:
            offset = page_size
            while len(rows_all) < total:
                out = await gbr_trips(
                    date_from=dateFrom,
                    date_to=dateTo,
                    gbr_name=(gbrName or None),
                    object_id=(objectId or None),
                    limit=page_size,
                    offset=offset,
                    session=session,
                    _perm=current,
                )
                batch = out.get("data") or []
                if not batch:
                    break
                rows_all.extend(batch)
                offset += page_size
                if len(batch) < page_size:
                    break

    # Enrich eventservice rows with alarm id (parent_event_id) and object address.
    event_ids = sorted(
        {
            str(row.get("eventId") or "").strip()
            for row in rows_all
            if str(row.get("eventId") or "").strip() and not str(row.get("eventId") or "").startswith("archive:")
        }
    )
    event_alarm_ids: dict[str, str] = {}
    if event_ids:
        for chunk in _iter_chunks(event_ids):
            q = select(Event.id, Event.parent_event_id).where(Event.id.in_(chunk))
            for event_id, parent_event_id in (await session.execute(q)).all():
                eid = str(event_id or "").strip()
                if eid:
                    event_alarm_ids[eid] = str(parent_event_id or event_id or "").strip()

    object_ids = sorted({str(row.get("objectId") or "").strip() for row in rows_all if str(row.get("objectId") or "").strip()})
    object_meta_by_id: dict[str, tuple[str, str]] = {}
    if object_ids:
        for chunk in _iter_chunks(object_ids):
            q = select(Object.id, Object.name, Object.address).where(Object.id.in_(chunk))
            for object_id, object_name, address in (await session.execute(q)).all():
                object_meta_by_id[str(object_id)] = (str(object_name or "").strip(), str(address or "").strip())

    normalized_rows: list[dict[str, Any]] = []
    for row in rows_all:
        item = dict(row)
        event_id = str(item.get("eventId") or "").strip()
        object_id = str(item.get("objectId") or "").strip()
        item["alarmId"] = _public_alarm_id(
            item.get("alarmId"),
            event_alarm_ids.get(event_id),
            item.get("agencyEventId"),
            event_id,
        )

        object_name = str(item.get("objectName") or "").strip()
        address = str(item.get("address") or "").strip()
        if object_id in object_meta_by_id:
            meta_name, meta_address = object_meta_by_id[object_id]
            object_name = meta_name or object_name
            address = meta_address or address
        item["objectName"] = object_name
        item["address"] = address
        item["resultInspection"] = str(item.get("resultInspection") or item.get("resultText") or item.get("tripStatus") or "").strip()
        normalized_rows.append(item)

    # Business rule: count only trips with "Выезд + Прибытие".
    rows_all = [
        row
        for row in normalized_rows
        if row.get("calledAt") and row.get("arrivedAt") and not row.get("cancelledAt") and not _gbr_report_row_is_excluded(row)
    ]

    # Business rule: 1 тревога = 1 выезд, even if multiple triggers/records exist.
    # Prefer alarm id, then a stable agency alarm id, then crew+object+time, then event id.
    def _pick_better(a: dict[str, Any], b: dict[str, Any]) -> dict[str, Any]:
        # Prefer record that has arrivedAt; then cancelledAt; then calledAt.
        def _rank(x: dict[str, Any]) -> tuple[int, int, int, int]:
            return (
                1 if x.get("arrivedAt") else 0,
                1 if x.get("cancelledAt") else 0,
                1 if x.get("calledAt") else 0,
                1 if x.get("lastActionAt") else 0,
            )

        best = a if _rank(a) >= _rank(b) else b
        other = b if best is a else a

        # Merge: keep earliest calledAt, earliest arrivedAt/cancelledAt, latest lastActionAt.
        def _min_ts(x: str | None, y: str | None) -> str | None:
            if not x:
                return y
            if not y:
                return x
            return x if x <= y else y

        def _max_ts(x: str | None, y: str | None) -> str | None:
            if not x:
                return y
            if not y:
                return x
            return x if x >= y else y

        best["calledAt"] = _min_ts(best.get("calledAt"), other.get("calledAt"))
        best["arrivedAt"] = _min_ts(best.get("arrivedAt"), other.get("arrivedAt"))
        best["cancelledAt"] = _min_ts(best.get("cancelledAt"), other.get("cancelledAt"))
        best["lastActionAt"] = _max_ts(best.get("lastActionAt"), other.get("lastActionAt"))

        # Fill missing descriptive fields.
        for k in [
            "alarmId",
            "objectId",
            "objectName",
            "address",
            "clientName",
            "responsibleName",
            "calledOperator",
            "meterCount",
            "resultText",
            "tripStatus",
            "agencyEventId",
            "eventId",
        ]:
            if not best.get(k) and other.get(k):
                best[k] = other.get(k)

        # Recompute travelSeconds if we now have better timestamps.
        try:
            ca = best.get("calledAt")
            aa = best.get("arrivedAt")
            if ca and aa:
                ca_dt = datetime.fromisoformat(str(ca))
                aa_dt = datetime.fromisoformat(str(aa))
                best["travelSeconds"] = abs((aa_dt - ca_dt).total_seconds())
        except Exception:
            pass
        return best

    dedup: dict[tuple[str, ...], dict[str, Any]] = {}
    for r in rows_all:
        key = _gbr_trip_dedupe_key(r)
        if key is None:
            continue
        if key in dedup:
            dedup[key] = _pick_better(dedup[key], r)
        else:
            dedup[key] = r

    rows_all = list(dedup.values())
    # Stable ordering: by calledAt desc (like analytics), then arrived/cancelled.
    def _sort_key(x: dict[str, Any]) -> tuple[str, str, str]:
        return (
            str(x.get("calledAt") or ""),
            str(x.get("arrivedAt") or ""),
            str(x.get("cancelledAt") or ""),
        )

    rows_all.sort(key=_sort_key)

    trips["data"] = rows_all
    trips["total"] = len(rows_all)

    # Build XLSX similarly to analytics export
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    def clean_excel_text(value: object) -> Any:
        if value is None:
            return ""
        if isinstance(value, str):
            return re.sub(ILLEGAL_CHARACTERS_RE, "", value)
        return value

    columns = [
        "№ объекта",
        "Адрес",
        "Название объекта",
        "Шлейф",
        "Инженер",
        "Результат",
        "Дата",
        "ГБР",
        "Вызов",
        "Прибыл",
        "Время в пути",
        "Результат осмотра",
        "Оператор",
        "Заявка",
        "Штраф",
        "Сработок за полгода",
        "ID события (аг.)",
        "Параметр (MeterCount)",
        "Пометка оператора (Result_Text)",
        "Статус",
        "Результат выезда",
    ]

    wb = Workbook()
    ws = cast(Any, wb.active)
    ws.title = "Рапорт"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    header_title = "Рапорт"
    if (gbrName or "").strip():
        header_title = f"Рапорт ГБР по {(gbrName or '').strip()}"
    ws.cell(row=1, column=1, value=clean_excel_text(header_title)).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    period_text = f"За период: {from_dt.strftime('%d.%m.%Y %H:%M')} — {to_dt.strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws.cell(row=2, column=1, value=clean_excel_text(period_text)).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(columns))
    ws.cell(row=3, column=1, value="оперативная обстановка следующая:").alignment = Alignment(
        horizontal="center"
    )

    header_row = 5
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = [12, 28, 28, 10, 16, 16, 12, 14, 18, 18, 12, 22, 16, 14, 10, 18, 16, 28, 45, 14, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = w

    for col_idx, title in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=title)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    def fmt_ts(ts: str | None) -> str:
        if not ts:
            return ""
        return ts.replace("T", " ")[:19]

    def fmt_date(ts: str | None) -> str:
        if not ts:
            return ""
        return ts[:10].replace("-", ".")

    def fmt_travel(seconds) -> str:
        try:
            if seconds is None:
                return ""
            total = int(round(float(seconds)))
            if total < 0:
                return ""
            h = total // 3600
            m = (total % 3600) // 60
            s = total % 60
            return f"{h:d}:{m:02d}:{s:02d}" if h > 0 else f"{m:d}:{s:02d}"
        except Exception:
            return ""

    rows = trips.get("data") or []
    start_row = header_row + 1
    for i, r in enumerate(rows, start=0):
        row_idx = start_row + i
        called_at = r.get("calledAt")
        arrived_at = r.get("arrivedAt")
        cancelled_at = r.get("cancelledAt")
        values = [
            r.get("objectId") or "",
            r.get("address") or "",
            r.get("objectName") or r.get("clientName") or "",
            "",
            "",
            r.get("tripStatus") or "",
            fmt_date(called_at),
            r.get("gbrName") or "",
            fmt_ts(called_at),
            fmt_ts(arrived_at) if arrived_at else ("Отмена" if cancelled_at else ""),
            fmt_travel(r.get("travelSeconds")),
            r.get("resultInspection") or r.get("resultText") or r.get("tripStatus") or "",
            "",
            "",
            "",
            "",
            r.get("agencyEventId") or "",
            r.get("meterCount") or "",
            r.get("resultText") or "",
            r.get("tripStatus") or "",
            r.get("resultText") or r.get("tripStatus") or "",
        ]
        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=clean_excel_text(v))
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    summary_row_idx = start_row + len(rows)
    ws.merge_cells(start_row=summary_row_idx, start_column=1, end_row=summary_row_idx, end_column=len(columns))
    summary_cell = ws.cell(summary_row_idx, 1, value=clean_excel_text(f"Итого отработанных тревог: {len(rows)}"))
    summary_cell.font = Font(bold=True)
    summary_cell.alignment = Alignment(horizontal="right", vertical="center")
    summary_cell.border = border

    ws.freeze_panes = ws["A6"]
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    out = BytesIO()
    wb.save(out)
    data = out.getvalue()

    report_id = reportId or str(uuid4())
    gbr_part = (gbrName or "").strip()
    if gbr_part:
        filename = f"raport-gbr-{gbr_part}-{ps}-{pe}.xlsx"
    else:
        filename = f"raport-gbr-{ps}-{pe}.xlsx"
    return await _store_generated_report(
        session,
        report_id=report_id,
        report_type="gbrRaportXlsx",
        period_start=ps,
        period_end=pe,
        filename=filename,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=data,
        events_count=len(rows_all),
        critical_count=0,
        params={"dateFrom": dateFrom, "dateTo": dateTo, "gbrName": gbrName, "objectId": objectId, "source": trips_source},
    )


@router.post("/generate/events-raport-xlsx")
async def generate_events_raport_xlsx(
    dateFrom: str = Query(description="ISO datetime"),
    dateTo: str = Query(description="ISO datetime"),
    type: str | None = Query(default=None, description="Event.type filter"),  # noqa: A002
    objectId: str | None = Query(default=None, description="Event.object_id filter"),
    severity: str | None = Query(default=None, description="Event.severity filter"),
    status: str | None = Query(default=None, description="Event.status filter"),
    search: str | None = Query(default=None, description="Free-text search (description/object/client/location)"),
    includeNoise: bool = Query(False, description="Include access/noise events"),
    includeSystem: bool = Query(False, description="Include system-handled events (no operator)"),
    includeCancelled: bool = Query(False, description="Include cancelled events"),
    onlyWithOperatorComment: bool = Query(
        True,
        description="Show only events that have an operator comment (Result_Text)",
    ),
    limit: int = Query(50000, ge=1, le=200000),
    reportId: str | None = None,
    session: AsyncSession = Depends(get_session),
    _current: dict = Depends(get_current_user),
) -> dict:
    from_dt = _parse_dt(dateFrom)
    to_dt = _parse_dt(dateTo)
    if not from_dt or not to_dt or to_dt < from_dt:
        raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Invalid date range"})

    ps = from_dt.date().isoformat()
    pe = to_dt.date().isoformat()
    if reportId is None:
        report_id = str(uuid4())
        pending = await _create_pending_report(
            session,
            report_id=report_id,
            report_type="eventsRaportXlsx",
            period_start=ps,
            period_end=pe,
            params={
                "dateFrom": dateFrom,
                "dateTo": dateTo,
                "type": type,
                "objectId": objectId,
                "severity": severity,
                "status": status,
                "search": search,
                "includeNoise": includeNoise,
                "includeSystem": includeSystem,
                "includeCancelled": includeCancelled,
                "onlyWithOperatorComment": onlyWithOperatorComment,
                "limit": limit,
            },
        )
        _start_report_worker(
            report_id=report_id,
            worker=lambda bg_session: generate_events_raport_xlsx(
                dateFrom=dateFrom,
                dateTo=dateTo,
                type=type,
                objectId=objectId,
                severity=severity,
                status=status,
                search=search,
                includeNoise=includeNoise,
                includeSystem=includeSystem,
                includeCancelled=includeCancelled,
                onlyWithOperatorComment=onlyWithOperatorComment,
                limit=limit,
                reportId=report_id,
                session=bg_session,
                _current=_current,
            ),
        )
        return _as_report_out_dict(pending)

    # Reuse events raport builder (no HTTP call)
    from app.api.v1.events import build_events_raport_xlsx_bytes

    xlsx, events_count = await build_events_raport_xlsx_bytes(
        dateFrom=dateFrom,
        dateTo=dateTo,
        type=type,
        objectId=objectId,
        severity=severity,
        status=status,
        search=search,
        includeNoise=includeNoise,
        includeSystem=includeSystem,
        includeCancelled=includeCancelled,
        onlyWithOperatorComment=onlyWithOperatorComment,
        limit=limit,
        session=session,
    )

    report_id = reportId or str(uuid4())
    filename = f"raport-events-{ps}-{pe}.xlsx"
    return await _store_generated_report(
        session,
        report_id=report_id,
        report_type="eventsRaportXlsx",
        period_start=ps,
        period_end=pe,
        filename=filename,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=xlsx,
        events_count=int(events_count or 0),
        critical_count=0,
        params={
            "dateFrom": dateFrom,
            "dateTo": dateTo,
            "type": type,
            "objectId": objectId,
            "severity": severity,
            "status": status,
            "search": search,
            "includeNoise": includeNoise,
            "includeSystem": includeSystem,
            "includeCancelled": includeCancelled,
            "onlyWithOperatorComment": onlyWithOperatorComment,
            "limit": limit,
        },
    )


@router.post("/generate/alarm-messages-xlsx")
async def generate_alarm_messages_xlsx(
    dateFrom: str = Query(description="ISO datetime"),
    dateTo: str = Query(description="ISO datetime"),
    type: str | None = Query(default=None, description="Event.type filter"),  # noqa: A002
    objectId: str | None = Query(default=None, description="Event.object_id filter"),
    severity: str | None = Query(default=None, description="Event.severity filter"),
    status: str | None = Query(default=None, description="Event.status filter"),
    search: str | None = Query(default=None, description="Free-text search"),
    includeNoise: bool = Query(False, description="Include access/noise events"),
    includeSystem: bool = Query(False, description="Include system-handled events (no operator)"),
    includeCancelled: bool = Query(False, description="Include cancelled events"),
    onlyWithOperatorComment: bool = Query(
        False,
        description="Show only events that have an operator comment (Result_Text)",
    ),
    limit: int = Query(50000, ge=1, le=200000),
    reportId: str | None = None,
    session: AsyncSession = Depends(get_session),
    _current: dict = Depends(get_current_user),
) -> dict:
    from_dt = _parse_dt(dateFrom)
    to_dt = _parse_dt(dateTo)
    if not from_dt or not to_dt or to_dt < from_dt:
        raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Invalid date range"})

    ps = from_dt.date().isoformat()
    pe = to_dt.date().isoformat()
    if reportId is None:
        report_id = str(uuid4())
        pending = await _create_pending_report(
            session,
            report_id=report_id,
            report_type="alarmMessages",
            period_start=ps,
            period_end=pe,
            params={
                "dateFrom": dateFrom,
                "dateTo": dateTo,
                "type": type,
                "objectId": objectId,
                "severity": severity,
                "status": status,
                "search": search,
                "includeNoise": includeNoise,
                "includeSystem": includeSystem,
                "includeCancelled": includeCancelled,
                "onlyWithOperatorComment": onlyWithOperatorComment,
                "limit": limit,
            },
        )
        _start_report_worker(
            report_id=report_id,
            worker=lambda bg_session: generate_alarm_messages_xlsx(
                dateFrom=dateFrom,
                dateTo=dateTo,
                type=type,
                objectId=objectId,
                severity=severity,
                status=status,
                search=search,
                includeNoise=includeNoise,
                includeSystem=includeSystem,
                includeCancelled=includeCancelled,
                onlyWithOperatorComment=onlyWithOperatorComment,
                limit=limit,
                reportId=report_id,
                session=bg_session,
                _current=_current,
            ),
        )
        return _as_report_out_dict(pending)

    from app.api.v1.events import build_alarm_messages_xlsx_bytes

    xlsx, events_count = await build_alarm_messages_xlsx_bytes(
        dateFrom=dateFrom,
        dateTo=dateTo,
        type=type,
        objectId=objectId,
        severity=severity,
        status=status,
        search=search,
        includeNoise=includeNoise,
        includeSystem=includeSystem,
        includeCancelled=includeCancelled,
        onlyWithOperatorComment=onlyWithOperatorComment,
        limit=limit,
        session=session,
    )

    report_id = reportId or str(uuid4())
    filename = f"trevozhnye-soobshcheniya-{ps}-{pe}.xlsx"
    return await _store_generated_report(
        session,
        report_id=report_id,
        report_type="alarmMessages",
        period_start=ps,
        period_end=pe,
        filename=filename,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=xlsx,
        events_count=int(events_count or 0),
        critical_count=0,
        params={
            "dateFrom": dateFrom,
            "dateTo": dateTo,
            "type": type,
            "objectId": objectId,
            "severity": severity,
            "status": status,
            "search": search,
            "includeNoise": includeNoise,
            "includeSystem": includeSystem,
            "includeCancelled": includeCancelled,
            "onlyWithOperatorComment": onlyWithOperatorComment,
            "limit": limit,
        },
    )


@router.post("/generate/pcn-ledger-xlsx")
async def generate_pcn_ledger_xlsx(
    dateFrom: str = Query(description="YYYY-MM-DD"),
    dateTo: str = Query(description="YYYY-MM-DD"),
    dayStart: str | None = Query(default="08:45", description="HH:MM (start of day shift)"),
    dayEnd: str | None = Query(default=None, description="HH:MM (end of day shift)"),
    nightStart: str | None = Query(default="19:50", description="HH:MM (start of night shift)"),
    nightEnd: str | None = Query(default=None, description="HH:MM (end of night shift)"),
    actionName: str | None = Query(default="Прием на обработку", description="EventAction.action_name match"),
    operatorQuery: str | None = Query(default=None, description="Filter by operator name (substring)"),
    manualOperators: list[str] | None = Query(default=None, description="Repeatable operator full names for the whole shift"),
    hideOperatorNames: bool = Query(
        default=False,
        description="If true, operator names are hidden in XLSX (personal data protection)",
    ),
    pay0: int = Query(default=0, ge=0, le=100000, description="Payout for lowest bracket"),
    pay1: int = Query(default=330, ge=0, le=100000, description="Payout for bracket 2"),
    pay2: int = Query(default=430, ge=0, le=100000, description="Payout for bracket 3"),
    pay3: int = Query(default=480, ge=0, le=100000, description="Payout for top bracket"),
    thr3_1: int = Query(default=29, ge=0, le=100, description="3 dispatchers: threshold 1"),
    thr3_2: int = Query(default=36, ge=0, le=100, description="3 dispatchers: threshold 2"),
    thr3_3: int = Query(default=40, ge=0, le=100, description="3 dispatchers: threshold 3"),
    thr4_1: int = Query(default=21, ge=0, le=100, description="4 dispatchers: threshold 1"),
    thr4_2: int = Query(default=27, ge=0, le=100, description="4 dispatchers: threshold 2"),
    thr4_3: int = Query(default=30, ge=0, le=100, description="4 dispatchers: threshold 3"),
    thr5_1: int = Query(default=17, ge=0, le=100, description="5 dispatchers: threshold 1"),
    thr5_2: int = Query(default=23, ge=0, le=100, description="5 dispatchers: threshold 2"),
    thr5_3: int = Query(default=26, ge=0, le=100, description="5 dispatchers: threshold 3"),
    bonusDefault: int = Query(default=500, ge=0, le=100000, description="Default bonus per operator row"),
    bonusOverride: list[str] | None = Query(
        default=None,
        description="Repeatable override: 'ФИО:1000'",
    ),
    includePresenceOnly: bool = Query(
        default=True,
        description="Include operators who were present in the shift even if they have 0 alarms",
    ),
    dispatchersSource: str = Query(
        default="auto",
        description="Dispatcher count source: auto|presence|actions",
    ),
    minPresenceMinutes: int = Query(
        default=30,
        ge=0,
        le=24 * 60,
        description="Minimum presence minutes inside a shift to count as dispatcher",
    ),
    presenceGraceMinutes: int = Query(
        default=15,
        ge=0,
        le=24 * 60,
        description="Grace minutes added after lastSeenAt to close non-ended presence sessions",
    ),
    handoverMinutes: int = Query(
        default=60,
        ge=0,
        le=6 * 60,
        description=(
            "Minutes before shift start treated as handover overlap. "
            "Used to attribute early actions (e.g. 19:48 or 08:30) to the operator's actual shift."
        ),
    ),
    reportId: str | None = None,
    session: AsyncSession = Depends(get_session),
    _current: dict = Depends(get_current_user),
) -> dict:
    # Stored XLSX report: "Ведомость учета работы операторов ПЦН".

    report_log_id = str(reportId or "direct")
    pcn_started_at = monotonic()

    def _pcn_stage(stage: str, **payload: Any) -> None:
        report_worker_logger.info(
            "PCN stage report_id=%s stage=%s elapsed_ms=%s data=%s",
            report_log_id,
            stage,
            int((monotonic() - pcn_started_at) * 1000),
            _report_log_value(payload),
        )

    ds = (dispatchersSource or "auto").strip().lower()
    if ds not in {"auto", "presence", "actions"}:
        raise HTTPException(
            status_code=400,
            detail={"code": "BAD_REQUEST", "message": "dispatchersSource must be one of: auto|presence|actions"},
        )

    def _has_time_component(v: str | None) -> bool:
        return bool(re.search(r"\d{1,2}:\d{2}", (v or "")))

    dt_from: datetime | None = _parse_dt(dateFrom) if _has_time_component(dateFrom) else None
    dt_to: datetime | None = _parse_dt(dateTo) if _has_time_component(dateTo) else None

    d_from: date_type | None = None
    d_to: date_type | None = None
    if dt_from is None:
        parsed_from_dt = _parse_dt(dateFrom)
        d_from = _parse_date(dateFrom) or (parsed_from_dt.date() if parsed_from_dt else None)
    if dt_to is None:
        parsed_to_dt = _parse_dt(dateTo)
        d_to = _parse_date(dateTo) or (parsed_to_dt.date() if parsed_to_dt else None)

    day_start = _parse_hhmm(dayStart or "", default=time_type(8, 45))
    night_start = _parse_hhmm(nightStart or "", default=time_type(19, 50))
    default_day_end = _minutes_to_time(_time_to_minutes(night_start) - 5)
    day_end = _parse_hhmm(dayEnd or "", default=default_day_end)
    night_end = _parse_hhmm(nightEnd or "", default=day_start)

    day_start_m = _time_to_minutes(day_start)
    day_end_m = _time_to_minutes(day_end)
    night_start_m = _time_to_minutes(night_start)
    night_end_m = _time_to_minutes(night_end)
    if not (day_start_m < day_end_m <= night_start_m):
        raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Invalid day shift window"})
    if night_end_m > day_start_m:
        raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Invalid night shift window"})

    exact_window = dt_from is not None or dt_to is not None
    selected_shift_keys: set[tuple[date_type, str]] | None = None
    if exact_window:
        if dt_from is None or dt_to is None or dt_to < dt_from:
            raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Invalid date range"})
        window_start = dt_from
        window_end = dt_to
        period_start_date = dt_from.date()
        period_end_date = dt_to.date()

        # For exact windows we do NOT clamp by shift_date, otherwise early-hours actions
        # (before dayStart) would be incorrectly dropped (they belong to previous day's night shift).
        clamp_shift_dates: tuple[date_type, date_type] | None = None

        # For presence bucketing we need a date span that covers possible shift_dates.
        # shift_bucket can produce (ts.date() - 1) for times before dayStart.
        presence_span_start = window_start.date() - timedelta(days=1)
        presence_span_end = window_end.date()
        if (window_end - window_start) <= timedelta(hours=16):
            selected_shift_keys = {
                _shift_bucket(
                    window_start,
                    day_start=day_start,
                    day_end=day_end,
                    night_start=night_start,
                    night_end=night_end,
                )
            }
    else:
        if not d_from or not d_to or d_to < d_from:
            raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Invalid date range"})
        # Build query window that covers the last night shift up to next day's dayStart.
        handover = timedelta(minutes=int(handoverMinutes))
        window_start = datetime.combine(d_from, day_start) - handover
        window_end = datetime.combine(d_to + timedelta(days=1), day_start) + handover
        period_start_date = d_from
        period_end_date = d_to
        clamp_shift_dates = (d_from, d_to)
        # Presence bucketing needs a span that covers possible shift_dates, including
        # handover overlaps which may attribute early morning actions to previous night.
        presence_span_start = d_from - timedelta(days=1)
        presence_span_end = d_to

    if period_end_date < period_start_date:
        raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Invalid date range"})

    _pcn_stage(
        "window_ready",
        exactWindow=exact_window,
        windowStart=window_start,
        windowEnd=window_end,
        periodStart=period_start_date,
        periodEnd=period_end_date,
        dispatchersSource=ds,
        includePresenceOnly=includePresenceOnly,
    )

    ps = period_start_date.isoformat()
    pe = period_end_date.isoformat()
    if reportId is None:
        report_id = str(uuid4())
        pending = await _create_pending_report(
            session,
            report_id=report_id,
            report_type="pcnLedger",
            period_start=ps,
            period_end=pe,
            params={
                "dateFrom": dateFrom,
                "dateTo": dateTo,
                "dayStart": dayStart,
                "nightStart": nightStart,
                "actionName": actionName,
                "operatorQuery": operatorQuery,
                "manualOperators": sorted(
                    {
                        part.strip()
                        for raw in (manualOperators or [])
                        for part in re.split(r"[\r\n,;]+", str(raw or ""))
                        if part.strip()
                    },
                    key=str.lower,
                ),
                "hideOperatorNames": hideOperatorNames,
                "payouts": {"pay0": pay0, "pay1": pay1, "pay2": pay2, "pay3": pay3},
                "thresholds": {
                    "3": [thr3_1, thr3_2, thr3_3],
                    "4": [thr4_1, thr4_2, thr4_3],
                    "5": [thr5_1, thr5_2, thr5_3],
                },
                "bonusDefault": bonusDefault,
                "bonusOverride": bonusOverride or [],
                "includePresenceOnly": includePresenceOnly,
                "dispatchersSource": dispatchersSource,
                "minPresenceMinutes": minPresenceMinutes,
                "presenceGraceMinutes": presenceGraceMinutes,
            },
        )
        _start_report_worker(
            report_id=report_id,
            worker=lambda bg_session: generate_pcn_ledger_xlsx(
                dateFrom=dateFrom,
                dateTo=dateTo,
                dayStart=dayStart,
                dayEnd=dayEnd,
                nightStart=nightStart,
                nightEnd=nightEnd,
                actionName=actionName,
                operatorQuery=operatorQuery,
                manualOperators=manualOperators,
                hideOperatorNames=hideOperatorNames,
                pay0=pay0,
                pay1=pay1,
                pay2=pay2,
                pay3=pay3,
                thr3_1=thr3_1,
                thr3_2=thr3_2,
                thr3_3=thr3_3,
                thr4_1=thr4_1,
                thr4_2=thr4_2,
                thr4_3=thr4_3,
                thr5_1=thr5_1,
                thr5_2=thr5_2,
                thr5_3=thr5_3,
                bonusDefault=bonusDefault,
                bonusOverride=bonusOverride,
                includePresenceOnly=includePresenceOnly,
                dispatchersSource=dispatchersSource,
                minPresenceMinutes=minPresenceMinutes,
                presenceGraceMinutes=presenceGraceMinutes,
                handoverMinutes=handoverMinutes,
                reportId=report_id,
                session=bg_session,
                _current=_current,
            ),
        )
        return _as_report_out_dict(pending)

    payouts: tuple[int, int, int, int] = (int(pay0), int(pay1), int(pay2), int(pay3))
    thresholds: dict[int, tuple[int, int, int]] = {
        3: (int(thr3_1), int(thr3_2), int(thr3_3)),
        4: (int(thr4_1), int(thr4_2), int(thr4_3)),
        5: (int(thr5_1), int(thr5_2), int(thr5_3)),
    }

    def _validate_thresholds(label: str, t: tuple[int, int, int]) -> None:
        a, b, c = t
        if not (0 <= a <= b <= c <= 100):
            raise HTTPException(
                status_code=400,
                detail={"code": "BAD_REQUEST", "message": f"Invalid thresholds for {label}"},
            )

    _validate_thresholds("3 dispatchers", thresholds[3])
    _validate_thresholds("4 dispatchers", thresholds[4])
    _validate_thresholds("5 dispatchers", thresholds[5])

    # Bonus overrides parsing
    overrides: dict[str, int] = {}
    for item in bonusOverride or []:
        try:
            if not item:
                continue
            if ":" not in item:
                continue
            name, val = item.split(":", 1)
            name = name.strip()
            val_i = int(val.strip())
            if name:
                overrides[name] = val_i
        except Exception:
            continue

    manual_operator_names = {
        part.strip()
        for raw in (manualOperators or [])
        for part in re.split(r"[\r\n,;]+", str(raw or ""))
        if part.strip()
    }

    try:
        dialect_name = getattr(getattr(session.get_bind(), "dialect", None), "name", None)
    except Exception:
        dialect_name = None

    # Load only relevant actions first. Resolving EventAction -> Event in a second step
    # is much cheaper on large production datasets than a single OR-join across whole tables.
    act = (actionName or "").strip()
    action_stmt = (
        select(
            EventAction.event_id.label("event_id"),
            EventAction.raw_event_id.label("raw_event_id"),
            EventAction.date_key.label("date_key"),
            EventAction.operator_name.label("operator_name"),
            EventAction.action_time.label("ts"),
        )
        .select_from(EventAction)
        .where(EventAction.operator_name.is_not(None))
        .where(EventAction.action_time >= window_start)
    )

    if exact_window:
        action_stmt = action_stmt.where(EventAction.action_time <= window_end)
    else:
        action_stmt = action_stmt.where(EventAction.action_time < window_end)

    action_stmt = action_stmt.where(_pcn_accept_action_predicate(act))

    oq = (operatorQuery or "").strip()
    if oq:
        action_stmt = action_stmt.where(EventAction.operator_name.ilike(f"%{oq}%"))

    action_stmt = action_stmt.order_by(EventAction.action_time.asc(), EventAction.event_id.asc())
    _pcn_stage("fetch_actions_begin")
    action_rows = (await session.execute(action_stmt)).all()
    _pcn_stage("fetch_actions_done", actionRows=len(action_rows))

    exact_event_ids = sorted({str(event_id) for event_id, *_rest in action_rows if event_id})
    events_by_id: dict[str, dict[str, Any]] = {}
    for chunk in _iter_chunks(exact_event_ids):
        events_stmt = (
            select(
                Event.id,
                Event.parent_event_id,
                Event.object_id,
                Event.object_name,
                Event.location,
                Event.meter_count,
                Event.result_text,
            )
            .where(Event.id.in_(chunk))
            .where(Event.type == "alarm")
            .where(~_pcn_excluded_alarm_predicate())
        )
        for event_id, parent_event_id, object_id, object_name, location, meter_count, result_text in (
            await session.execute(events_stmt)
        ).all():
            events_by_id[str(event_id)] = {
                "eventId": str(event_id),
                "parentEventId": str(parent_event_id or "").strip() or None,
                "objectId": str(object_id or "").strip(),
                "objectName": str(object_name or "").strip(),
                "address": str(location or "").strip(),
                "meterCount": str(meter_count or "").strip(),
                "resultText": str(result_text or "").strip(),
            }
    _pcn_stage(
        "resolve_exact_events_done",
        exactEventIds=len(exact_event_ids),
        resolvedExactEvents=len(events_by_id),
    )

    unresolved_pairs = {
        (int(raw_event_id), int(date_key))
        for event_id, raw_event_id, date_key, _operator_name, _ts in action_rows
        if raw_event_id is not None and date_key is not None and str(event_id or "") not in events_by_id
    }
    events_by_raw_key: dict[tuple[int, int], dict[str, Any]] = {}
    if unresolved_pairs:
        _pcn_stage("resolve_raw_events_begin", unresolvedPairs=len(unresolved_pairs))
        raw_ids = sorted({raw_event_id for raw_event_id, _date_key in unresolved_pairs})
        raw_date_keys = sorted({date_key for _raw_event_id, date_key in unresolved_pairs})
        numeric_event_id_expr = sql_cast(Event.id, Integer)
        date_key_expr = _event_date_key_expr(dialect_name)

        for date_key in raw_date_keys:
            for raw_chunk in _iter_chunks(raw_ids):
                raw_stmt = (
                    select(
                        Event.id,
                        Event.parent_event_id,
                        Event.object_id,
                        Event.object_name,
                        Event.location,
                        Event.meter_count,
                        Event.result_text,
                        numeric_event_id_expr.label("raw_event_id"),
                        date_key_expr.label("date_key"),
                    )
                    .where(Event.type == "alarm")
                    .where(~_pcn_excluded_alarm_predicate())
                    .where(_numeric_event_id_predicate(dialect_name))
                    .where(date_key_expr == date_key)
                    .where(numeric_event_id_expr.in_(list(raw_chunk)))
                )
                for (
                    event_id,
                    parent_event_id,
                    object_id,
                    object_name,
                    location,
                    meter_count,
                    result_text,
                    raw_event_id,
                    resolved_date_key,
                ) in (await session.execute(raw_stmt)).all():
                    events_by_raw_key[(int(raw_event_id), int(resolved_date_key))] = {
                        "eventId": str(event_id),
                        "parentEventId": str(parent_event_id or "").strip() or None,
                        "objectId": str(object_id or "").strip(),
                        "objectName": str(object_name or "").strip(),
                        "address": str(location or "").strip(),
                        "meterCount": str(meter_count or "").strip(),
                        "resultText": str(result_text or "").strip(),
                    }
        _pcn_stage(
            "resolve_raw_events_done",
            rawDateKeys=len(raw_date_keys),
            rawIds=len(raw_ids),
            resolvedRawEvents=len(events_by_raw_key),
        )
    else:
        _pcn_stage("resolve_raw_events_skipped", unresolvedPairs=0)

    object_ids = sorted(
        {
            event_data["objectId"]
            for event_data in list(events_by_id.values()) + list(events_by_raw_key.values())
            if event_data.get("objectId")
        }
    )
    objects_by_id: dict[str, tuple[str, str]] = {}
    for chunk in _iter_chunks(object_ids):
        obj_stmt = select(Object.id, Object.name, Object.address).where(Object.id.in_(chunk))
        for object_id, object_name, address in (await session.execute(obj_stmt)).all():
            objects_by_id[str(object_id)] = (str(object_name or "").strip(), str(address or "").strip())
    _pcn_stage("fetch_objects_done", objectIds=len(object_ids), resolvedObjects=len(objects_by_id))

    rows: list[tuple[str, str, str, datetime, str, str, str, str, str]] = []
    for event_id, raw_event_id, date_key, op, ts in action_rows:
        event_data = events_by_id.get(str(event_id or "").strip())
        if event_data is None and raw_event_id is not None and date_key is not None:
            event_data = events_by_raw_key.get((int(raw_event_id), int(date_key)))
        if event_data is None:
            continue

        object_id = str(event_data.get("objectId") or "").strip()
        object_name = str(event_data.get("objectName") or "").strip()
        address = str(event_data.get("address") or "").strip()
        if object_id in objects_by_id:
            obj_name, obj_address = objects_by_id[object_id]
            object_name = obj_name or object_name
            address = obj_address or address

        resolved_event_id = str(event_data.get("eventId") or "").strip()
        alarm_id = _public_alarm_id(event_data.get("parentEventId"), raw_event_id, resolved_event_id)
        rows.append(
            (
            str(alarm_id or ""),
                resolved_event_id,
                str(op or "").strip(),
                ts,
                object_id,
                object_name,
                address,
                str(event_data.get("meterCount") or "").strip(),
                str(event_data.get("resultText") or "").strip(),
            )
        )
    _pcn_stage("build_source_rows_done", sourceRows=len(rows))

    # Presence (who was logged in / "in the system")
    # Used to compute the dispatcher count per shift (staffing), independent from actions.
    presence_ops_by_shift: dict[tuple[date_type, str], set[str]] = {}
    presence_seconds_by_shift_op: dict[tuple[date_type, str, str], int] = {}
    try:
        from app.models.user_presence_session import UserPresenceSession

        grace = timedelta(minutes=int(presenceGraceMinutes))
        min_seconds = int(minPresenceMinutes) * 60

        # Query a slightly wider window to account for grace.
        q_start = window_start - grace
        q_end = window_end + grace

        pres_stmt = (
            select(
                UserPresenceSession.username,
                UserPresenceSession.started_at,
                UserPresenceSession.last_seen_at,
                UserPresenceSession.ended_at,
            )
            .where(UserPresenceSession.started_at < q_end)
            .where(
                or_(
                    and_(UserPresenceSession.ended_at.is_not(None), UserPresenceSession.ended_at >= q_start),
                    and_(UserPresenceSession.ended_at.is_(None), UserPresenceSession.last_seen_at >= q_start),
                )
            )
        )

        _pcn_stage("fetch_presence_begin", qStart=q_start, qEnd=q_end)
        pres_rows = (await session.execute(pres_stmt)).all()
        _pcn_stage("fetch_presence_done", presenceRows=len(pres_rows))

        # Prepare shift windows for the date range.
        shift_windows: list[tuple[date_type, str, datetime, datetime]] = []
        d = presence_span_start
        while d <= presence_span_end:
            day_s = datetime.combine(d, day_start)
            day_e = datetime.combine(d, day_end)
            night_s = datetime.combine(d, night_start)
            night_e = datetime.combine(d + timedelta(days=1), night_end)
            shift_windows.append((d, "день", day_s, day_e))
            shift_windows.append((d, "ночь", night_s, night_e))
            d = d + timedelta(days=1)

        # Accumulate overlap seconds per (shift_date, shift_name, operator)
        presence_seconds: dict[tuple[date_type, str, str], int] = {}
        for username, started_at, last_seen_at, ended_at in pres_rows:
            op = str(username or "").strip()
            if not op or not isinstance(started_at, datetime) or not isinstance(last_seen_at, datetime):
                continue

            ses_start = started_at
            ses_end = ended_at if isinstance(ended_at, datetime) else (last_seen_at + grace)
            if ses_end <= ses_start:
                continue

            for sd, sh, s_start, s_end in shift_windows:
                # fast reject
                if ses_end <= s_start or ses_start >= s_end:
                    continue
                overlap = min(ses_end, s_end) - max(ses_start, s_start)
                sec = int(overlap.total_seconds())
                if sec <= 0:
                    continue
                key = (sd, sh, op)
                presence_seconds[key] = presence_seconds.get(key, 0) + sec

            presence_seconds_by_shift_op = dict(presence_seconds)

        for (sd, sh, op), sec in presence_seconds.items():
            if min_seconds <= 0 or sec >= min_seconds:
                presence_ops_by_shift.setdefault((sd, sh), set()).add(op)
        _pcn_stage(
            "presence_processed",
            presenceShiftOps=sum(len(ops) for ops in presence_ops_by_shift.values()),
            presenceShiftKeys=len(presence_ops_by_shift),
        )
    except Exception:
        report_worker_logger.exception(
            "PCN presence fallback report_id=%s stage=presence_failed",
            report_log_id,
        )
        # Presence is optional; fallback to action-based dispatcher count.
        presence_ops_by_shift = {}
        presence_seconds_by_shift_op = {}

    # Aggregate unique alarms per shift and operator.
    operator_alarm_ids: dict[tuple[date_type, str, str], set[str]] = {}
    shift_alarm_ids: dict[tuple[date_type, str], set[str]] = {}
    shift_alarm_details: dict[tuple[date_type, str, str], dict[str, Any]] = {}
    detail_event_ids: set[str] = set()

    handover_td = timedelta(minutes=int(handoverMinutes or 0))

    def _presence_seconds(sd: date_type, sh: str, op: str) -> int:
        try:
            return int(presence_seconds_by_shift_op.get((sd, sh, op), 0) or 0)
        except Exception:
            return 0

    def _handover_boundary_kind(ts: datetime) -> tuple[str, date_type] | None:
        if not handover_td or handover_td.total_seconds() <= 0:
            return None
        d = ts.date()
        day_start_dt = datetime.combine(d, day_start)
        night_start_dt = datetime.combine(d, night_start)
        if day_start_dt - handover_td <= ts < day_start_dt + handover_td:
            return ("morning", d)
        if night_start_dt - handover_td <= ts < night_start_dt + handover_td:
            return ("evening", d)
        return None

    # Two-pass handover attribution:
    # - Baseline: count alarms outside the handover windows.
    # - Boundary actions (near 09:00/20:00): assign to the operator's dominant shift by baseline alarm counts.
    #   Presence seconds are used as a tie-breaker if available.
    enriched_rows: list[
        tuple[
            str,
            str,
            str,
            datetime,
            str,
            str,
            str,
            str,
            str,
            date_type,
            str,
            tuple[str, date_type] | None,
        ]
    ] = []

    exact_selected_shift: tuple[date_type, str] | None = None
    if exact_window and selected_shift_keys and len(selected_shift_keys) == 1:
        exact_selected_shift = next(iter(selected_shift_keys))

    for alarm_id, event_id, op, ts, object_id, object_name, address, meter_count, result_text in rows:
        if not isinstance(ts, datetime) or not op:
            continue
        alarm_id_s = str(alarm_id or event_id or "").strip()
        if not alarm_id_s:
            continue
        op_s = str(op)
        if exact_selected_shift is not None and window_start <= ts <= window_end:
            shift_date, shift_name = exact_selected_shift
            boundary = None
        else:
            shift_date, shift_name = _shift_bucket(
                ts,
                day_start=day_start,
                day_end=day_end,
                night_start=night_start,
                night_end=night_end,
            )
            boundary = None if exact_window else _handover_boundary_kind(ts)
        enriched_rows.append(
            (
                alarm_id_s,
                str(event_id or "").strip(),
                op_s,
                ts,
                str(object_id or "").strip(),
                str(object_name or "").strip(),
                str(address or "").strip(),
                str(meter_count or "").strip(),
                str(result_text or "").strip(),
                shift_date,
                shift_name,
                boundary,
            )
        )

    baseline_alarm_ids_by_shift_op: dict[tuple[date_type, str, str], set[str]] = {}
    for alarm_id_s, _event_id_s, op_s, ts, *_rest, shift_date, shift_name, boundary in enriched_rows:
        if boundary is not None:
            continue
        baseline_alarm_ids_by_shift_op.setdefault((shift_date, shift_name, op_s), set()).add(alarm_id_s)

    def _baseline_cnt(sd: date_type, sh: str, op: str) -> int:
        return len(baseline_alarm_ids_by_shift_op.get((sd, sh, op), set()))

    for (
        alarm_id_s,
        event_id_s,
        op_s,
        ts,
        object_id_s,
        object_name_s,
        address_s,
        meter_count_s,
        result_text_s,
        shift_date,
        shift_name,
        boundary,
    ) in enriched_rows:
        if boundary is not None:
            kind, d = boundary
            if kind == "morning":
                cand_a = (d - timedelta(days=1), "ночь")
                cand_b = (d, "день")
            else:
                cand_a = (d, "день")
                cand_b = (d, "ночь")

            a_cnt = _baseline_cnt(cand_a[0], cand_a[1], op_s)
            b_cnt = _baseline_cnt(cand_b[0], cand_b[1], op_s)
            if a_cnt != b_cnt:
                shift_date, shift_name = (cand_a if a_cnt > b_cnt else cand_b)
            else:
                a_pres = _presence_seconds(cand_a[0], cand_a[1], op_s)
                b_pres = _presence_seconds(cand_b[0], cand_b[1], op_s)
                if a_pres != b_pres:
                    shift_date, shift_name = (cand_a if a_pres > b_pres else cand_b)
                else:
                    # If we have no evidence at all (no baseline alarms and no presence),
                    # avoid dragging early actions of the next DAY shift into the previous NIGHT.
                    if kind == "morning" and shift_name == "ночь" and shift_date == (d - timedelta(days=1)):
                        shift_date, shift_name = (d, "день")
                    # else: keep the original shift bucket

        if clamp_shift_dates is not None:
            dmin, dmax = clamp_shift_dates
            if shift_date < dmin or shift_date > dmax:
                continue
        operator_alarm_ids.setdefault((shift_date, shift_name, op_s), set()).add(alarm_id_s)
        shift_alarm_ids.setdefault((shift_date, shift_name), set()).add(alarm_id_s)
        detail_key = (shift_date, shift_name, alarm_id_s)
        detail = shift_alarm_details.get(detail_key)
        if detail is None:
            detail = {
                "alarmId": alarm_id_s,
                "eventId": event_id_s or None,
                "acceptedAt": ts,
                "objectId": object_id_s,
                "objectName": object_name_s,
                "address": address_s,
                "meterCount": meter_count_s,
                "resultText": result_text_s,
                "operators": {op_s},
            }
            shift_alarm_details[detail_key] = detail
        else:
            detail["operators"].add(op_s)
            accepted_at = detail.get("acceptedAt")
            if isinstance(accepted_at, datetime) and ts < accepted_at:
                detail["acceptedAt"] = ts
            if not detail.get("objectId") and object_id_s:
                detail["objectId"] = object_id_s
            if not detail.get("objectName") and object_name_s:
                detail["objectName"] = object_name_s
            if not detail.get("address") and address_s:
                detail["address"] = address_s
            if not detail.get("meterCount") and meter_count_s:
                detail["meterCount"] = meter_count_s
            if not detail.get("resultText") and result_text_s:
                detail["resultText"] = result_text_s

        if event_id_s:
            detail_event_ids.add(event_id_s)
    _pcn_stage(
        "aggregate_done",
        operatorAlarmGroups=len(operator_alarm_ids),
        shiftTotals=len(shift_alarm_ids),
        detailEventIds=len(detail_event_ids),
    )

    # Payroll rule (date-based reports only): if the same operator appears in both day and night shifts
    # for the same shift_date, attribute them to the dominant shift (by alarm count)
    # and move the smaller-shift alarms to the dominant shift.
    #
    # IMPORTANT: do NOT apply this rule for exact windows ("report for a single shift")
    # otherwise the selected shift can become empty after re-attribution.
    def _move_operator_alarm(
        *,
        op: str,
        alarm_id: str,
        from_sd: date_type,
        from_sh: str,
        to_sd: date_type,
        to_sh: str,
    ) -> None:
        # Update shift totals (best-effort). If other operators are still present on the
        # alarm in the source shift, we keep it there too.
        from_key = (from_sd, from_sh, alarm_id)
        src_detail = shift_alarm_details.get(from_key)
        snapshot: dict[str, Any] | None = None
        if isinstance(src_detail, dict) and op in (src_detail.get("operators") or set()):
            snapshot = {
                **src_detail,
                "operators": set(src_detail.get("operators") or set()),
            }
            src_detail["operators"].discard(op)
            if not src_detail["operators"]:
                shift_alarm_details.pop(from_key, None)
                shift_alarm_ids.get((from_sd, from_sh), set()).discard(alarm_id)

        # Ensure destination detail has this operator.
        to_key = (to_sd, to_sh, alarm_id)
        dst_detail = shift_alarm_details.get(to_key)
        if dst_detail is None:
            base = snapshot or {
                "alarmId": alarm_id,
                "eventId": None,
                "acceptedAt": None,
                "objectId": "",
                "objectName": "",
                "address": "",
                "meterCount": "",
                "resultText": "",
                "operators": set(),
            }
            shift_alarm_details[to_key] = {
                "alarmId": str(base.get("alarmId") or alarm_id),
                "eventId": base.get("eventId"),
                "acceptedAt": base.get("acceptedAt"),
                "objectId": str(base.get("objectId") or ""),
                "objectName": str(base.get("objectName") or ""),
                "address": str(base.get("address") or ""),
                "meterCount": str(base.get("meterCount") or ""),
                "resultText": str(base.get("resultText") or ""),
                "operators": {op},
            }
        else:
            dst_detail.setdefault("operators", set())
            if not isinstance(dst_detail["operators"], set):
                dst_detail["operators"] = set(dst_detail["operators"] or [])
            dst_detail["operators"].add(op)
            if snapshot is not None:
                a = snapshot.get("acceptedAt")
                b = dst_detail.get("acceptedAt")
                if isinstance(a, datetime) and (not isinstance(b, datetime) or a < b):
                    dst_detail["acceptedAt"] = a

        shift_alarm_ids.setdefault((to_sd, to_sh), set()).add(alarm_id)

    if not exact_window:
        # Merge for each date/operator.
        # Note: shift_date for "ночь" is the night start date (20:00 of that date).
        ops_by_date: dict[date_type, set[str]] = {}
        for (sd, sh, op), _alarms in operator_alarm_ids.items():
            if sh in {"день", "ночь"}:
                ops_by_date.setdefault(sd, set()).add(op)

        for sd, ops in ops_by_date.items():
            for op in ops:
                day_key = (sd, "день", op)
                night_key = (sd, "ночь", op)
                day_alarms = operator_alarm_ids.get(day_key) or set()
                night_alarms = operator_alarm_ids.get(night_key) or set()
                if not day_alarms or not night_alarms:
                    continue

                day_cnt = len(day_alarms)
                night_cnt = len(night_alarms)
                if day_cnt == night_cnt:
                    day_pres = _presence_seconds(sd, "день", op)
                    night_pres = _presence_seconds(sd, "ночь", op)
                    if day_pres == night_pres:
                        continue
                    dominant = "день" if day_pres > night_pres else "ночь"
                else:
                    dominant = "день" if day_cnt > night_cnt else "ночь"

                if dominant == "ночь":
                    from_key = day_key
                    to_key = night_key
                else:
                    from_key = night_key
                    to_key = day_key

                moved = set(operator_alarm_ids.get(from_key) or set())
                if not moved:
                    continue
                operator_alarm_ids.setdefault(to_key, set()).update(moved)
                operator_alarm_ids.pop(from_key, None)

                from_sd, from_sh, _ = from_key
                to_sd, to_sh, _ = to_key
                for alarm_id in moved:
                    _move_operator_alarm(
                        op=op,
                        alarm_id=alarm_id,
                        from_sd=from_sd,
                        from_sh=from_sh,
                        to_sd=to_sd,
                        to_sh=to_sh,
                    )

    counts: dict[tuple[date_type, str, str], int] = {
        key: len(alarm_ids)
        for key, alarm_ids in operator_alarm_ids.items()
    }

    # Totals per shift
    shift_totals: dict[tuple[date_type, str], int] = {
        key: len(alarm_ids)
        for key, alarm_ids in shift_alarm_ids.items()
    }
    shift_ops: dict[tuple[date_type, str], set[str]] = {}
    for (sd, sh, op), c in counts.items():
        shift_ops.setdefault((sd, sh), set()).add(op)

    alarm_trip_meta_by_event_id: dict[str, dict[str, str]] = {}
    if detail_event_ids:
        called_match = or_(
            EventAction.action_name.ilike("%Вызван%"),
            EventAction.action_name.ilike("%Вызов%"),
            EventAction.action_name.ilike("%Направ%"),
            EventAction.action_name.ilike("%Отправ%"),
            EventAction.action_name.ilike("%Выезд%"),
            EventAction.action_name.ilike("%Следу%"),
        )
        trip_stmt = (
            select(
                EventAction.event_id,
                func.coalesce(
                    func.min(case((called_match, EventAction.gbr_name), else_=None)),
                    func.min(EventAction.gbr_name),
                ).label("gbr_name"),
            )
            .where(EventAction.event_id.in_(sorted(detail_event_ids)))
            .group_by(EventAction.event_id)
        )
        for event_id, gbr_name in (await session.execute(trip_stmt)).all():
            alarm_trip_meta_by_event_id[str(event_id)] = {
                "gbrName": str(gbr_name or "").strip(),
            }
    _pcn_stage("fetch_trip_meta_done", tripMetaEvents=len(alarm_trip_meta_by_event_id))

    # Build ordered output rows
    ordered_shift_keys = set(shift_totals.keys())
    if selected_shift_keys is not None:
        ordered_shift_keys.update(selected_shift_keys)

    ordered_shifts = sorted(ordered_shift_keys, key=lambda x: (x[0].toordinal(), 0 if x[1] == "день" else 1))
    if selected_shift_keys is not None:
        ordered_shifts = [key for key in ordered_shifts if key in selected_shift_keys]

    operator_label = (operatorQuery or "").strip()
    operator_filter_ops: set[str] | None = None
    if operator_label:
        # Match operators by substring (case-insensitive), consistent with UI's "ФИО или часть".
        # Note: we filter OUTPUT rows only; dispatcher count and shift totals are computed for the whole shift.
        wanted = operator_label.lower()
        candidate_ops = {str(op) for (_sd, _sh, op) in counts.keys() if str(op).strip()}
        operator_filter_ops = {op for op in candidate_ops if wanted in op.lower()}

    out_rows: list[dict[str, Any]] = []
    control_rows: list[dict[str, Any]] = []
    detail_rows: list[dict[str, Any]] = []
    for sd, sh in ordered_shifts:
        total = int(shift_totals.get((sd, sh)) or 0)
        ops_from_actions = shift_ops.get((sd, sh)) or set()
        ops_from_presence = presence_ops_by_shift.get((sd, sh)) or set()
        manual_matched_ops, manual_unmatched_ops = _resolve_manual_operator_names(
            manual_operator_names,
            set(map(str, ops_from_actions)) | set(map(str, ops_from_presence)),
        )

        # Operators for this shift by accepted alarm actions.
        alarms_by_op: dict[str, int] = {}
        for (sd2, sh2, op), c in counts.items():
            if sd2 == sd and sh2 == sh:
                alarms_by_op[str(op)] = int(c or 0)

        ranked_action_ops = [
            op
            for op, count_value in sorted(alarms_by_op.items(), key=lambda x: (-x[1], x[0].lower()))
            if int(count_value or 0) > 0
        ]
        effective_action_ops = ranked_action_ops[:5]

        ranked_presence_ops = sorted(map(str, ops_from_presence), key=str.lower)
        effective_presence_ops = ranked_presence_ops[:5]

        dispatchers_presence = len(effective_presence_ops)
        dispatchers_actions = len(effective_action_ops)

        # Dispatcher count source selection.
        # For PЦН business logic we primarily care about who actually handled alarms
        # in the shift. Presence remains auxiliary/debug information on the control sheet.
        if manual_operator_names:
            used_ops = set(manual_matched_ops)
            dispatchers = len(manual_operator_names)
        elif ds == "presence":
            used_ops = set(effective_presence_ops)
            dispatchers = dispatchers_presence
        elif ds == "actions":
            used_ops = set(effective_action_ops)
            dispatchers = dispatchers_actions
        else:
            # auto: prefer archive actions, fallback to presence only if no actions exist.
            used_ops = set(effective_action_ops or effective_presence_ops)
            dispatchers = len(used_ops)

        total_used = sum(int(alarms_by_op.get(op, 0) or 0) for op in used_ops)
        total_for_sheet = int(total if manual_operator_names else (total_used or total))

        if sh == "день":
            shift_window = f"{day_start.strftime('%H:%M')}–{day_end.strftime('%H:%M')}"
        else:
            shift_window = f"{night_start.strftime('%H:%M')}–{night_end.strftime('%H:%M')}"

        control_rows.append(
            {
                "date": sd,
                "shift": sh,
                "shiftWindow": shift_window,
                "totalAlarms": total,
                "totalAlarmsUsed": total_for_sheet,
                "dispatchersPresence": dispatchers_presence,
                "operatorsPresence": ", ".join(sorted(map(str, ops_from_presence), key=str.lower)),
                "dispatchersActions": dispatchers_actions,
                "operatorsActions": ", ".join(ranked_action_ops),
                "dispatchersUsed": dispatchers,
                "operatorsUsed": ", ".join(
                    sorted(
                        list(map(str, used_ops)) + list(map(str, manual_unmatched_ops)),
                        key=str.lower,
                    )
                ),
            }
        )

        if manual_operator_names:
            op_names = set(manual_matched_ops) | set(manual_unmatched_ops)
        elif ds == "presence":
            if includePresenceOnly:
                op_names = set(effective_presence_ops)
            else:
                op_names = set(op for op in effective_presence_ops if int(alarms_by_op.get(op, 0)) > 0)
        else:
            op_names = set(effective_action_ops)

        if operator_filter_ops is not None:
            op_names = {op for op in op_names if op in operator_filter_ops}

        ops = [(op, int(alarms_by_op.get(op, 0))) for op in op_names]
        ops.sort(key=lambda x: (-x[1], x[0].lower()))

        for op, c in ops:
            percent = (float(c) * 100.0 / float(total_for_sheet)) if total_for_sheet > 0 else 0.0
            payout = _pcn_ledger_payout(dispatchers, percent, payouts=payouts, thresholds=thresholds)
            bonus = int(overrides.get(op, bonusDefault))
            out_rows.append(
                {
                    "date": sd,
                    "dispatchers": dispatchers,
                    "shift": sh,
                    "operator": op,
                    "alarms": c,
                    "percent": percent,
                    "payout": payout,
                    "bonus": bonus,
                    "total": payout + bonus,
                }
            )

        manual_filter_ops = set(manual_matched_ops) if manual_operator_names else set()
        shift_details = [
            detail
            for (detail_sd, detail_sh, _alarm_id), detail in shift_alarm_details.items()
            if detail_sd == sd and detail_sh == sh
        ]
        shift_details.sort(
            key=lambda item: (
                (lambda accepted_at: accepted_at.isoformat() if isinstance(accepted_at, datetime) else "")(
                    item.get("acceptedAt")
                ),
                str(item.get("objectId") or ""),
            )
        )

        for detail in shift_details:
            detail_ops_raw = detail.get("operators")
            detail_ops = set(detail_ops_raw) if isinstance(detail_ops_raw, set) else set()
            if manual_filter_ops and detail_ops.isdisjoint(manual_filter_ops):
                continue
            if operator_filter_ops is not None and detail_ops.isdisjoint(operator_filter_ops):
                continue
            event_id = str(detail.get("eventId") or "").strip()
            trip_meta = alarm_trip_meta_by_event_id.get(event_id, {})
            detail_rows.append(
                {
                    "date": sd,
                    "shift": sh,
                    "acceptedAt": detail.get("acceptedAt"),
                    "objectId": detail.get("objectId") or "",
                    "objectName": detail.get("objectName") or "",
                    "address": detail.get("address") or "",
                    "meterCount": detail.get("meterCount") or "",
                    "gbrName": trip_meta.get("gbrName") or "",
                    "resultText": detail.get("resultText") or "",
                    "operators": ", ".join(sorted(detail_ops, key=str.lower)),
                    "alarmId": detail.get("alarmId") or "",
                }
            )
    _pcn_stage(
        "output_rows_done",
        outRows=len(out_rows),
        controlRows=len(control_rows),
        detailRows=len(detail_rows),
        orderedShifts=len(ordered_shifts),
    )

    # Build XLSX
    from io import BytesIO

    try:
        from openpyxl import Workbook
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        from openpyxl.styles import Alignment, Border, Font, Side
    except ModuleNotFoundError as e:
        raise HTTPException(
            status_code=500,
            detail={
                "code": "MISSING_DEP",
                "message": "Не установлен openpyxl (нужен для XLSX). Установите зависимости: pip install -r backend/requirements.txt",
            },
        ) from e

    def clean_excel_text(value: object) -> Any:
        if value is None:
            return ""
        if isinstance(value, str):
            return re.sub(ILLEGAL_CHARACTERS_RE, "", value)
        return value

    wb = Workbook()
    ws = cast(Any, wb.active)
    ws.title = "Ведомость"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Threshold legend (like sample)
    ws["C3"].value = f"{payouts[0]} р."
    ws["D3"].value = f"{payouts[1]} р."
    ws["E3"].value = f"{payouts[2]} р."
    ws["F3"].value = f"{payouts[3]} р."

    t31, t32, t33 = thresholds[3]
    t41, t42, t43 = thresholds[4]
    t51, t52, t53 = thresholds[5]

    ws["B4"].value = "3 диспетчера"
    ws["C4"].value = f"< {t31} %"
    ws["D4"].value = f"{t31}– {t32} %"
    ws["E4"].value = f"{t32} – {t33} %"
    ws["F4"].value = f">{t33} %"

    ws["B5"].value = "4 диспетчера"
    ws["C5"].value = f"< {t41} %"
    ws["D5"].value = f"{t41} – {t42} %"
    ws["E5"].value = f"{t42} – {t43} %"
    ws["F5"].value = f">{t43} %"

    ws["B6"].value = "5 диспетчеров"
    ws["C6"].value = f"< {t51} %"
    ws["D6"].value = f"{t51} – {t52} %"
    ws["E6"].value = f"{t52} – {t53} %"
    ws["F6"].value = f">{t53} %"

    for r in range(3, 7):
        for c in range(2, 7):
            cell = ws.cell(r, c)
            cell.alignment = center
            cell.border = border
            cell.font = Font(size=10, bold=(r == 3))

    # Formula note (so the sheet always shows the exact rule used)
    ws.merge_cells(start_row=7, start_column=2, end_row=8, end_column=10)
    formula_note = "".join(
        [
            "Формула: % = (тревоги оператора * 100) / (все тревоги смены). ",
            "Выплата определяется по порогам выше (по числу диспетчеров в смену).\n",
            f"Рабочие окна смен: день {day_start.strftime('%H:%M')}–{day_end.strftime('%H:%M')}, ночь {night_start.strftime('%H:%M')}–{night_end.strftime('%H:%M')}. ",
            f"Отработка тревоги: действие '{actionName or ''}'. ",
            f"Диспетчеры в смену: {ds} (auto/actions = до 5 операторов по архивным действиям; presence>= {int(minPresenceMinutes)} мин, grace {int(presenceGraceMinutes)} мин). ",
            f"Исключены нетиповые результаты: {', '.join(_PCN_EXCLUDED_ALARM_PATTERNS)}. ",
            "ФИО скрыты. " if hideOperatorNames else "",
            "Сравнение presence/actions — на листе 'Контроль'.",
        ]
    )
    ws.cell(7, 2, clean_excel_text(formula_note)).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(7, 2).font = Font(size=10)

    # Title
    ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=10)
    title = _pcn_ledger_title(period_start_date, period_end_date, operatorQuery)
    ws.cell(9, 2, clean_excel_text(title)).font = Font(bold=True, size=12)
    ws.cell(9, 2).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=10, start_column=2, end_row=10, end_column=10)
    scope_bits = [f"Оператор: {operator_label}" if operator_label else "Оператор: все"]
    if hideOperatorNames:
        scope_bits.append("ФИО скрыты")
    ws.cell(10, 2, clean_excel_text(" | ".join(scope_bits))).alignment = Alignment(horizontal="center")
    ws.cell(10, 2).font = Font(size=10, italic=True)

    ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=10)
    ws.cell(
        11,
        2,
        clean_excel_text(
            f"Дневная смена: {day_start.strftime('%H:%M')}–{day_end.strftime('%H:%M')} | "
            f"Ночная смена: {night_start.strftime('%H:%M')}–{night_end.strftime('%H:%M')}"
        ),
    ).alignment = Alignment(horizontal="center")
    ws.cell(11, 2).font = Font(size=10)

    headers = [
        "Дата",
        "Количество диспетчеров в смену",
        "Смена",
        "ФИО",
        "Тревоги",
        "%",
        "Сумма",
        "Доплата",
        "Всего в смену, руб.",
    ]

    start_row = 13
    for idx, h in enumerate(headers, start=2):
        cell = ws.cell(start_row, idx, clean_excel_text(h))
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border

    cur = start_row + 1
    # Group by shift
    from collections import defaultdict

    grouped: dict[tuple[date_type, str], list[dict[str, Any]]] = defaultdict(list)
    for r in out_rows:
        grouped[(r["date"], r["shift"])].append(r)

    for sd, sh in ordered_shifts:
        items = grouped.get((sd, sh)) or []
        if not items:
            continue

        # Operator rows
        total_alarms = sum(int(x.get("alarms") or 0) for x in items)
        total_payout = sum(int(x.get("payout") or 0) for x in items)
        total_bonus = sum(int(x.get("bonus") or 0) for x in items)
        total_total = sum(int(x.get("total") or 0) for x in items)

        for i, x in enumerate(items):
            ws.cell(cur, 2, sd).number_format = "DD.MM.YYYY" if i == 0 else ""
            ws.cell(cur, 3, int(x.get("dispatchers") or 0) if i == 0 else "")
            ws.cell(cur, 4, sh if i == 0 else "")
            ws.cell(cur, 5, "" if hideOperatorNames else clean_excel_text(x.get("operator")))
            ws.cell(cur, 6, int(x.get("alarms") or 0))
            ws.cell(cur, 7, float(x.get("percent") or 0.0))
            ws.cell(cur, 7).number_format = "0.00"
            ws.cell(cur, 8, int(x.get("payout") or 0))
            ws.cell(cur, 9, int(x.get("bonus") or 0))
            ws.cell(cur, 10, int(x.get("total") or 0))

            for col in range(2, 11):
                cell = ws.cell(cur, col)
                cell.border = border
                cell.alignment = center if col != 5 else Alignment(horizontal="left", vertical="center")
            cur += 1

        # Summary row
        ws.cell(cur, 3, "Всего в смену:").font = Font(bold=True)
        ws.cell(cur, 6, int(total_alarms)).font = Font(bold=True)
        ws.cell(cur, 8, int(total_payout)).font = Font(bold=True)
        ws.cell(cur, 9, int(total_bonus)).font = Font(bold=True)
        ws.cell(cur, 10, int(total_total)).font = Font(bold=True)
        for col in range(2, 11):
            ws.cell(cur, col).border = border
            ws.cell(cur, col).alignment = center if col != 5 else Alignment(horizontal="left", vertical="center")
        cur += 1

    # Column widths (approx)
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 16

    # Control sheet: compare dispatcher counts from presence vs from actions
    try:
        ws2 = wb.create_sheet("Контроль")
        ws2_headers = [
            "Дата",
            "Смена",
            "Границы смены",
            "Всего тревог (все)",
            "Всего тревог (в расчете)",
            "Диспетчеры (presence)",
            "Операторы (presence)",
            "Диспетчеры (actions)",
            "Операторы (actions)",
            "Использовано",
            "Операторы (использовано)",
        ]
        for idx, h in enumerate(ws2_headers, start=1):
            cell = ws2.cell(1, idx, clean_excel_text(h))
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border

        r = 2
        for x in control_rows:
            sd = x.get("date")
            ws2.cell(r, 1, sd).number_format = "DD.MM.YYYY"
            ws2.cell(r, 2, clean_excel_text(x.get("shift")))
            ws2.cell(r, 3, clean_excel_text(x.get("shiftWindow") or ""))
            ws2.cell(r, 4, int(x.get("totalAlarms") or 0))
            ws2.cell(r, 5, int(x.get("totalAlarmsUsed") or 0))
            ws2.cell(r, 6, int(x.get("dispatchersPresence") or 0))
            ws2.cell(r, 7, clean_excel_text(x.get("operatorsPresence") or ""))
            ws2.cell(r, 8, int(x.get("dispatchersActions") or 0))
            ws2.cell(r, 9, clean_excel_text(x.get("operatorsActions") or ""))
            ws2.cell(r, 10, int(x.get("dispatchersUsed") or 0))
            ws2.cell(r, 11, clean_excel_text(x.get("operatorsUsed") or ""))
            for col in range(1, 12):
                c = ws2.cell(r, col)
                c.border = border
                c.alignment = center if col not in {7, 9, 11} else Alignment(horizontal="left", vertical="center", wrap_text=True)
            r += 1

        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 10
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 12
        ws2.column_dimensions["E"].width = 18
        ws2.column_dimensions["F"].width = 20
        ws2.column_dimensions["G"].width = 42
        ws2.column_dimensions["H"].width = 18
        ws2.column_dimensions["I"].width = 42
        ws2.column_dimensions["J"].width = 14
        ws2.column_dimensions["K"].width = 42
    except Exception:
        # If something goes wrong, do not fail the report generation.
        pass

    try:
        ws3 = wb.create_sheet("Тревоги")
        ws3_headers = [
            "Дата",
            "Смена",
            "Принята",
            "№ объекта",
            "Название",
            "Адрес",
            "№ шлейфа",
            "ГБР",
            "Результат",
            "Оператор(ы)",
            "№ тревоги",
        ]
        for idx, h in enumerate(ws3_headers, start=1):
            cell = ws3.cell(1, idx, clean_excel_text(h))
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border

        row_idx = 2
        for item in detail_rows:
            accepted_at = item.get("acceptedAt")
            values = [
                item.get("date"),
                item.get("shift") or "",
                accepted_at.strftime("%d.%m.%Y %H:%M:%S") if isinstance(accepted_at, datetime) else "",
                item.get("objectId") or "",
                item.get("objectName") or "",
                item.get("address") or "",
                item.get("meterCount") or "",
                item.get("gbrName") or "",
                item.get("resultText") or "",
                item.get("operators") or "",
                item.get("alarmId") or "",
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws3.cell(row_idx, col_idx, clean_excel_text(value))
                cell.border = border
                cell.alignment = center if col_idx not in {5, 6, 9, 10, 11} else Alignment(horizontal="left", vertical="center", wrap_text=True)
            if isinstance(item.get("date"), date_type):
                ws3.cell(row_idx, 1).number_format = "DD.MM.YYYY"
            row_idx += 1

        ws3.column_dimensions["A"].width = 12
        ws3.column_dimensions["B"].width = 10
        ws3.column_dimensions["C"].width = 20
        ws3.column_dimensions["D"].width = 14
        ws3.column_dimensions["E"].width = 34
        ws3.column_dimensions["F"].width = 42
        ws3.column_dimensions["G"].width = 14
        ws3.column_dimensions["H"].width = 18
        ws3.column_dimensions["I"].width = 36
        ws3.column_dimensions["J"].width = 28
        ws3.column_dimensions["K"].width = 20
    except Exception:
        pass

    try:
        ws4 = wb.create_sheet("Тревоги по операторам")
        ws4_headers = [
            "Дата",
            "Смена",
            "Оператор",
            "Принята",
            "№ объекта",
            "Название",
            "Адрес",
            "№ шлейфа",
            "ГБР",
            "Результат",
            "№ тревоги",
        ]
        for idx, h in enumerate(ws4_headers, start=1):
            cell = ws4.cell(1, idx, clean_excel_text(h))
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border

        row_idx = 2
        for item in detail_rows:
            accepted_at = item.get("acceptedAt")
            operator_names = [
                name.strip()
                for name in str(item.get("operators") or "").split(",")
                if name.strip()
            ]
            if not operator_names:
                operator_names = [""]
            for operator_name in operator_names:
                values = [
                    item.get("date"),
                    item.get("shift") or "",
                    operator_name,
                    accepted_at.strftime("%d.%m.%Y %H:%M:%S") if isinstance(accepted_at, datetime) else "",
                    item.get("objectId") or "",
                    item.get("objectName") or "",
                    item.get("address") or "",
                    item.get("meterCount") or "",
                    item.get("gbrName") or "",
                    item.get("resultText") or "",
                    item.get("alarmId") or "",
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws4.cell(row_idx, col_idx, clean_excel_text(value))
                    cell.border = border
                    cell.alignment = center if col_idx not in {3, 6, 7, 10, 11} else Alignment(horizontal="left", vertical="center", wrap_text=True)
                if isinstance(item.get("date"), date_type):
                    ws4.cell(row_idx, 1).number_format = "DD.MM.YYYY"
                row_idx += 1

        ws4.column_dimensions["A"].width = 12
        ws4.column_dimensions["B"].width = 10
        ws4.column_dimensions["C"].width = 28
        ws4.column_dimensions["D"].width = 20
        ws4.column_dimensions["E"].width = 14
        ws4.column_dimensions["F"].width = 34
        ws4.column_dimensions["G"].width = 42
        ws4.column_dimensions["H"].width = 14
        ws4.column_dimensions["I"].width = 18
        ws4.column_dimensions["J"].width = 36
        ws4.column_dimensions["K"].width = 20
    except Exception:
        pass

    bio = BytesIO()
    _pcn_stage("workbook_save_begin")
    wb.save(bio)
    data = bio.getvalue()
    _pcn_stage("workbook_save_done", fileBytes=len(data))

    report_id = reportId or str(uuid4())
    filename = f"pcn-ledger-{ps}-{pe}.xlsx"

    # For the operator-filtered report, show the operator's alarm count in the UI list.
    # For the full shift report, keep the previous behavior: total alarms used in calculations.
    if operator_filter_ops is not None:
        events_count_for_ui = sum(int(x.get("alarms") or 0) for x in out_rows)
    else:
        events_count_for_ui = sum(int(x.get("totalAlarmsUsed") or 0) for x in control_rows) if control_rows else 0

    _pcn_stage("store_report_begin", eventsCountForUi=events_count_for_ui)

    return await _store_generated_report(
        session,
        report_id=report_id,
        report_type="pcnLedger",
        period_start=ps,
        period_end=pe,
        filename=filename,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=data,
        events_count=events_count_for_ui,
        critical_count=0,
        params={
            "dateFrom": dateFrom,
            "dateTo": dateTo,
            "dayStart": dayStart,
            "dayEnd": day_end.strftime('%H:%M'),
            "nightStart": nightStart,
            "nightEnd": night_end.strftime('%H:%M'),
            "actionName": actionName,
            "operatorQuery": operatorQuery,
            "manualOperators": sorted(manual_operator_names, key=str.lower),
            "hideOperatorNames": hideOperatorNames,
            "payouts": {"pay0": pay0, "pay1": pay1, "pay2": pay2, "pay3": pay3},
            "thresholds": {
                "3": [thr3_1, thr3_2, thr3_3],
                "4": [thr4_1, thr4_2, thr4_3],
                "5": [thr5_1, thr5_2, thr5_3],
            },
            "bonusDefault": bonusDefault,
            "bonusOverride": bonusOverride or [],
            "includePresenceOnly": includePresenceOnly,
            "dispatchersSource": dispatchersSource,
            "minPresenceMinutes": minPresenceMinutes,
            "presenceGraceMinutes": presenceGraceMinutes,
        },
    )


@router.get("/{report_id}/download")
async def download_report(
    report_id: str,
    format: str | None = Query(default=None, description="csv|xlsx"),
    session: AsyncSession = Depends(get_session),
    current: dict = Depends(get_current_user),
) -> Response:
    r = (
        await session.execute(select(Report).where(Report.id == report_id).limit(1))
    ).scalars().first()
    if r is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "Report not found"})

    # Permission gate for analytics-based reports
    if str(r.type) == "gbrRaportXlsx":
        have = set(map(str, current.get("permissions") or []))
        if "analytics:read" not in have and current.get("role") != "admin":
            raise HTTPException(status_code=403, detail={"code": "FORBIDDEN", "message": "Missing permissions"})

    # If pending is stale, fail it now so UI stops showing an infinite spinner.
    if _pending_is_stale(r, now_utc_naive=datetime.utcnow(), stale_seconds=_REPORT_PENDING_STALE_SECONDS):
        r.status = "failed"
        r.error_message = (
            "Отчёт завис в статусе 'Ожидает' (возможно, сервер перезапускался). "
            "Перегенерируйте отчёт из меню."
        )
        r.generated_at = _utcnow_iso()
        await session.commit()

    if str(r.status) == "pending":
        raise HTTPException(status_code=409, detail={"code": "PENDING", "message": "Отчёт ещё формируется"})
    if str(r.status) == "failed":
        raise HTTPException(
            status_code=409,
            detail={"code": "FAILED", "message": "Отчёт не сформирован", "error": (r.error_message or "")},
        )

    if not r.storage_path or not r.file_name:
        raise HTTPException(status_code=409, detail={"code": "NO_FILE", "message": "Report has no stored file"})

    resolved = _resolve_and_validate_store_path(str(r.storage_path))

    requested = (format or "").strip().lower() or None
    if requested not in {None, "xlsx"}:
        raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Invalid format"})

    stored_ext = _file_ext(r.file_name)
    # Always return XLSX to users.
    if stored_ext == "xlsx":
        return FileResponse(
            path=str(resolved),
            media_type=r.mime_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=r.file_name,
        )

    src = resolved.read_bytes()
    if stored_ext == "csv":
        out = _csv_bytes_to_xlsx_bytes(src)
        filename = (r.file_name.rsplit(".", 1)[0] + ".xlsx") if r.file_name else "report.xlsx"
        return Response(
            content=out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # fallback: return as is
    media = r.mime_type or "application/octet-stream"
    return FileResponse(
        path=str(resolved),
        media_type=media,
        filename=r.file_name,
    )


@router.delete("/{report_id}")
async def delete_report(
    report_id: str,
    session: AsyncSession = Depends(get_session),
    current: dict = Depends(get_current_user),
) -> Response:
    _ensure_reports_manage_perm(current)

    r = (
        await session.execute(select(Report).where(Report.id == report_id).limit(1))
    ).scalars().first()
    if r is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "Report not found"})

    # Permission gate for analytics-based reports
    if str(r.type) == "gbrRaportXlsx":
        have = set(map(str, current.get("permissions") or []))
        if "analytics:read" not in have and current.get("role") != "admin":
            raise HTTPException(status_code=403, detail={"code": "FORBIDDEN", "message": "Missing permissions"})

    if r.storage_path:
        try:
            resolved = _resolve_and_validate_store_path(str(r.storage_path))
            resolved.unlink(missing_ok=True)
        except HTTPException:
            # If file is missing/bad path, still allow deleting DB record.
            pass
        except Exception:
            pass

    await session.delete(r)
    await session.commit()
    return Response(status_code=204)


@router.get("/{report_id}/params")
async def get_report_params(
    report_id: str,
    session: AsyncSession = Depends(get_session),
    current: dict = Depends(get_current_user),
) -> dict:
    """Return generation/view parameters for a stored report.

    Used by the UI for 'Просмотреть параметры' and 'Перегенерировать'.
    """

    r = (
        await session.execute(select(Report).where(Report.id == report_id).limit(1))
    ).scalars().first()
    if r is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "Report not found"})

    # Permission gate for analytics-based reports
    if str(r.type) == "gbrRaportXlsx":
        have = set(map(str, current.get("permissions") or []))
        if "analytics:read" not in have and current.get("role") != "admin":
            raise HTTPException(status_code=403, detail={"code": "FORBIDDEN", "message": "Missing permissions"})

    if _pending_is_stale(r, now_utc_naive=datetime.utcnow(), stale_seconds=_REPORT_PENDING_STALE_SECONDS):
        r.status = "failed"
        r.error_message = (
            "Отчёт завис в статусе 'Ожидает' (возможно, сервер перезапускался). "
            "Перегенерируйте отчёт из меню."
        )
        r.generated_at = _utcnow_iso()
        await session.commit()

    params: dict[str, Any] = {}
    try:
        if r.params_json:
            loaded = json.loads(r.params_json)
            if isinstance(loaded, dict):
                params = loaded
    except Exception:
        params = {}

    out = _as_report_out_dict(r)
    out["params"] = params
    return out


@router.get("/{report_id}/preview")
async def preview_report(
    report_id: str,
    maxRows: int = Query(200, ge=1, le=5000, description="Max rows for table preview"),
    maxCols: int = Query(50, ge=1, le=200, description="Max cols for table preview"),
    sheetName: str | None = Query(default=None, description="Worksheet name for XLSX preview"),
    session: AsyncSession = Depends(get_session),
    current: dict = Depends(get_current_user),
) -> dict:
    r = (
        await session.execute(select(Report).where(Report.id == report_id).limit(1))
    ).scalars().first()
    if r is None:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "Report not found"})

    # Permission gate for analytics-based reports
    if str(r.type) == "gbrRaportXlsx":
        have = set(map(str, current.get("permissions") or []))
        if "analytics:read" not in have and current.get("role") != "admin":
            raise HTTPException(status_code=403, detail={"code": "FORBIDDEN", "message": "Missing permissions"})

    if _pending_is_stale(r, now_utc_naive=datetime.utcnow(), stale_seconds=_REPORT_PENDING_STALE_SECONDS):
        r.status = "failed"
        r.error_message = (
            "Отчёт завис в статусе 'Ожидает' (возможно, сервер перезапускался). "
            "Перегенерируйте отчёт из меню."
        )
        r.generated_at = _utcnow_iso()
        await session.commit()

    if str(r.status) == "pending":
        raise HTTPException(status_code=409, detail={"code": "PENDING", "message": "Отчёт ещё формируется"})
    if str(r.status) == "failed":
        raise HTTPException(
            status_code=409,
            detail={"code": "FAILED", "message": "Отчёт не сформирован", "error": (r.error_message or "")},
        )

    # Generic preview for stored files (CSV/XLSX)
    if not r.storage_path or not r.file_name:
        raise HTTPException(status_code=409, detail={"code": "NO_FILE", "message": "Report has no stored file"})

    resolved = _resolve_and_validate_store_path(str(r.storage_path))
    content = resolved.read_bytes()
    ext = _file_ext(r.file_name) or ("csv" if (r.mime_type or "").startswith("text/csv") else "")

    if ext == "csv":
        return _preview_table_from_csv_bytes(content, max_rows=int(maxRows))
    if ext == "xlsx":
        return _preview_table_from_xlsx_bytes(
            content,
            max_rows=int(maxRows),
            max_cols=int(maxCols),
            sheet_name=sheetName,
        )

    raise HTTPException(status_code=400, detail={"code": "BAD_REQUEST", "message": "Preview not supported"})


@router.get("/export/daily", include_in_schema=False)
async def export_daily(
    date: str = Query(default_factory=today_str, description="YYYY-MM-DD"),
    session: AsyncSession = Depends(get_session),
) -> Response:
    raise HTTPException(
        status_code=410,
        detail={
            "code": "DISABLED",
            "message": "Суточный отчёт отключён. Формируйте отчёты только по согласованным событиям/объектам.",
        },
    )

    content = await export_daily_report_csv(session=session, date=date)
    xlsx = _csv_bytes_to_xlsx_bytes(content)
    filename = f"daily-report-{date}.xlsx"
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/daily/xlsx", include_in_schema=False)
async def export_daily_xlsx(
    date: str = Query(default_factory=today_str, description="YYYY-MM-DD"),
    session: AsyncSession = Depends(get_session),
) -> Response:
    return await export_daily(date=date, session=session)


@router.get("/export/phrase-counts")
async def export_phrase_counts(
    # Filters
    year: int | None = Query(default=None, ge=1970, le=2100, description="Год, например 2025"),
    dateFrom: str | None = Query(default=None, description="ISO datetime, например 2025-01-01T00:00:00"),
    dateTo: str | None = Query(default=None, description="ISO datetime, например 2025-12-31T23:59:59"),
    clientName: str | None = Query(default=None, description="Контрагент/клиент (поиск по подстроке)"),
    # What to count
    phraseA: str = Query(default="Снятие не по расписанию", min_length=1),
    phraseB: str = Query(default="Объект не поставлен под охрану по расписанию", min_length=1),
    # Output
    limit: int = Query(default=50000, ge=1, le=200000),
    session: AsyncSession = Depends(get_session),
) -> Response:
    """Экспорт агрегированного отчёта (по объектам) по двум ключевым фразам.

    Нужен для периодических запросов вида:
    - за год N по контрагенту X: сколько было событий типа A и B.
    """
    dt_from: datetime | None = None
    dt_to: datetime | None = None

    if year is not None:
        dt_from = datetime(year, 1, 1, 0, 0, 0)
        dt_to = datetime(year, 12, 31, 23, 59, 59, 999999)

    if dateFrom:
        parsed = _parse_dt(dateFrom)
        if parsed:
            dt_from = parsed
    if dateTo:
        parsed = _parse_dt(dateTo)
        if parsed:
            dt_to = parsed

    filters: list[Any] = []
    if dt_from is not None:
        filters.append(Event.timestamp >= dt_from)
    if dt_to is not None:
        filters.append(Event.timestamp <= dt_to)

    client = (clientName or "").strip()
    if client:
        needle = f"%{client}%"
        filters.append(or_(Object.client_name.ilike(needle), Event.client_name.ilike(needle)))

    # Only keep rows that match at least one phrase (for performance + relevance)
    p_a = f"%{phraseA.strip()}%"
    p_b = f"%{phraseB.strip()}%"

    where = and_(*filters) if filters else None

    # Prefer objects snapshot for better names/addresses when event has only object_id.
    obj_name = func.coalesce(Object.name, Event.object_name)
    obj_addr = func.coalesce(Object.address, Event.location)

    a_count = func.sum(case((or_(Event.description.ilike(p_a), Event.code_text.ilike(p_a)), 1), else_=0))
    b_count = func.sum(case((or_(Event.description.ilike(p_b), Event.code_text.ilike(p_b)), 1), else_=0))

    stmt = (
        select(
            Event.object_id.label("object_id"),
            obj_name.label("object_name"),
            obj_addr.label("address"),
            a_count.label("phrase_a_count"),
            b_count.label("phrase_b_count"),
        )
        .select_from(Event)
        .outerjoin(Object, Object.id == Event.object_id)
        .group_by(Event.object_id, Object.name, Event.object_name, Object.address, Event.location)
        .having(or_(a_count > 0, b_count > 0))
        .order_by(obj_name.asc())
        .limit(limit)
    )
    if where is not None:
        stmt = stmt.where(where)

    rows = (await session.execute(stmt)).all()

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = cast(Any, wb.active)
    ws.title = "Фразы"

    headers = [
        "Номер объекта",
        "Название объекта",
        "Адрес",
        phraseA,
        phraseB,
        "Примечание",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for object_id, object_name, address, c_a, c_b in rows:
        ws.append(
            [
                object_id or "",
                object_name or "",
                address or "",
                int(c_a or 0),
                int(c_b or 0),
                "",
            ]
        )

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20

    out = BytesIO()
    wb.save(out)
    content = out.getvalue()
    y = str(year) if year is not None else "custom"
    safe_client = client.replace('"', "").replace("'", "").strip() or "all"
    filename = f"phrase-counts-{y}-{safe_client}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/event-codes")
async def list_event_codes(
    query: str | None = Query(default=None, description="Поиск по коду или расшифровке"),
    limit: int = Query(default=100, ge=1, le=500),
    session: AsyncSession = Depends(get_session),
) -> list[dict]:
    """Справочник кодов событий для UI.

    Возвращает коды, которые реально встречаются в локальной таблице events,
    чтобы UI мог выбирать по коду (например E1001), показывая расшифровку.
    """
    q = (query or "").strip()

    stmt = (
        select(
            Event.code.label("code"),
            func.max(Event.code_text).label("codeText"),
            func.count(func.distinct(func.coalesce(Event.parent_event_id, Event.id))).label("count"),
        )
        .where(Event.code.isnot(None))
        .group_by(Event.code)
        .order_by(func.count(func.distinct(func.coalesce(Event.parent_event_id, Event.id))).desc())
        .limit(limit)
    )

    if q:
        needle = f"%{q}%"
        stmt = stmt.where(or_(Event.code.ilike(needle), Event.code_text.ilike(needle)))

    rows = (await session.execute(stmt)).all()
    return [
        {
            "code": code,
            "codeText": code_text,
            "count": int(count or 0),
        }
        for code, code_text, count in rows
        if code
    ]


@router.get("/export/objects-by-code")
async def export_objects_by_code(
    eventCode: str = Query(min_length=1, max_length=16, description="Код события, например E1001"),
    dateFrom: str | None = Query(default=None, description="ISO datetime или YYYY-MM-DD"),
    dateTo: str | None = Query(default=None, description="ISO datetime или YYYY-MM-DD"),
    year: int | None = Query(default=None, ge=1970, le=2100, description="Год (если указан — задаёт период)"),
    clientName: str | None = Query(default=None, description="Контрагент/клиент (поиск по подстроке)"),
    objectQuery: str | None = Query(default=None, description="Поиск по объекту/адресу/ID"),
    limit: int = Query(default=50000, ge=1, le=200000),
    session: AsyncSession = Depends(get_session),
) -> Response:
    """XLSX: по выбранному коду события — сколько раз и по каким объектам за период."""
    dt_from: datetime | None = None
    dt_to: datetime | None = None

    if year is not None:
        dt_from = datetime(year, 1, 1, 0, 0, 0)
        dt_to = datetime(year, 12, 31, 23, 59, 59, 999999)

    if dateFrom:
        parsed_date = _parse_date(dateFrom)
        parsed = _parse_dt(dateFrom) or (
            datetime.combine(parsed_date, datetime.min.time()) if parsed_date else None
        )
        if parsed:
            dt_from = parsed
    if dateTo:
        parsed_date = _parse_date(dateTo)
        parsed = _parse_dt(dateTo) or (
            datetime.combine(parsed_date, datetime.max.time()) if parsed_date else None
        )
        if parsed:
            dt_to = parsed

    filters: list[Any] = [Event.code == eventCode]
    if dt_from is not None:
        filters.append(Event.timestamp >= dt_from)
    if dt_to is not None:
        filters.append(Event.timestamp <= dt_to)

    client = (clientName or "").strip()
    if client:
        needle = f"%{client}%"
        filters.append(or_(Object.client_name.ilike(needle), Event.client_name.ilike(needle)))

    obj_q = (objectQuery or "").strip()
    if obj_q:
        needle = f"%{obj_q}%"
        obj_name = func.coalesce(Object.name, Event.object_name)
        obj_addr = func.coalesce(Object.address, Event.location)
        filters.append(
            or_(
                obj_name.ilike(needle),
                obj_addr.ilike(needle),
                Event.object_id.ilike(needle),
            )
        )

    where = and_(*filters)

    obj_name = func.coalesce(Object.name, Event.object_name)
    obj_addr = func.coalesce(Object.address, Event.location)

    base_stmt = (
        select(
            Event.id.label("event_id"),
            func.coalesce(Event.parent_event_id, Event.id).label("alarm_id"),
            Event.object_id.label("object_id"),
            obj_name.label("object_name"),
            obj_addr.label("address"),
            Event.timestamp.label("timestamp"),
            Event.result_text.label("result_text"),
            Event.meter_count.label("meter_count"),
        )
        .select_from(Event)
        .outerjoin(Object, Object.id == Event.object_id)
        .where(where)
    )
    base = base_stmt.subquery("base")

    agg_stmt = (
        select(
            base.c.object_id,
            base.c.object_name,
            base.c.address,
            func.count(func.distinct(base.c.alarm_id)).label("events_count"),
            func.min(base.c.timestamp).label("first_time"),
            func.max(base.c.timestamp).label("last_time"),
        )
        .group_by(base.c.object_id, base.c.object_name, base.c.address)
        .order_by(func.min(base.c.timestamp).asc(), base.c.object_name.asc())
        .limit(limit)
    )
    agg = agg_stmt.subquery("agg")

    rn_stmt = select(
        base.c.object_id.label("object_id"),
        base.c.alarm_id.label("alarm_id"),
        base.c.result_text.label("result_text"),
        base.c.meter_count.label("meter_count"),
        base.c.timestamp.label("timestamp"),
        func.row_number().over(partition_by=base.c.object_id, order_by=base.c.timestamp.desc()).label("rn"),
    )
    rn = rn_stmt.subquery("rn")
    last_note = (
        select(
            rn.c.object_id.label("object_id"),
            rn.c.alarm_id.label("last_event_id"),
            rn.c.result_text.label("last_result_text"),
            rn.c.meter_count.label("last_meter_count"),
        )
        .where(rn.c.rn == 1)
        .subquery("last_note")
    )

    stmt = (
        select(
            agg.c.object_id,
            agg.c.object_name,
            agg.c.address,
            agg.c.events_count,
            agg.c.first_time,
            agg.c.last_time,
            last_note.c.last_event_id,
            last_note.c.last_meter_count,
            last_note.c.last_result_text,
        )
        .select_from(agg)
        .outerjoin(last_note, last_note.c.object_id == agg.c.object_id)
        .order_by(agg.c.object_name.asc())
    )

    rows = (await session.execute(stmt)).all()

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = cast(Any, wb.active)
    ws.title = "Объекты"

    headers = [
        "Код события",
        "Номер объекта",
        "Название объекта",
        "Адрес",
        "Количество событий",
        "Первое срабатывание",
        "Последнее срабатывание",
        "ID события (аг.)",
        "Параметр (MeterCount)",
        "Комментарий оператора (Result_Text)",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for object_id, object_name, address, events_count, first_time, last_time, last_event_id, last_meter_count, last_result_text in rows:
        ws.append(
            [
                eventCode,
                object_id or "",
                object_name or "",
                address or "",
                int(events_count or 0),
                first_time.isoformat() if first_time else "",
                last_time.isoformat() if last_time else "",
                _agency_event_id(last_event_id) or "",
                last_meter_count or "",
                last_result_text or "",
            ]
        )

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 32
    ws.column_dimensions["J"].width = 35

    out = BytesIO()
    wb.save(out)
    content = out.getvalue()
    safe_code = eventCode.replace("/", "_").replace("\\", "_")
    filename = f"objects-by-code-{safe_code}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/objects-by-code/xlsx")
async def export_objects_by_code_xlsx(
    eventCode: str = Query(min_length=1, max_length=16, description="Код события, например E1001"),
    dateFrom: str | None = Query(default=None, description="ISO datetime или YYYY-MM-DD"),
    dateTo: str | None = Query(default=None, description="ISO datetime или YYYY-MM-DD"),
    year: int | None = Query(default=None, ge=1970, le=2100, description="Год (если указан — задаёт период)"),
    clientName: str | None = Query(default=None, description="Контрагент/клиент (поиск по подстроке)"),
    objectQuery: str | None = Query(default=None, description="Поиск по объекту/адресу/ID"),
    limit: int = Query(default=50000, ge=1, le=200000),
    session: AsyncSession = Depends(get_session),
) -> Response:
    return await export_objects_by_code(
        eventCode=eventCode,
        dateFrom=dateFrom,
        dateTo=dateTo,
        year=year,
        clientName=clientName,
        objectQuery=objectQuery,
        limit=limit,
        session=session,
    )
