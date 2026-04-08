from __future__ import annotations

import asyncio
import re
from datetime import datetime, timedelta
from typing import Any, Annotated

from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy import and_, case, extract, func, or_, select
from sqlalchemy.sql import desc
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.responses import Response

from app.api.v1.deps import require_permissions
from app.core.config import settings
from app.db.session import get_session
from app.integrations.agency_mssql import (
    fetch_alarm_stands_analysis as fetch_alarm_stands_analysis_mssql,
    fetch_gbr_archive_trips as fetch_gbr_archive_trips_mssql,
    fetch_gbr_group_statuses as fetch_gbr_group_statuses_mssql,
)
from app.integrations.agency_sqlite import (
    fetch_gbr_archive_trips as fetch_gbr_archive_trips_sqlite,
    fetch_gbr_group_statuses,
)
from app.models.event import Event
from app.models.event_action import EventAction
from app.models.object import Responsible

router = APIRouter(prefix="/analytics")


def _seconds_between(end_ts, start_ts):
    """Return duration in seconds between two timestamps.

    - SQLite: julianday() diff * 86400
    - Postgres: EXTRACT(EPOCH FROM (end - start))
    """

    if settings.database_url.lower().startswith("postgresql"):
        return extract("epoch", end_ts - start_ts)
    return (func.julianday(end_ts) - func.julianday(start_ts)) * 86400.0


def _parse_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        v = str(value).strip()
        # Python's datetime.fromisoformat() does NOT accept trailing 'Z'.
        # Frontend commonly sends ISO strings like "2026-03-01T00:00:00.000Z".
        if v.endswith("Z"):
            v = v[:-1] + "+00:00"

        dt = datetime.fromisoformat(v)
        # DB stores timestamps as naive datetimes (no timezone). If the frontend sends
        # tz-aware ISO strings (e.g. trailing 'Z' or '+03:00'), normalize them to the
        # server's *local* timezone and drop tzinfo so comparisons match stored values.
        # This also satisfies asyncpg which rejects tz-aware datetimes for TIMESTAMP WITHOUT TIME ZONE.
        if dt.tzinfo is not None:
            return dt.astimezone().replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _action_name_matches(col, patterns: list[str]):
    # Use ILIKE semantics where available; SQLAlchemy will emulate it on SQLite.
    # Patterns should include % wildcards.
    return or_(*[col.ilike(p) for p in patterns])


def _gbr_called_match(col):
    return _action_name_matches(
        col,
        [
            "%Вызван%груп%",
            "%Вызван%реаг%",
            "%Вызван%ГБР%",
            "%Вызов%груп%",
            "%Вызов%реаг%",
            "%Вызов%ГБР%",
            "%Направ%груп%",
            "%Направ%реаг%",
            "%Направ%ГБР%",
            "%Отправ%груп%",
            "%Отправ%реаг%",
            "%Отправ%ГБР%",
            "%Выезд%груп%",
            "%Выезд%реаг%",
            "%Выезд%ГБР%",
        ],
    )


def _gbr_called_loose_match(col):
    return _action_name_matches(col, ["%Вызван%", "%Направ%", "%Отправ%", "%Выезд%", "%Следу%"])


def _gbr_arrived_match(col):
    return _action_name_matches(
        col,
        [
            "%Приб%груп%",
            "%Приб%реаг%",
            "%Приб%ГБР%",
            "%На объект%",
            "%Доех%объект%",
        ],
    )


def _gbr_arrived_loose_match(col):
    return _action_name_matches(col, ["%Прибыт%", "%Прибыл%", "%На объект%", "%Доех%"]) 


def _gbr_cancelled_match(col):
    return _action_name_matches(
        col,
        [
            "%Отмен%груп%",
            "%Отмен%реаг%",
            "%Отмен%ГБР%",
            "%Отбой%груп%",
            "%Отбой%реаг%",
            "%Отбой%ГБР%",
        ],
    )


def _gbr_cancelled_loose_match(col):
    return _action_name_matches(col, ["%Отмен%", "%Отбой%", "%Ложн%тревог%", "%Ложный%"]) 


_GBR_ARCHIVE_CANCEL_PATTERNS = (
    "отмен",
    "ложн",
    "свобод",
)


def _gbr_archive_is_cancelled(status_reason: object) -> bool:
    text = str(status_reason or "").strip().lower()
    if not text:
        return False
    return any(pattern in text for pattern in _GBR_ARCHIVE_CANCEL_PATTERNS)


def _gbr_archive_row_to_trip(row: dict[str, Any]) -> dict[str, Any]:
    archive_id = str(row.get("id") or "").strip()
    called_at = row.get("StartTime")
    raw_end = row.get("EndTime")
    status_reason = str(row.get("StatusReason") or "").strip() or None
    cancelled = _gbr_archive_is_cancelled(status_reason)
    arrived_at = raw_end if raw_end is not None and not cancelled else None
    cancelled_at = raw_end if raw_end is not None and cancelled else None

    if arrived_at is not None:
        trip_status = "На объекте"
    elif cancelled_at is not None:
        trip_status = "Свободна"
    else:
        trip_status = "На выезде"

    object_name = str(row.get("ObjectName") or "").strip() or None
    object_address = str(row.get("ObjectAddress") or "").strip() or None
    duration_seconds = row.get("DurationSeconds")
    try:
        duration_value = float(duration_seconds) if duration_seconds is not None else None
    except Exception:
        duration_value = None

    return {
        "eventId": f"archive:{archive_id}" if archive_id else "archive:unknown",
        "agencyEventId": archive_id or None,
        "gbrName": str(row.get("GroupName") or "").strip() or "Не указан",
        "calledAt": called_at.isoformat() if isinstance(called_at, datetime) else None,
        "arrivedAt": arrived_at.isoformat() if isinstance(arrived_at, datetime) else None,
        "cancelledAt": cancelled_at.isoformat() if isinstance(cancelled_at, datetime) else None,
        "lastActionAt": raw_end.isoformat() if isinstance(raw_end, datetime) else (called_at.isoformat() if isinstance(called_at, datetime) else None),
        "objectId": str(row.get("Panel_id") or "").strip() or None,
        "objectName": object_address or object_name,
        "clientName": object_name,
        "responsibleName": None,
        "calledOperator": None,
        "travelSeconds": duration_value,
        "resultText": status_reason,
        "meterCount": None,
        "timeMeterCount": None,
        "tripStatus": trip_status,
    }


async def _fetch_gbr_archive_trips_payload(
    *,
    date_from: str | None,
    date_to: str | None,
    object_id: str | None,
    max_rows: int,
) -> dict[str, Any] | None:
    url = (settings.agency_database_url or "").strip()
    if not url:
        return None

    scheme = (url.split(":", 1)[0] or "").lower()
    if not (scheme.startswith("sqlite") or scheme.startswith("mssql")):
        return None

    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)
    limit = max(1, min(int(max_rows), 50000))

    if scheme.startswith("mssql"):
        return await asyncio.to_thread(
            fetch_gbr_archive_trips_mssql,
            url,
            date_from=dt_from,
            date_to=dt_to,
            group_id=None,
            panel_id=(object_id or None),
            limit=limit,
        )

    return await asyncio.to_thread(
        fetch_gbr_archive_trips_sqlite,
        url,
        date_from=dt_from,
        date_to=dt_to,
        group_id=None,
        panel_id=(object_id or None),
        limit=limit,
    )


def _filter_gbr_archive_trips(
    rows: list[dict[str, Any]],
    *,
    gbr_name: str | None,
    status: str | None,
) -> list[dict[str, Any]]:
    gbr_name_norm = (gbr_name or "").strip().lower()
    status_norm = (status or "all").strip().lower()

    items = [_gbr_archive_row_to_trip(row) for row in rows]
    if gbr_name_norm:
        items = [item for item in items if str(item.get("gbrName") or "").strip().lower() == gbr_name_norm]

    if status_norm in {"all", "arrived"}:
        items = [item for item in items if item.get("arrivedAt")]
    elif status_norm == "cancelled":
        items = [item for item in items if item.get("cancelledAt") and not item.get("arrivedAt")]
    elif status_norm == "called":
        items = [item for item in items if item.get("calledAt") and not item.get("arrivedAt") and not item.get("cancelledAt")]

    items.sort(
        key=lambda item: (
            str(item.get("calledAt") or ""),
            str(item.get("arrivedAt") or ""),
            str(item.get("cancelledAt") or ""),
        ),
        reverse=True,
    )
    return items


def _xlsx_response(data: bytes, filename: str) -> Response:
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _format_seconds_hhmmss(seconds: float | int | None) -> str:
    if seconds is None:
        return ""
    try:
        total = int(round(float(seconds)))
    except Exception:
        return ""
    if total < 0:
        return ""
    h = total // 3600
    m = (total % 3600) // 60
    s = total % 60
    return f"{h:d}:{m:02d}:{s:02d}" if h > 0 else f"{m:d}:{s:02d}"


@router.get("/gbr/statuses")
async def gbr_statuses(
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> dict[str, Any]:
    """Текущие статусы ГБР (живой статус группы, не статус выезда).

    Источник: агентский SQLite-слепок (AGENCY_DATABASE_URL=sqlite:///.../agency_raw.db),
    таблицы GroupResponse + StatusGroupResponse.
    """

    url = (settings.agency_database_url or "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="AGENCY_DATABASE_URL не задан (нужно sqlite:///.../agency_raw.db)")

    scheme = (url.split(":", 1)[0] or "").lower()
    if not (scheme.startswith("sqlite") or scheme.startswith("mssql")):
        raise HTTPException(
            status_code=400,
            detail="Эндпоинт /analytics/gbr/statuses поддерживает только AGENCY_DATABASE_URL=sqlite:///... или mssql+pyodbc://...",
        )

    try:
        if scheme.startswith("mssql"):
            return await asyncio.to_thread(fetch_gbr_group_statuses_mssql, url)
        return await asyncio.to_thread(fetch_gbr_group_statuses, url)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        # Keep message readable for UI; detailed trace is already logged by global handler.
        raise HTTPException(status_code=500, detail=f"Не удалось прочитать статусы ГБР из agency DB: {e}")


@router.get("/gbr/archive-trips")
async def gbr_archive_trips(
    date_from: str | None = Query(default=None, alias="dateFrom"),
    date_to: str | None = Query(default=None, alias="dateTo"),
    group_id: int | None = Query(default=None, ge=1, alias="groupId"),
    panel_id: str | None = Query(default=None, alias="panelId"),
    limit: int = Query(default=500, ge=1, le=5000),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> dict[str, Any]:
    """История выездов ГБР из ArchiveGroupResponse.

    Работает только при AGENCY_DATABASE_URL=sqlite:///.../agency_raw.db
    (слепок из дампов).
    """

    url = (settings.agency_database_url or "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="AGENCY_DATABASE_URL не задан (нужно sqlite:///.../agency_raw.db)")

    scheme = (url.split(":", 1)[0] or "").lower()
    if not (scheme.startswith("sqlite") or scheme.startswith("mssql")):
        raise HTTPException(
            status_code=400,
            detail="Эндпоинт /analytics/gbr/archive-trips поддерживает только AGENCY_DATABASE_URL=sqlite:///... или mssql+pyodbc://...",
        )

    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)
    if date_from and dt_from is None:
        raise HTTPException(status_code=400, detail="Некорректный dateFrom (ожидается ISO дата/время)")
    if date_to and dt_to is None:
        raise HTTPException(status_code=400, detail="Некорректный dateTo (ожидается ISO дата/время)")

    try:
        if scheme.startswith("mssql"):
            return await asyncio.to_thread(
                fetch_gbr_archive_trips_mssql,
                url,
                date_from=dt_from,
                date_to=dt_to,
                group_id=group_id,
                panel_id=panel_id,
                limit=limit,
            )

        return await asyncio.to_thread(
            fetch_gbr_archive_trips_sqlite,
            url,
            date_from=dt_from,
            date_to=dt_to,
            group_id=group_id,
            panel_id=panel_id,
            limit=limit,
        )
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Не удалось прочитать историю выездов ГБР: {e}")


@router.get("/gbr/archive-summary")
async def gbr_archive_summary(
    date_from: str | None = Query(default=None, alias="dateFrom"),
    date_to: str | None = Query(default=None, alias="dateTo"),
    gbr_name: str | None = Query(default=None, alias="gbrName"),
    panel_id: str | None = Query(default=None, alias="panelId"),
    limit: int = Query(default=5000, ge=1, le=20000),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> dict[str, Any]:
    """Сводка по экипажам ГБР из ArchiveGroupResponse.

    Это более надёжный источник для пункта 5, чем eventservice:
    считаем реальные архивные выезды и их длительность по каждому экипажу.
    """

    url = (settings.agency_database_url or "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="AGENCY_DATABASE_URL не задан (нужно sqlite:///.../agency_raw.db или mssql+pyodbc://...)")

    scheme = (url.split(":", 1)[0] or "").lower()
    if not (scheme.startswith("sqlite") or scheme.startswith("mssql")):
        raise HTTPException(
            status_code=400,
            detail="Эндпоинт /analytics/gbr/archive-summary поддерживает только AGENCY_DATABASE_URL=sqlite:///... или mssql+pyodbc://...",
        )

    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)
    if date_from and dt_from is None:
        raise HTTPException(status_code=400, detail="Некорректный dateFrom (ожидается ISO дата/время)")
    if date_to and dt_to is None:
        raise HTTPException(status_code=400, detail="Некорректный dateTo (ожидается ISO дата/время)")

    try:
        if scheme.startswith("mssql"):
            payload = await asyncio.to_thread(
                fetch_gbr_archive_trips_mssql,
                url,
                date_from=dt_from,
                date_to=dt_to,
                group_id=None,
                panel_id=panel_id,
                limit=limit,
            )
        else:
            payload = await asyncio.to_thread(
                fetch_gbr_archive_trips_sqlite,
                url,
                date_from=dt_from,
                date_to=dt_to,
                group_id=None,
                panel_id=panel_id,
                limit=limit,
            )
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Не удалось построить сводку выездов ГБР: {e}")

    rows = list(payload.get("rows") or [])

    gbr_name_norm = (gbr_name or "").strip().lower()
    if gbr_name_norm:
        rows = [r for r in rows if str(r.get("GroupName") or "").strip().lower() == gbr_name_norm]

    summary: dict[tuple[int | None, str], dict[str, Any]] = {}
    for row in rows:
        group_id = row.get("Group_id")
        group_name = str(row.get("GroupName") or "").strip() or (f"Группа #{group_id}" if group_id else "Не указан")
        key = (int(group_id) if group_id is not None else None, group_name)
        start_time = row.get("StartTime")
        duration = row.get("DurationSeconds")
        try:
            duration_sec = int(duration) if duration is not None else None
        except Exception:
            duration_sec = None

        item = summary.setdefault(
            key,
            {
                "groupId": key[0],
                "gbrName": group_name,
                "tripsCount": 0,
                "objectsCount": 0,
                "totalDurationSeconds": 0,
                "avgDurationSeconds": None,
                "minDurationSeconds": None,
                "maxDurationSeconds": None,
                "firstStartTime": None,
                "lastStartTime": None,
                "_objects": set(),
                "_durations_count": 0,
            },
        )

        item["tripsCount"] += 1
        panel = str(row.get("Panel_id") or "").strip()
        if panel:
            item["_objects"].add(panel)

        if isinstance(start_time, datetime):
            current_first = item["firstStartTime"]
            current_last = item["lastStartTime"]
            if current_first is None or start_time < current_first:
                item["firstStartTime"] = start_time
            if current_last is None or start_time > current_last:
                item["lastStartTime"] = start_time

        if duration_sec is not None and duration_sec >= 0:
            item["totalDurationSeconds"] += duration_sec
            item["_durations_count"] += 1
            if item["minDurationSeconds"] is None or duration_sec < item["minDurationSeconds"]:
                item["minDurationSeconds"] = duration_sec
            if item["maxDurationSeconds"] is None or duration_sec > item["maxDurationSeconds"]:
                item["maxDurationSeconds"] = duration_sec

    result_rows: list[dict[str, Any]] = []
    for item in summary.values():
        item["objectsCount"] = len(item.pop("_objects"))
        durations_count = int(item.pop("_durations_count") or 0)
        if durations_count > 0:
            item["avgDurationSeconds"] = round(float(item["totalDurationSeconds"]) / durations_count, 2)
        else:
            item["avgDurationSeconds"] = None

        first_start = item.get("firstStartTime")
        last_start = item.get("lastStartTime")
        item["firstStartTime"] = first_start.isoformat() if isinstance(first_start, datetime) else None
        item["lastStartTime"] = last_start.isoformat() if isinstance(last_start, datetime) else None
        result_rows.append(item)

    result_rows.sort(key=lambda r: (-int(r.get("tripsCount") or 0), str(r.get("gbrName") or "")))

    return {
        "snapshotAt": payload.get("snapshotAt"),
        "totalTrips": len(rows),
        "rows": result_rows,
    }


@router.get("/alarms/stands")
async def alarms_stands(
    date_from: str | None = Query(default=None, alias="dateFrom"),
    date_to: str | None = Query(default=None, alias="dateTo"),
    q: str | None = Query(default=None, description="Search by panel/object/address/code"),
    limit: int = Query(default=200, ge=1, le=1000),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> dict[str, Any]:
    """Анализ тревог по объектам из dbo.Stands (MSSQL агентства).

    Возвращает активные записи Stands (standorkey=0, TimeEnd is null/active)
    и статистику по архиву событий за период.
    """

    url = (settings.agency_database_url or "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="AGENCY_DATABASE_URL не задан (нужно mssql+pyodbc://...)")

    scheme = (url.split(":", 1)[0] or "").lower()
    if not scheme.startswith("mssql"):
        raise HTTPException(
            status_code=400,
            detail="Эндпоинт /analytics/alarms/stands поддерживает только AGENCY_DATABASE_URL=mssql+pyodbc://...",
        )

    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)
    if date_from and dt_from is None:
        raise HTTPException(status_code=400, detail="Некорректный dateFrom (ожидается ISO дата/время)")
    if date_to and dt_to is None:
        raise HTTPException(status_code=400, detail="Некорректный dateTo (ожидается ISO дата/время)")

    now = datetime.utcnow()
    if dt_to is None:
        dt_to = now
    if dt_from is None:
        dt_from = now - timedelta(hours=24)

    try:
        return await asyncio.to_thread(
            fetch_alarm_stands_analysis_mssql,
            url,
            archives_db_name=settings.agency_archives_db_name,
            date_from=dt_from,
            date_to=dt_to,
            q=q,
            limit=limit,
        )
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Не удалось выполнить анализ стендов: {e}")


@router.get("/operators/live")
async def operators_live(
    window_minutes: int = Query(60, ge=1, le=1440, alias="windowMinutes"),
    online_minutes: int = Query(10, ge=1, le=240, alias="onlineMinutes"),
    session: AsyncSession = Depends(get_session),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> list[dict[str, Any]]:
    """Операторы в реальном времени (эффективность/онлайн).

    Источник: event_actions.

    Важно: "онлайн" вычисляется по последнему действию (event_actions.action_time)
    и порогу onlineMinutes.
    """

    now = datetime.utcnow()
    dt_window = now - timedelta(minutes=int(window_minutes))
    dt_15m = now - timedelta(minutes=15)
    dt_5m = now - timedelta(minutes=5)
    dt_base = now - timedelta(minutes=max(int(window_minutes), 15))

    operator_col = EventAction.operator_name

    # Last action per operator (time) -> join back to fetch name/computer.
    last_ts_sq = (
        select(
            operator_col.label("operator"),
            func.max(EventAction.action_time).label("lastActionAt"),
        )
        .where(operator_col.is_not(None))
        .group_by(operator_col)
        .subquery("last_ts")
    )

    last_rows_sq = (
        select(
            EventAction.operator_name.label("operator"),
            last_ts_sq.c.lastActionAt.label("lastActionAt"),
            EventAction.action_name.label("lastActionName"),
            EventAction.computer.label("computer"),
        )
        .select_from(EventAction)
        .join(
            last_ts_sq,
            and_(
                EventAction.operator_name == last_ts_sq.c.operator,
                EventAction.action_time == last_ts_sq.c.lastActionAt,
            ),
        )
        .subquery("last_rows")
    )

    last_sq = (
        select(
            last_rows_sq.c.operator,
            func.max(last_rows_sq.c.lastActionAt).label("lastActionAt"),
            func.max(last_rows_sq.c.lastActionName).label("lastActionName"),
            func.max(last_rows_sq.c.computer).label("computer"),
        )
        .group_by(last_rows_sq.c.operator)
        .subquery("last")
    )

    # Activity counters for the requested window + fixed short windows.
    actions_5m = func.sum(case((EventAction.action_time >= dt_5m, 1), else_=0)).label("actions5m")
    actions_15m = func.sum(case((EventAction.action_time >= dt_15m, 1), else_=0)).label("actions15m")
    actions_window = func.sum(case((EventAction.action_time >= dt_window, 1), else_=0)).label(
        "actionsWindow"
    )

    event_id_in_window = case((EventAction.action_time >= dt_window, EventAction.event_id), else_=None)
    events_window = func.count(func.distinct(event_id_in_window)).label("eventsWindow")

    counts_sq = (
        select(
            operator_col.label("operator"),
            actions_5m,
            actions_15m,
            actions_window,
            events_window,
        )
        .where(operator_col.is_not(None))
        .where(EventAction.action_time >= dt_base)
        .group_by(operator_col)
        .subquery("counts")
    )

    # Handling time: only events fully handled inside window.
    accept_ts = func.min(
        case(
            (EventAction.action_name.like("Прием%"), EventAction.action_time),
            else_=None,
        )
    ).label("accept_ts")
    end_ts = func.max(
        case(
            (EventAction.action_name == "Окончание обработки", EventAction.action_time),
            else_=None,
        )
    ).label("end_ts")

    per_event = (
        select(
            EventAction.event_id.label("event_id"),
            operator_col.label("operator"),
            accept_ts,
            end_ts,
        )
        .where(operator_col.is_not(None))
        .where(EventAction.action_time >= dt_window)
        .group_by(EventAction.event_id, operator_col)
    ).subquery("per_event")

    duration_seconds = _seconds_between(per_event.c.end_ts, per_event.c.accept_ts).label(
        "duration_seconds"
    )

    handling_sq = (
        select(
            per_event.c.operator.label("operator"),
            func.avg(duration_seconds).label("avgHandlingSeconds"),
            func.count().label("handledEvents"),
        )
        .where(per_event.c.accept_ts.is_not(None))
        .where(per_event.c.end_ts.is_not(None))
        .where(duration_seconds >= 0)
        .group_by(per_event.c.operator)
        .subquery("handling")
    )

    q = (
        select(
            last_sq.c.operator,
            last_sq.c.lastActionAt,
            last_sq.c.lastActionName,
            last_sq.c.computer,
            func.coalesce(counts_sq.c.actions5m, 0).label("actions5m"),
            func.coalesce(counts_sq.c.actions15m, 0).label("actions15m"),
            func.coalesce(counts_sq.c.actionsWindow, 0).label("actionsWindow"),
            func.coalesce(counts_sq.c.eventsWindow, 0).label("eventsWindow"),
            func.coalesce(handling_sq.c.avgHandlingSeconds, None).label("avgHandlingSeconds"),
            func.coalesce(handling_sq.c.handledEvents, 0).label("handledEvents"),
        )
        .select_from(last_sq)
        .outerjoin(counts_sq, counts_sq.c.operator == last_sq.c.operator)
        .outerjoin(handling_sq, handling_sq.c.operator == last_sq.c.operator)
        .order_by(desc(func.coalesce(counts_sq.c.actionsWindow, 0)))
    )

    rows = (await session.execute(q)).all()
    online_cutoff = now - timedelta(minutes=int(online_minutes))

    out: list[dict[str, Any]] = []
    for (
        operator,
        last_action_at,
        last_action_name,
        computer,
        actions5m,
        actions15m,
        actionsWindow,
        eventsWindow,
        avgHandlingSeconds,
        handledEvents,
    ) in rows:
        if not operator:
            continue

        last_dt: datetime | None = last_action_at if isinstance(last_action_at, datetime) else None
        seconds_since = None
        if last_dt is not None:
            seconds_since = int((now - last_dt).total_seconds())

        online = bool(last_dt is not None and last_dt >= online_cutoff)

        out.append(
            {
                "operator": operator,
                "computer": computer,
                "online": online,
                "lastActionAt": last_dt.isoformat() if last_dt else None,
                "lastActionName": last_action_name,
                "secondsSinceLastAction": seconds_since,
                "actions5m": int(actions5m or 0),
                "actions15m": int(actions15m or 0),
                "actionsWindow": int(actionsWindow or 0),
                "eventsWindow": int(eventsWindow or 0),
                "avgHandlingSeconds": float(avgHandlingSeconds) if avgHandlingSeconds is not None else None,
                "handledEvents": int(handledEvents or 0),
                "windowMinutes": int(window_minutes),
                "onlineMinutes": int(online_minutes),
            }
        )

    out.sort(key=lambda r: (1 if r.get("online") else 0, int(r.get("actionsWindow") or 0)), reverse=True)
    return out


@router.get("/operators/handling")
async def operators_handling_time(
    date_from: str | None = Query(None, alias="dateFrom"),
    date_to: str | None = Query(None, alias="dateTo"),
    operator: str | None = Query(None),
    gbr_name: str | None = Query(None, alias="gbrName"),
    session: AsyncSession = Depends(get_session),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> list[dict[str, Any]]:
    """Оценка скорости обработки по операторам.

    Считаем для каждой пары (event_id, operator):
    - accept_time = min(action_time) для action_name, начинающегося с "Прием"
    - end_time = max(action_time) для action_name == "Окончание обработки"
    duration_seconds = end_time - accept_time

    Затем агрегируем по operator.
    """

    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)

    accept_ts = func.min(
        case(
            (EventAction.action_name.like("Прием%"), EventAction.action_time),
            else_=None,
        )
    ).label("accept_ts")

    end_ts = func.max(
        case(
            (EventAction.action_name == "Окончание обработки", EventAction.action_time),
            else_=None,
        )
    ).label("end_ts")

    per_event = (
        select(
            EventAction.event_id.label("event_id"),
            EventAction.operator_name.label("operator"),
            accept_ts,
            end_ts,
        )
        .group_by(EventAction.event_id, EventAction.operator_name)
    )

    if dt_from is not None:
        per_event = per_event.where(EventAction.action_time >= dt_from)
    if dt_to is not None:
        per_event = per_event.where(EventAction.action_time <= dt_to)
    if operator:
        per_event = per_event.where(EventAction.operator_name == operator)
    if gbr_name:
        per_event = per_event.where(EventAction.gbr_name == gbr_name)

    per_event_sq = per_event.subquery("per_event")

    duration_seconds = _seconds_between(per_event_sq.c.end_ts, per_event_sq.c.accept_ts).label(
        "duration_seconds"
    )

    agg = (
        select(
            per_event_sq.c.operator.label("operator"),
            func.count().label("events"),
            func.avg(duration_seconds).label("avgSeconds"),
            func.min(duration_seconds).label("minSeconds"),
            func.max(duration_seconds).label("maxSeconds"),
        )
        .where(per_event_sq.c.operator.is_not(None))
        .where(per_event_sq.c.accept_ts.is_not(None))
        .where(per_event_sq.c.end_ts.is_not(None))
        .where(duration_seconds >= 0)
        .group_by(per_event_sq.c.operator)
        .order_by(func.count().desc())
    )

    rows = (await session.execute(agg)).all()
    out: list[dict[str, Any]] = []
    for operator, events, avg_s, min_s, max_s in rows:
        out.append(
            {
                "operator": operator,
                "events": int(events or 0),
                "avgSeconds": float(avg_s or 0.0),
                "minSeconds": float(min_s or 0.0),
                "maxSeconds": float(max_s or 0.0),
            }
        )
    return out


@router.get("/filters")
async def analytics_filters(
    session: AsyncSession = Depends(get_session),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> dict[str, Any]:
    operators = (
        await session.execute(
            select(EventAction.operator_name)
            .where(EventAction.operator_name.is_not(None))
            .distinct()
            .order_by(EventAction.operator_name.asc())
        )
    ).scalars().all()

    action_names = (
        await session.execute(
            select(EventAction.action_name)
            .where(EventAction.action_name.is_not(None))
            .distinct()
            .order_by(EventAction.action_name.asc())
        )
    ).scalars().all()

    gbr_names = (
        await session.execute(
            select(EventAction.gbr_name)
            .where(EventAction.gbr_name.is_not(None))
            .distinct()
            .order_by(EventAction.gbr_name.asc())
        )
    ).scalars().all()

    min_ts, max_ts = (
        await session.execute(select(func.min(EventAction.action_time), func.max(EventAction.action_time)))
    ).one()

    return {
        "operators": [o for o in operators if o],
        "actionNames": [a for a in action_names if a],
        "gbrNames": [g for g in gbr_names if g],
        "dateMin": min_ts.isoformat() if isinstance(min_ts, datetime) else None,
        "dateMax": max_ts.isoformat() if isinstance(max_ts, datetime) else None,
    }


@router.get("/operators/activity")
async def operators_activity(
    bucket: str = Query("day", pattern="^(day|month)$"),
    date_from: str | None = Query(None, alias="dateFrom"),
    date_to: str | None = Query(None, alias="dateTo"),
    operator: str | None = Query(None),
    session: AsyncSession = Depends(get_session),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> list[dict[str, Any]]:
    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)

    # SQLite-friendly bucketing; for Postgres this still works in many cases but is not perfect.
    if bucket == "month":
        bucket_expr = func.strftime("%Y-%m", EventAction.action_time)
    else:
        bucket_expr = func.date(EventAction.action_time)

    q = (
        select(
            bucket_expr.label("bucket"),
            EventAction.operator_name.label("operator"),
            func.count().label("actions"),
        )
        .where(EventAction.operator_name.is_not(None))
        .group_by(bucket_expr, EventAction.operator_name)
        .order_by(bucket_expr.asc(), func.count().desc())
    )

    if dt_from is not None:
        q = q.where(EventAction.action_time >= dt_from)
    if dt_to is not None:
        q = q.where(EventAction.action_time <= dt_to)
    if operator:
        q = q.where(EventAction.operator_name == operator)

    rows = (await session.execute(q)).all()
    return [{"bucket": b, "operator": o, "actions": int(c or 0)} for (b, o, c) in rows]


@router.get("/gbr/trips")
async def gbr_trips(
    date_from: Annotated[str | None, Query(alias="dateFrom")] = None,
    date_to: Annotated[str | None, Query(alias="dateTo")] = None,
    gbr_name: Annotated[str | None, Query(alias="gbrName")] = None,
    object_id: Annotated[str | None, Query(alias="objectId")] = None,
    status: Annotated[str | None, Query(pattern="^(all|arrived|cancelled|called)$")] = None,
    limit: Annotated[int, Query(ge=1, le=2000)] = 200,
    offset: Annotated[int, Query(ge=0)] = 0,
    session: AsyncSession = Depends(get_session),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> dict[str, Any]:
    """Отчёт по выездам ГБР.

    Строится из event_actions (eventservice) по действиям:
    - "Вызвана группа реагирования"
    - "Прибытие группы реагирования"
    - "Отмена вызова группы реагирования" (и похожие)

    Возвращает поездки (event_id + gbr_name) с временами и, если есть, объектом из таблицы events.
    """

    try:
        archive_payload = await _fetch_gbr_archive_trips_payload(
            date_from=date_from,
            date_to=date_to,
            object_id=object_id,
            max_rows=50000,
        )
    except Exception:
        archive_payload = None

    if archive_payload is not None:
        filtered = _filter_gbr_archive_trips(
            list(archive_payload.get("rows") or []),
            gbr_name=gbr_name,
            status=status,
        )
        total = len(filtered)
        page = filtered[int(offset) : int(offset) + int(limit)]
        return {"data": page, "total": total, "limit": int(limit), "offset": int(offset)}

    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)

    # Exclude non-GBR dispatch types that may still carry gbr_name in raw data.
    # These are NOT GBR trips and must not be counted in the GBR report.
    # Examples from ops:
    # - "Оповещение х/о" (not a trip)
    # - "Физохрана" (physical security, not a GBR trip)
    non_gbr_exclude = and_(
        # Notify responsible person (Х/О) is not a GBR trip
        ~EventAction.action_name.ilike("%оповещ%х/о%"),
        ~EventAction.action_name.ilike("%х/о%"),
        # Physical security is not a GBR trip
        ~EventAction.action_name.ilike("%физохран%"),
        ~EventAction.action_name.ilike("%физ%охран%"),
    )

    # Match actions robustly but avoid false positives.
    # We use a strict match (mentions group/GBR/react) with a fallback loose match.
    called_match_strict = _gbr_called_match(EventAction.action_name)
    called_match_loose = and_(_gbr_called_loose_match(EventAction.action_name), non_gbr_exclude)

    arrived_match_strict = _gbr_arrived_match(EventAction.action_name)
    arrived_match_loose = and_(_gbr_arrived_loose_match(EventAction.action_name), non_gbr_exclude)

    cancelled_match_strict = _gbr_cancelled_match(EventAction.action_name)
    cancelled_match_loose = and_(_gbr_cancelled_loose_match(EventAction.action_name), non_gbr_exclude)

    any_trip_match = or_(
        called_match_strict,
        called_match_loose,
        arrived_match_strict,
        arrived_match_loose,
        cancelled_match_strict,
        cancelled_match_loose,
    )

    called_ts_strict = func.min(case((called_match_strict, EventAction.action_time), else_=None))
    called_ts_loose = func.min(case((called_match_loose, EventAction.action_time), else_=None))
    any_trip_ts = func.min(case((any_trip_match, EventAction.action_time), else_=None))
    called_ts = func.coalesce(called_ts_strict, called_ts_loose, any_trip_ts).label("called_ts")

    called_op_strict = func.min(case((called_match_strict, EventAction.operator_name), else_=None))
    called_op_loose = func.min(case((called_match_loose, EventAction.operator_name), else_=None))
    any_trip_operator = func.min(case((any_trip_match, EventAction.operator_name), else_=None))
    called_operator = func.coalesce(called_op_strict, called_op_loose, any_trip_operator).label("called_operator")

    arrived_ts_strict = func.min(case((arrived_match_strict, EventAction.action_time), else_=None))
    arrived_ts_loose = func.min(case((arrived_match_loose, EventAction.action_time), else_=None))
    arrived_ts = func.coalesce(arrived_ts_strict, arrived_ts_loose).label("arrived_ts")

    cancelled_ts_strict = func.min(case((cancelled_match_strict, EventAction.action_time), else_=None))
    cancelled_ts_loose = func.min(case((cancelled_match_loose, EventAction.action_time), else_=None))
    cancelled_ts = func.coalesce(cancelled_ts_strict, cancelled_ts_loose).label("cancelled_ts")
    last_action_ts = func.max(EventAction.action_time).label("last_action_ts")

    base = (
        select(
            EventAction.event_id.label("event_id"),
            EventAction.gbr_name.label("gbr_name"),
            called_ts,
            called_operator,
            arrived_ts,
            cancelled_ts,
            last_action_ts,
        )
        .where(EventAction.gbr_name.is_not(None))
        .where(any_trip_match)
        .group_by(EventAction.event_id, EventAction.gbr_name)
    )

    if dt_from is not None:
        base = base.where(EventAction.action_time >= dt_from)
    if dt_to is not None:
        base = base.where(EventAction.action_time <= dt_to)
    if gbr_name:
        base = base.where(EventAction.gbr_name == gbr_name)

    sq = base.subquery("gbr_trips")

    # Enrich with a single (first) responsible per object (Panel_id).
    # We pick a deterministic value via MIN(name) which works on SQLite/Postgres.
    resp_sq = (
        select(
            Responsible.object_id.label("object_id"),
            func.min(Responsible.name).label("responsible_name"),
        )
        .group_by(Responsible.object_id)
        .subquery("resp")
    )

    travel_seconds = _seconds_between(sq.c.arrived_ts, sq.c.called_ts).label("travel_seconds")

    def _agency_event_id(event_id: str | None) -> str | None:
        if not event_id:
            return None
        parts = str(event_id).split(":")
        if len(parts) >= 3:
            return parts[-1] or None
        return None

    def _trip_status(called: object, arrived: object, cancelled: object) -> str | None:
        # Best-effort status derived from the trip state.
        # Full GBR statuses like "На СТО/АЗС/..." require a separate source table.
        if arrived is not None:
            return "На объекте"
        if cancelled is not None:
            return "Свободна"
        if called is not None:
            return "На выезде"
        return None

    q = (
        select(
            sq.c.event_id,
            sq.c.gbr_name,
            sq.c.called_ts,
            sq.c.arrived_ts,
            sq.c.cancelled_ts,
            sq.c.last_action_ts,
            Event.object_id,
            Event.object_name,
            Event.client_name,
            Event.result_text,
            Event.meter_count,
            Event.time_meter_count,
            resp_sq.c.responsible_name,
            sq.c.called_operator,
            travel_seconds,
        )
        .select_from(sq)
        .outerjoin(Event, Event.id == sq.c.event_id)
        .outerjoin(resp_sq, resp_sq.c.object_id == Event.object_id)
        .where(or_(sq.c.called_ts.is_not(None), sq.c.arrived_ts.is_not(None), sq.c.cancelled_ts.is_not(None)))
        .order_by(sq.c.called_ts.desc())
        .offset(offset)
        .limit(limit)
    )

    status_norm = (status or "all").strip().lower()
    if status_norm == "all":
        q = q.where(sq.c.arrived_ts.is_not(None))
    elif status_norm == "arrived":
        q = q.where(sq.c.arrived_ts.is_not(None))
    elif status_norm == "cancelled":
        q = q.where(sq.c.arrived_ts.is_(None)).where(sq.c.cancelled_ts.is_not(None))
    elif status_norm == "called":
        q = q.where(sq.c.arrived_ts.is_(None)).where(sq.c.cancelled_ts.is_(None))

    if object_id:
        q = q.where(Event.object_id == object_id)

    rows = (await session.execute(q)).all()
    items: list[dict[str, Any]] = []
    for (
        event_id,
        gbr,
        called,
        arrived,
        cancelled,
        last_action,
        obj_id,
        obj_name,
        client_name,
        result_text,
        meter_count,
        time_meter_count,
        responsible_name,
        called_operator_name,
        travel_s,
    ) in rows:
        travel_seconds_val: float | None
        travel_seconds_val = None
        if travel_s is not None:
            try:
                travel_seconds_val = float(travel_s)
            except Exception:
                travel_seconds_val = None
        if travel_seconds_val is None and isinstance(called, datetime) and isinstance(arrived, datetime):
            try:
                travel_seconds_val = float((arrived - called).total_seconds())
            except Exception:
                travel_seconds_val = None
        if travel_seconds_val is not None and travel_seconds_val < 0:
            # Data can contain inverted timestamps due to inconsistent action naming;
            # for reporting we show absolute travel time.
            travel_seconds_val = abs(travel_seconds_val)

        items.append(
            {
                "eventId": event_id,
                "agencyEventId": _agency_event_id(event_id),
                "gbrName": gbr,
                "calledAt": called.isoformat() if isinstance(called, datetime) else None,
                "arrivedAt": arrived.isoformat() if isinstance(arrived, datetime) else None,
                "cancelledAt": cancelled.isoformat() if isinstance(cancelled, datetime) else None,
                "lastActionAt": last_action.isoformat() if isinstance(last_action, datetime) else None,
                "objectId": obj_id,
                "objectName": obj_name,
                "clientName": client_name,
                "responsibleName": responsible_name,
                "calledOperator": called_operator_name,
                "travelSeconds": travel_seconds_val,
                "resultText": result_text,
                "meterCount": meter_count,
                "timeMeterCount": time_meter_count.isoformat() if isinstance(time_meter_count, datetime) else None,
                "tripStatus": _trip_status(called, arrived, cancelled),
            }
        )

    # Count (for pagination) - count distinct pairs from the same grouped view.
    count_inner = (
        select(sq.c.event_id)
        .select_from(sq)
        .outerjoin(Event, Event.id == sq.c.event_id)
        .where(or_(sq.c.called_ts.is_not(None), sq.c.arrived_ts.is_not(None), sq.c.cancelled_ts.is_not(None)))
    )

    if status_norm == "all":
        count_inner = count_inner.where(sq.c.arrived_ts.is_not(None))
    elif status_norm == "arrived":
        count_inner = count_inner.where(sq.c.arrived_ts.is_not(None))
    elif status_norm == "cancelled":
        count_inner = count_inner.where(sq.c.arrived_ts.is_(None)).where(sq.c.cancelled_ts.is_not(None))
    elif status_norm == "called":
        count_inner = count_inner.where(sq.c.arrived_ts.is_(None)).where(sq.c.cancelled_ts.is_(None))
    if object_id:
        count_inner = count_inner.where(Event.object_id == object_id)
    count_q = select(func.count()).select_from(count_inner.subquery())
    total = (await session.execute(count_q)).scalar_one()

    return {"data": items, "total": int(total or 0), "limit": limit, "offset": offset}


async def _fetch_all_gbr_trips(
    *,
    date_from: str | None,
    date_to: str | None,
    gbr_name: str | None,
    object_id: str | None,
    status: str | None,
    session: AsyncSession,
    _perm: Any,
    page_size: int = 2000,
    max_rows: int = 50000,
) -> dict[str, Any]:
    """Fetch all GBR trips for export/reporting.

    `gbr_trips` is API-shaped and enforces `limit<=2000`, so we page through results.
    We also guard against producing enormous XLSX files silently.
    """

    first = await gbr_trips(
        date_from=date_from,
        date_to=date_to,
        gbr_name=gbr_name,
        object_id=object_id,
        status=status,
        limit=int(page_size),
        offset=0,
        session=session,
        _perm=_perm,
    )

    total = int(first.get("total") or 0)
    if max_rows and total > int(max_rows):
        raise HTTPException(
            status_code=400,
            detail={
                "code": "TOO_MANY_ROWS",
                "message": f"Слишком много строк для выгрузки: {total}. Сузьте период/фильтры.",
            },
        )

    items: list[dict[str, Any]] = list(first.get("data") or [])
    if total <= len(items) or len(items) < int(page_size):
        first["data"] = items
        first["total"] = total
        return first

    offset = int(page_size)
    while len(items) < total:
        out = await gbr_trips(
            date_from=date_from,
            date_to=date_to,
            gbr_name=gbr_name,
            object_id=object_id,
            status=status,
            limit=int(page_size),
            offset=offset,
            session=session,
            _perm=_perm,
        )
        batch = out.get("data") or []
        if not batch:
            break
        items.extend(batch)
        offset += int(page_size)
        if len(batch) < int(page_size):
            break

    first["data"] = items
    first["total"] = total
    first["limit"] = total
    first["offset"] = 0
    return first


@router.get("/gbr/trips/export")
async def gbr_trips_export(
    date_from: str | None = Query(None, alias="dateFrom"),
    date_to: str | None = Query(None, alias="dateTo"),
    gbr_name: str | None = Query(None, alias="gbrName"),
    object_id: str | None = Query(None, alias="objectId"),
    status: str | None = Query(None, pattern="^(all|arrived|cancelled|called)$"),
    session: AsyncSession = Depends(get_session),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> Response:
    # CSV больше не поддерживаем: отдаём XLSX (как в таблице UI).
    return await gbr_trips_export_table_xlsx(
        date_from=date_from,
        date_to=date_to,
        gbr_name=gbr_name,
        object_id=object_id,
        status=status,
        session=session,
        _perm=_perm,
    )

@router.get("/gbr/trips/export/xlsx")
async def gbr_trips_export_xlsx(
    date_from: str | None = Query(None, alias="dateFrom"),
    date_to: str | None = Query(None, alias="dateTo"),
    gbr_name: str | None = Query(None, alias="gbrName"),
    object_id: str | None = Query(None, alias="objectId"),
    session: AsyncSession = Depends(get_session),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> Response:
    """XLSX: «Рапорт» по выездам ГБР (шаблонный вид).

    Важно: исходный шаблон у пользователя в .xls. Мы формируем .xlsx,
    повторяя структуру (шапка + таблица) и заполняя доступные поля.
    """

    result = await _fetch_all_gbr_trips(
        date_from=date_from,
        date_to=date_to,
        gbr_name=gbr_name,
        object_id=object_id,
        status=None,
        session=session,
        _perm=_perm,
    )

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    def clean_excel_text(value: object) -> object:
        if value is None:
            return ""
        if isinstance(value, str):
            return re.sub(ILLEGAL_CHARACTERS_RE, "", value)
        return value

    wb = Workbook()
    ws = wb.active
    ws.title = "Рапорт"

    # Columns to match the visual template width.
    columns = [
        "№ объекта",
        "Адрес",
        "Шлейф",
        "Инженер",
        "Результат",
        "Дата",
        "ГБР",
        "Вызов",
        "Прибыл",
        "Время в пути",
        "Результат осмотра",
        "Оператор",
        "Заявка",
        "Штраф",
        "Сработок за полгода",
        "ID события (аг.)",
        "Параметр (MeterCount)",
        "Пометка оператора (Result_Text)",
        "Статус",
    ]

    # Header area similar to the screenshot
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value="Рапорт").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Period string
    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)
    period_text = ""
    if dt_from and dt_to:
        period_text = f"За период: {dt_from.strftime('%d.%m.%Y %H:%M')} — {dt_to.strftime('%d.%m.%Y %H:%M')}"
    elif dt_from:
        period_text = f"С: {dt_from.strftime('%d.%m.%Y %H:%M')}"
    elif dt_to:
        period_text = f"До: {dt_to.strftime('%d.%m.%Y %H:%M')}"

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws.cell(row=2, column=1, value=period_text).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(columns))
    ws.cell(row=3, column=1, value="оперативная обстановка следующая:").alignment = Alignment(
        horizontal="center"
    )

    header_row = 5
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Set column widths (approximate)
    widths = [
        12,  # № объекта
        28,  # Адрес
        10,  # Шлейф
        16,  # Инженер
        16,  # Результат
        12,  # Дата
        14,  # ГБР
        18,  # Вызов
        18,  # Прибыл
        12,  # Время в пути
        18,  # Результат осмотра
        16,  # Оператор
        14,  # Заявка
        10,  # Штраф
        18,  # Сработок
        16,  # ID события (аг.)
        28,  # Параметр (MeterCount)
        45,  # Пометка оператора (Result_Text)
        14,  # Статус
    ]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = w

    for col_idx, title in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=title)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    rows = result.get("data") or []
    start_row = header_row + 1
    for i, r in enumerate(rows, start=0):
        row_idx = start_row + i

        called_at = r.get("calledAt")
        arrived_at = r.get("arrivedAt")
        cancelled_at = r.get("cancelledAt")
        gbr = r.get("gbrName") or ""
        obj_id = r.get("objectId") or ""
        obj_name = r.get("objectName") or ""
        client = r.get("clientName") or ""

        # Template columns: fill what we have, rest leave empty.
        values = [
            obj_id,
            (obj_name or client),
            "",  # шлейф
            "",  # инженер
            "",  # результат
            (called_at or "")[:10].replace("-", ".") if called_at else "",  # дата
            gbr,
            (called_at or "")[:19].replace("T", " ") if called_at else "",
            (
                (arrived_at or "")[:19].replace("T", " ")
                if arrived_at
                else ("Отмена" if cancelled_at else "")
            ),
            _format_seconds_hhmmss(r.get("travelSeconds")),
            "",  # результат осмотра
            r.get("calledOperator") or "",
            "",  # заявка
            "",  # штраф
            "",  # сработок
            r.get("agencyEventId") or "",
            r.get("meterCount") or "",
            r.get("resultText") or "",
            r.get("tripStatus") or "",
        ]

        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=clean_excel_text(v))
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # Improve print layout
    ws.freeze_panes = ws["A6"]
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    out = BytesIO()
    wb.save(out)

    name = f"raport-gbr-{datetime.utcnow().date().isoformat()}.xlsx"
    return _xlsx_response(out.getvalue(), name)


@router.get("/gbr/trips/export/table/xlsx")
async def gbr_trips_export_table_xlsx(
    date_from: str | None = Query(None, alias="dateFrom"),
    date_to: str | None = Query(None, alias="dateTo"),
    gbr_name: str | None = Query(None, alias="gbrName"),
    object_id: str | None = Query(None, alias="objectId"),
    status: str | None = Query(None, pattern="^(all|arrived|cancelled|called)$"),
    session: AsyncSession = Depends(get_session),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> Response:
    """XLSX: выгрузка «как в таблице» на странице Отчёт ГБР."""

    result = await _fetch_all_gbr_trips(
        date_from=date_from,
        date_to=date_to,
        gbr_name=gbr_name,
        object_id=object_id,
        status=status,
        session=session,
        _perm=_perm,
    )

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    def clean_excel_text(value: object) -> object:
        if value is None:
            return ""
        if isinstance(value, str):
            return re.sub(ILLEGAL_CHARACTERS_RE, "", value)
        return value

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт ГБР"

    headers = [
        "Вызов",
        "Прибытие",
        "Статус",
        "ГБР",
        "№ объекта",
        "Объект",
        "Ответственный",
        "Оператор",
        "В пути",
        "ID события (аг.)",
        "Параметр (MeterCount)",
        "Пометка оператора (Result_Text)",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    def _fmt_iso(s: str | None) -> str:
        if not s:
            return ""
        try:
            return s.replace("T", " ")[:19]
        except Exception:
            return str(s)

    def _arrival_cell(row: dict[str, Any]) -> str:
        if row.get("arrivedAt"):
            return _fmt_iso(row.get("arrivedAt"))
        if row.get("cancelledAt"):
            return "Отмена"
        return "—"

    for r in (result.get("data") or []):
        ws.append(
            [
                clean_excel_text(_fmt_iso(r.get("calledAt"))),
                clean_excel_text(_arrival_cell(r)),
                clean_excel_text(r.get("tripStatus") or "—"),
                clean_excel_text(r.get("gbrName") or ""),
                clean_excel_text(r.get("objectId") or ""),
                clean_excel_text(r.get("objectName") or ""),
                clean_excel_text((r.get("responsibleName") or r.get("clientName") or "")),
                clean_excel_text(r.get("calledOperator") or ""),
                clean_excel_text((_format_seconds_hhmmss(r.get("travelSeconds")) or "—")),
                clean_excel_text(r.get("agencyEventId") or ""),
                clean_excel_text(r.get("meterCount") or ""),
                clean_excel_text(r.get("resultText") or ""),
            ]
        )

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 30
    ws.column_dimensions["L"].width = 45

    out = BytesIO()
    wb.save(out)

    name = f"gbr-trips-table-{datetime.utcnow().date().isoformat()}.xlsx"
    return _xlsx_response(out.getvalue(), name)


@router.get("/objects/events/summary")
async def object_events_summary(
    object_id: str = Query(..., alias="objectId"),
    date_from: str | None = Query(None, alias="dateFrom"),
    date_to: str | None = Query(None, alias="dateTo"),
    days: int | None = Query(None, ge=1, le=3650, description="Alternative to dateFrom/dateTo"),
    code: str | None = Query(None),
    code_group: int | None = Query(None, alias="codeGroup"),
    session: AsyncSession = Depends(get_session),
    _perm: Any = Depends(require_permissions("analytics:read")),
) -> dict[str, Any]:
    """Сводка по объекту за период: сколько событий/какие коды/статусы/серьёзность."""

    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)
    if days is not None and dt_from is None and dt_to is None:
        dt_to = datetime.utcnow()
        dt_from = dt_to - timedelta(days=int(days))

    filters: list[Any] = [Event.object_id == object_id]
    if dt_from is not None:
        filters.append(Event.timestamp >= dt_from)
    if dt_to is not None:
        filters.append(Event.timestamp <= dt_to)
    if code:
        filters.append(Event.code == code)
    if code_group is not None:
        filters.append(Event.code_group == int(code_group))

    where = and_(*filters) if filters else None

    alarm_id_expr = func.coalesce(Event.parent_event_id, Event.id)

    total_q = select(func.count(func.distinct(alarm_id_expr))).select_from(Event)
    if where is not None:
        total_q = total_q.where(where)
    total = (await session.execute(total_q)).scalar_one()

    by_sev_q = select(Event.severity, func.count(func.distinct(alarm_id_expr))).select_from(Event)
    if where is not None:
        by_sev_q = by_sev_q.where(where)
    by_sev_q = by_sev_q.group_by(Event.severity)

    by_status_q = select(Event.status, func.count(func.distinct(alarm_id_expr))).select_from(Event)
    if where is not None:
        by_status_q = by_status_q.where(where)
    by_status_q = by_status_q.group_by(Event.status)

    by_code_q = select(Event.code_group, Event.code, Event.code_text, func.count(func.distinct(alarm_id_expr))).select_from(Event)
    if where is not None:
        by_code_q = by_code_q.where(where)
    by_code_q = by_code_q.group_by(Event.code_group, Event.code, Event.code_text).order_by(func.count(func.distinct(alarm_id_expr)).desc()).limit(200)

    by_sev = {str(k or ""): int(v or 0) for (k, v) in (await session.execute(by_sev_q)).all()}
    by_status = {str(k or ""): int(v or 0) for (k, v) in (await session.execute(by_status_q)).all()}
    by_code = [
        {"codeGroup": cg, "code": c, "codeText": ct, "count": int(cnt or 0)}
        for (cg, c, ct, cnt) in (await session.execute(by_code_q)).all()
    ]

    return {
        "objectId": object_id,
        "total": int(total or 0),
        "bySeverity": by_sev,
        "byStatus": by_status,
        "byCode": by_code,
    }
