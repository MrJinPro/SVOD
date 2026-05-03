"""
ПЦН отчёт - основной ведомость по тревогам (Дата, № Объекта, Адрес, ФИО, Шлейф, Инженер, ГБР, Оператор, Время вызова, Время прибытия, Результат, Штраф).
"""
from __future__ import annotations

import re
from datetime import datetime, timezone
from typing import Any, cast
from io import BytesIO

from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from app.models.event import Event
from app.models.event_action import EventAction
from app.models.object import Object


async def build_pcn_main_report_xlsx(
    session: AsyncSession,
    *,
    date_from: datetime,
    date_to: datetime,
    operator_name: str | None = None,
    object_id: str | None = None,
    gbr_name: str | None = None,
) -> tuple[bytes, int]:
    """
    Генерирует основной XLSX отчёт ПЦН с колонками:
    Дата тревоги | № Объекта | Адрес объекта | ФИО | Шлейф | Инженер | ГБР | Оператор | 
    Время вызова | Время прибытия | Результат осмотра | Штраф
    
    Returns: (xlsx_bytes, events_count)
    """

    # Фильтры
    filters: list[Any] = [
        Event.timestamp >= date_from,
        Event.timestamp <= date_to,
    ]

    if object_id and (object_id := str(object_id).strip()):
        filters.append(
            or_(
                Event.object_id.ilike(f"%{object_id}%"),
                Object.name.ilike(f"%{object_id}%"),
                Object.address.ilike(f"%{object_id}%"),
            )
        )

    # Query основных данных
    stmt = (
        select(
            Event.id.label("event_id"),
            Event.timestamp.label("timestamp"),
            Event.object_id.label("object_id"),
            func.coalesce(Object.address, Event.location).label("address"),
            func.coalesce(Object.client_name, Event.client_name).label("client_name"),
            Event.line.label("line"),
            Event.zone.label("zone"),
            Event.result_text.label("result_text"),
        )
        .select_from(Event)
        .outerjoin(Object, Object.id == Event.object_id)
        .where(and_(*filters))
        .order_by(Event.timestamp.asc())
        .limit(10000)
    )

    events = (await session.execute(stmt)).all()

    # Enrichment: EventAction данные (ГБР, Оператор, Время прибытия)
    event_ids = [str(e[0]) for e in events]
    action_data: dict[str, dict[str, Any]] = {}

    if event_ids:
        # Берём последний action для каждого события
        for chunk_size in [1000]:  # SQLite может не любить очень большие IN
            for i in range(0, len(event_ids), chunk_size):
                chunk = event_ids[i : i + chunk_size]
                action_stmt = (
                    select(
                        EventAction.event_id,
                        EventAction.operator_name,
                        EventAction.gbr_name,
                        func.max(EventAction.action_time).label("max_action_time"),
                    )
                    .where(EventAction.event_id.in_(chunk))
                    .group_by(EventAction.event_id)
                )
                for event_id, op_name, gbr, action_time in (await session.execute(action_stmt)).all():
                    if event_id not in action_data:
                        action_data[str(event_id)] = {}
                    action_data[str(event_id)]["operator"] = str(op_name or "").strip() or None
                    action_data[str(event_id)]["gbr"] = str(gbr or "").strip() or None
                    action_data[str(event_id)]["arrival_time"] = action_time

    # Фильтрация по оператору и ГБР (если задано)
    filtered_events = []
    for event in events:
        evt_id = str(event[0])
        actions = action_data.get(evt_id, {})
        op = actions.get("operator")
        gbr = actions.get("gbr")

        if operator_name and (operator_name := str(operator_name).strip()):
            if not op or operator_name.lower() not in op.lower():
                continue

        if gbr_name and (gbr_name := str(gbr_name).strip()):
            if not gbr or gbr_name.lower() not in gbr.lower():
                continue

        filtered_events.append((event, actions))

    # Build XLSX
    columns = [
        "Дата тревоги",
        "№ Объекта",
        "Адрес объекта",
        "ФИО",
        "Шлейф",
        "Инженер",
        "ГБР",
        "Оператор",
        "Время вызова",
        "Время прибытия",
        "Результат осмотра",
        "Штраф",
    ]

    wb = Workbook()
    ws = cast(Any, wb.active)
    ws.title = "ПЦН"

    # Header row
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, title in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=title)
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # Data rows
    def clean_text(value: object) -> str:
        if value is None:
            return ""
        text = str(value or "").strip()
        # Remove illegal chars for Excel
        return re.sub(ILLEGAL_CHARACTERS_RE, "", text)

    def fmt_datetime(dt: datetime | None) -> str:
        if not dt:
            return ""
        return dt.strftime("%d.%m.%Y %H:%M:%S")

    def fmt_time(dt: datetime | None) -> str:
        if not dt:
            return ""
        return dt.strftime("%H:%M:%S")

    row_idx = 2
    for event, actions in filtered_events:
        event_id, timestamp, object_id, address, client_name, line, zone_value, result_text = event
        op = actions.get("operator")
        gbr = actions.get("gbr")
        arrival_time = actions.get("arrival_time")

        loop_label = str(line or "").strip()
        if not loop_label and zone_value is not None:
            loop_label = str(zone_value).strip()

        values = [
            fmt_datetime(timestamp),  # Дата тревоги
            object_id or "",  # № Объекта
            clean_text(address) or "",  # Адрес объекта
            clean_text(client_name) or "",  # ФИО
            loop_label,  # Шлейф
            "",  # Инженер (пусто, нет в данных)
            clean_text(gbr) or "",  # ГБР
            clean_text(op) or "",  # Оператор
            fmt_time(timestamp),  # Время вызова
            fmt_time(arrival_time),  # Время прибытия
            clean_text(result_text) or "",  # Результат осмотра
            "0",  # Штраф
        ]

        for col_idx, value in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

        row_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 35
    ws.column_dimensions["L"].width = 10

    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)

    return out.getvalue(), len(filtered_events)
