from __future__ import annotations

import asyncio
import json
import logging
import re
from datetime import datetime, timedelta
from typing import Any

import csv
import io
from pydantic import BaseModel, Field

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from starlette.responses import StreamingResponse
from sqlalchemy import Integer, Select, and_, case, cast, exists, func, literal, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.deps import get_current_user
from app.core.config import settings
from app.db.session import get_session
from app.integrations.agency_mssql import (
    fetch_eventservice_actions_for_event_pairs as fetch_eventservice_actions_for_event_pairs_mssql,
)
from app.integrations.agency_sqlite import (
    fetch_eventservice_actions_for_event_pairs as fetch_eventservice_actions_for_event_pairs_sqlite,
)
from app.models.event import Event
from app.models.event_action import EventAction
from app.models.object import Object
from app.utils.search import (
    query_needles,
    query_prefix_needles,
    query_variants,
    should_search_related_text,
    should_use_light_search,
    tokenize_query,
)

router = APIRouter(prefix="/events")

logger = logging.getLogger(__name__)


class OperatorCommentIn(BaseModel):
    eventId: str = Field(min_length=1, max_length=64)
    resultText: str | None = Field(default=None, max_length=5000)


def _is_operator_handled_predicate() -> Any:
    """Heuristic: event is handled by an operator (not purely system)."""

    return and_(Event.operator_id.is_not(None), Event.operator_id != "")


def _has_operator_comment_predicate() -> Any:
    """Event has an operator comment/note (Result_Text)."""

    return and_(
        Event.result_text.is_not(None),
        func.length(func.trim(Event.result_text)) > 0,
    )


def _accept_action_predicate() -> Any:
    """Best-effort match for operator action: accepted for processing."""

    return or_(
        EventAction.action_name == "Прием на обработку",
        EventAction.action_name.ilike("%прин%в обработ%"),
        EventAction.action_name.ilike("%прием%обработ%"),
    )


def _accepted_action_exists() -> Any:
    return exists(
        select(1)
        .select_from(EventAction)
        .where(EventAction.event_id == Event.id)
        .where(EventAction.operator_name.is_not(None))
        .where(EventAction.operator_name != "")
        .where(_accept_action_predicate())
    )


def _numeric_event_id_predicate(dialect_name: str | None) -> Any:
    if (dialect_name or "").lower() == "postgresql":
        return Event.id.op("~")(r"^\d+$")

    # SQLite: emulate "all digits" via GLOB
    # - starts with a digit
    # - and does NOT contain any non-digit
    return and_(Event.id.op("GLOB")("[0-9]*"), ~Event.id.op("GLOB")("*[^0-9]*"))


def _operator_action_exists(*, dialect_name: str | None) -> Any:
    """Any operator action exists for the event (best-effort)."""

    # Some deployments store Event.id as "mssql:{date_key}:{event_id}", while others
    # store a numeric Event_id string. Support both ways of linking to EventAction.
    raw_id_expr = case(
        (_numeric_event_id_predicate(dialect_name), cast(Event.id, Integer)),
        else_=None,
    )
    if (dialect_name or "").lower() == "sqlite":
        date_key_expr = cast(func.strftime("%Y%m%d", Event.timestamp), Integer)
    else:
        date_key_expr = cast(func.to_char(Event.timestamp, "YYYYMMDD"), Integer)

    return exists(
        select(1)
        .select_from(EventAction)
        .where(
            and_(
                EventAction.operator_name.is_not(None),
                EventAction.operator_name != "",
                or_(
                    EventAction.event_id == Event.id,
                    and_(
                        raw_id_expr.is_not(None),
                        EventAction.raw_event_id == raw_id_expr,
                        EventAction.date_key == date_key_expr,
                    ),
                ),
            )
        )
    )


async def _actions_linked_present(session: AsyncSession) -> bool:
    """True if there is at least one EventAction that actually joins to an Event."""

    try:
        any_join = (
            await session.execute(
                select(EventAction.id)
                .select_from(EventAction)
                .join(Event, EventAction.event_id == Event.id)
                .limit(1)
            )
        ).scalar_one_or_none()
        return any_join is not None
    except Exception:
        # If diagnostics query fails, don't over-filter.
        return True


def _csv_bytes_to_xlsx_bytes(content: bytes) -> bytes:
    from openpyxl import Workbook

    text = content.decode("utf-8-sig", errors="replace")
    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    reader = csv.reader(io.StringIO(text), delimiter=";")
    for row in reader:
        ws.append(list(row))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _parse_dt(value: str) -> datetime | None:
    try:
        v = (value or "").strip()
        # Accept UTC ISO strings from frontend: "...Z".
        if v.endswith("Z"):
            v = v[:-1] + "+00:00"
        dt = datetime.fromisoformat(v)
        # DB stores timestamps as naive datetimes (no timezone). If frontend sends
        # tz-aware ISO strings (e.g. trailing 'Z'), normalize to server local time
        # and drop tzinfo so comparisons match stored values.
        if dt.tzinfo is not None:
            return dt.astimezone().replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _append_object_id_filter(filters: list[Any], object_id: str | None) -> None:
    if not object_id:
        return

    oid_variants = _query_variants(object_id)
    if len(oid_variants) == 1:
        filters.append(Event.object_id == oid_variants[0])
    elif oid_variants:
        filters.append(or_(*[Event.object_id == v for v in oid_variants]))


def _append_search_filter(filters: list[Any], search: str | None) -> None:
    if not search or not search.strip():
        return

    token_clauses: list[Any] = []
    for token in tokenize_query(search):
        variants = query_variants(token)
        prefix_needles = query_prefix_needles(token)
        contains_needles = query_needles(token)
        light_search = should_use_light_search(token)
        search_related = should_search_related_text(token)

        like_clauses: list[Any] = []
        for variant in variants:
            like_clauses.extend(
                [
                    Event.id == variant,
                    Event.object_id == variant,
                    Event.code == variant,
                ]
            )

        for needle in prefix_needles:
            like_clauses.extend(
                [
                    Event.id.ilike(needle),
                    Event.object_id.ilike(needle),
                    Event.object_name.ilike(needle),
                    Event.client_name.ilike(needle),
                    Event.code.ilike(needle),
                ]
            )

        if not light_search:
            for needle in contains_needles:
                like_clauses.extend(
                    [
                        Event.description.ilike(needle),
                        Event.object_name.ilike(needle),
                        Event.client_name.ilike(needle),
                        Event.location.ilike(needle),
                        Event.result_text.ilike(needle),
                        Event.code_text.ilike(needle),
                        Event.state_name.ilike(needle),
                    ]
                )

        if search_related:
            action_exists = exists(
                select(literal(1)).where(
                    and_(
                        EventAction.event_id == Event.id,
                        or_(
                            *[
                                or_(
                                    EventAction.action_name.ilike(needle),
                                    EventAction.operator_name.ilike(needle),
                                    EventAction.computer.ilike(needle),
                                    EventAction.gbr_name.ilike(needle),
                                )
                                for needle in contains_needles
                            ]
                        ),
                    )
                )
            )
            like_clauses.append(action_exists)

        token_clauses.append(or_(*like_clauses))

    if token_clauses:
        filters.append(and_(*token_clauses))


async def _build_event_filters(
    *,
    session: AsyncSession,
    dateFrom: str | None,
    dateTo: str | None,
    type: str | None,  # noqa: A002
    objectId: str | None,
    severity: str | None,
    status: str | None,
    search: str | None,
    includeNoise: bool,
    includeSystem: bool,
    includeCancelled: bool,
    onlyWithOperatorComment: bool,
    use_default_lookback: bool,
) -> list[Any]:
    filters: list[Any] = []

    try:
        dialect_name = getattr(getattr(session.get_bind(), "dialect", None), "name", None)
    except Exception:
        dialect_name = None

    if not includeNoise:
        filters.append(Event.type != "access")

    if not includeCancelled:
        filters.append(Event.status != "cancelled")

    if onlyWithOperatorComment:
        filters.append(_has_operator_comment_predicate())

    if not includeSystem:
        actions_linked = await _actions_linked_present(session)
        if actions_linked:
            handled_alarm = or_(
                _is_operator_handled_predicate(),
                _operator_action_exists(dialect_name=dialect_name),
            )
            filters.append(or_(Event.type != "alarm", handled_alarm))

    if type:
        filters.append(Event.type == type)
    _append_object_id_filter(filters, objectId)
    if severity:
        filters.append(Event.severity == severity)
    if status:
        filters.append(Event.status == status)

    if (
        use_default_lookback
        and not dateFrom
        and not dateTo
        and not type
        and not objectId
        and not severity
        and not status
        and not (search and search.strip())
        and int(settings.ui_events_default_lookback_hours) > 0
    ):
        dt_from = datetime.utcnow() - timedelta(hours=int(settings.ui_events_default_lookback_hours))
        filters.append(Event.timestamp >= dt_from)

    if dateFrom:
        dt_from = _parse_dt(dateFrom)
        if dt_from:
            filters.append(Event.timestamp >= dt_from)
    if dateTo:
        dt_to = _parse_dt(dateTo)
        if dt_to:
            filters.append(Event.timestamp <= dt_to)

    _append_search_filter(filters, search)
    return filters


async def build_events_raport_xlsx_bytes(
    *,
    dateFrom: str | None,
    dateTo: str | None,
    type: str | None,  # noqa: A002
    objectId: str | None,
    severity: str | None,
    status: str | None,
    search: str | None,
    includeNoise: bool,
    includeSystem: bool,
    includeCancelled: bool,
    onlyWithOperatorComment: bool,
    limit: int,
    session: AsyncSession,
) -> tuple[bytes, int]:
    """Build XLSX bytes for «Рапорт» by current events filters.

    Returns: (xlsx_bytes, events_count)
    """

    filters = await _build_event_filters(
        session=session,
        dateFrom=dateFrom,
        dateTo=dateTo,
        type=type,
        objectId=objectId,
        severity=severity,
        status=status,
        search=search,
        includeNoise=includeNoise,
        includeSystem=includeSystem,
        includeCancelled=includeCancelled,
        onlyWithOperatorComment=onlyWithOperatorComment,
        use_default_lookback=False,
    )

    where = and_(*filters) if filters else None

    stmt: Select[tuple[Event]] = select(Event).order_by(Event.timestamp.desc()).limit(limit)
    if where is not None:
        stmt = stmt.where(where)
    events_rows: list[Event] = (await session.execute(stmt)).scalars().all()

    event_ids = [str(e.id) for e in events_rows if e.id]
    object_ids = sorted({str(e.object_id) for e in events_rows if getattr(e, "object_id", None)}, key=str)

    # Aggregate event_actions per event_id: called/arrived/cancelled + GBR + operator.
    called_match_strict = or_(
        EventAction.action_name.ilike("%Вызван%груп%"),
        EventAction.action_name.ilike("%Вызван%реаг%"),
        EventAction.action_name.ilike("%Вызван%ГБР%"),
        EventAction.action_name.ilike("%Вызов%груп%"),
        EventAction.action_name.ilike("%Вызов%реаг%"),
        EventAction.action_name.ilike("%Вызов%ГБР%"),
    )
    called_match_loose = EventAction.action_name.ilike("%Вызван%")
    arrived_match_strict = or_(
        EventAction.action_name.ilike("%Приб%груп%"),
        EventAction.action_name.ilike("%Приб%реаг%"),
        EventAction.action_name.ilike("%Приб%ГБР%"),
    )
    arrived_match_loose = or_(
        EventAction.action_name.ilike("%Прибыт%"),
        EventAction.action_name.ilike("%Прибыл%"),
    )
    cancelled_match_strict = or_(
        EventAction.action_name.ilike("%Отмен%груп%"),
        EventAction.action_name.ilike("%Отмен%реаг%"),
        EventAction.action_name.ilike("%Отмен%ГБР%"),
    )
    cancelled_match_loose = EventAction.action_name.ilike("%Отмен%")

    called_ts = func.coalesce(
        func.min(case((called_match_strict, EventAction.action_time), else_=None)),
        func.min(case((called_match_loose, EventAction.action_time), else_=None)),
    ).label("called_ts")
    arrived_ts = func.coalesce(
        func.min(case((arrived_match_strict, EventAction.action_time), else_=None)),
        func.min(case((arrived_match_loose, EventAction.action_time), else_=None)),
    ).label("arrived_ts")
    cancelled_ts = func.coalesce(
        func.min(case((cancelled_match_strict, EventAction.action_time), else_=None)),
        func.min(case((cancelled_match_loose, EventAction.action_time), else_=None)),
    ).label("cancelled_ts")
    called_operator = func.coalesce(
        func.min(case((called_match_strict, EventAction.operator_name), else_=None)),
        func.min(case((called_match_loose, EventAction.operator_name), else_=None)),
    ).label("called_operator")
    gbr_name = func.coalesce(
        func.min(case((called_match_strict, EventAction.gbr_name), else_=None)),
        func.min(case((called_match_loose, EventAction.gbr_name), else_=None)),
        func.min(EventAction.gbr_name),
    ).label("gbr_name")

    actions_by_event: dict[str, dict[str, Any]] = {}
    if event_ids:
        actions_q = (
            select(
                EventAction.event_id,
                gbr_name,
                called_ts,
                arrived_ts,
                cancelled_ts,
                called_operator,
            )
            .where(EventAction.event_id.in_(event_ids))
            .group_by(EventAction.event_id)
        )
        for (eid, gbr, called, arrived, cancelled, op) in (await session.execute(actions_q)).all():
            actions_by_event[str(eid)] = {
                "gbr": gbr,
                "called": called,
                "arrived": arrived,
                "cancelled": cancelled,
                "operator": op,
            }

    # Count alarms for the last 6 months per object.
    half_year_counts: dict[str, int] = {}
    if object_ids:
        cutoff = datetime.utcnow() - timedelta(days=183)
        cnt_q = (
            select(Event.object_id, func.count())
            .where(Event.object_id.in_(object_ids))
            .where(Event.timestamp >= cutoff)
            .where(Event.type == "alarm")
            .group_by(Event.object_id)
        )
        for (oid, cnt) in (await session.execute(cnt_q)).all():
            if oid is not None:
                half_year_counts[str(oid)] = int(cnt or 0)

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Рапорт"

    columns = [
        "№ объекта",
        "Адрес",
        "Шлейф",
        "Инженер",
        "Результат",
        "Дата",
        "ГБР",
        "Вызов",
        "Прибыл",
        "Время в пути",
        "Результат осмотра",
        "Оператор",
        "Заявка",
        "Штраф",
        "Сработок за полгода",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value="Рапорт").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    dt_from = _parse_dt(dateFrom) if dateFrom else None
    dt_to = _parse_dt(dateTo) if dateTo else None
    period_text = ""
    if dt_from and dt_to:
        period_text = f"За период: {dt_from.strftime('%d.%m.%Y %H:%M')} — {dt_to.strftime('%d.%m.%Y %H:%M')}"
    elif dt_from:
        period_text = f"С: {dt_from.strftime('%d.%m.%Y %H:%M')}"
    elif dt_to:
        period_text = f"До: {dt_to.strftime('%d.%m.%Y %H:%M')}"

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws.cell(row=2, column=1, value=period_text).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(columns))
    ws.cell(row=3, column=1, value="оперативная обстановка следующая:").alignment = Alignment(
        horizontal="center"
    )

    header_row = 5
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = [12, 36, 10, 18, 22, 12, 18, 18, 18, 12, 22, 18, 18, 12, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col_idx, title in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=title)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    start_row = header_row + 1
    for i, e in enumerate(events_rows, start=0):
        row_idx = start_row + i

        a = actions_by_event.get(str(e.id)) or {}
        called = a.get("called")
        arrived = a.get("arrived")
        cancelled = a.get("cancelled")

        travel_seconds: float | None = None
        if isinstance(called, datetime) and isinstance(arrived, datetime):
            try:
                travel_seconds = float((arrived - called).total_seconds())
            except Exception:
                travel_seconds = None

        shleif = _extract_desc_value(e.description, "Шлейф")
        engineer = _extract_desc_value(e.description, "Инженер")
        osmotr = _extract_desc_value(e.description, "Осмотр")
        result_osmotr = _extract_desc_value(e.description, "Результат")
        zayavka = _extract_desc_value(e.description, "Заявка")
        shtraf = _extract_desc_value(e.description, "Штраф")

        operator = (
            (str(a.get("operator") or "").strip())
            or _extract_desc_value(e.description, "Оператор")
            or ""
        )
        gbr = (str(a.get("gbr") or "").strip()) or _extract_desc_value(e.description, "ГБР") or ""

        result_main = (getattr(e, "result_text", None) or "").strip()
        if not result_main:
            result_main = _extract_desc_value(e.description, "Заметки")

        date_cell = _fmt_date_ru(called if isinstance(called, datetime) else e.timestamp)

        arrived_cell = ""
        if isinstance(arrived, datetime):
            arrived_cell = _fmt_dt_ru(arrived)
        elif isinstance(cancelled, datetime):
            arrived_cell = "Отмена"

        values = [
            getattr(e, "object_id", None) or "",
            getattr(e, "location", None) or "",
            shleif,
            engineer,
            result_main,
            date_cell,
            gbr,
            _fmt_dt_ru(called if isinstance(called, datetime) else None),
            arrived_cell,
            _fmt_travel(travel_seconds),
            (result_osmotr or osmotr),
            operator,
            zayavka,
            shtraf,
            str(half_year_counts.get(str(getattr(e, "object_id", "") or ""), 0) or 0),
        ]

        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = ws["A6"]
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    out = BytesIO()
    wb.save(out)
    return (out.getvalue(), len(events_rows))


async def build_alarm_messages_xlsx_bytes(
    *,
    dateFrom: str | None,
    dateTo: str | None,
    type: str | None,  # noqa: A002
    objectId: str | None,
    severity: str | None,
    status: str | None,
    search: str | None,
    includeNoise: bool,
    includeSystem: bool,
    includeCancelled: bool,
    onlyWithOperatorComment: bool,
    limit: int,
    session: AsyncSession,
) -> tuple[bytes, int]:
    filters = await _build_event_filters(
        session=session,
        dateFrom=dateFrom,
        dateTo=dateTo,
        type=type,
        objectId=objectId,
        severity=severity,
        status=status,
        search=search,
        includeNoise=includeNoise,
        includeSystem=includeSystem,
        includeCancelled=includeCancelled,
        onlyWithOperatorComment=onlyWithOperatorComment,
        use_default_lookback=False,
    )

    where = and_(*filters) if filters else None

    stmt: Select[tuple[Event]] = select(Event).order_by(Event.timestamp.desc()).limit(limit)
    if where is not None:
        stmt = stmt.where(where)
    events_rows: list[Event] = (await session.execute(stmt)).scalars().all()

    event_ids = [e.id for e in events_rows if e.id]
    object_ids = sorted({e.object_id for e in events_rows if getattr(e, "object_id", None)})

    called_match_strict = or_(
        EventAction.action_name.ilike("%Вызван%груп%"),
        EventAction.action_name.ilike("%Вызван%реаг%"),
        EventAction.action_name.ilike("%Вызван%ГБР%"),
        EventAction.action_name.ilike("%Вызов%груп%"),
        EventAction.action_name.ilike("%Вызов%реаг%"),
        EventAction.action_name.ilike("%Вызов%ГБР%"),
        EventAction.action_name.ilike("%Направ%груп%"),
        EventAction.action_name.ilike("%Направ%реаг%"),
        EventAction.action_name.ilike("%Направ%ГБР%"),
        EventAction.action_name.ilike("%Отправ%груп%"),
        EventAction.action_name.ilike("%Отправ%реаг%"),
        EventAction.action_name.ilike("%Отправ%ГБР%"),
        EventAction.action_name.ilike("%Выезд%груп%"),
        EventAction.action_name.ilike("%Выезд%реаг%"),
        EventAction.action_name.ilike("%Выезд%ГБР%"),
    )
    called_match_loose = or_(
        EventAction.action_name.ilike("%Вызван%"),
        EventAction.action_name.ilike("%Направ%"),
        EventAction.action_name.ilike("%Отправ%"),
        EventAction.action_name.ilike("%Выезд%"),
        EventAction.action_name.ilike("%Следу%"),
    )
    arrived_match_strict = or_(
        EventAction.action_name.ilike("%Приб%груп%"),
        EventAction.action_name.ilike("%Приб%реаг%"),
        EventAction.action_name.ilike("%Приб%ГБР%"),
        EventAction.action_name.ilike("%На объект%"),
        EventAction.action_name.ilike("%Доех%объект%"),
    )
    arrived_match_loose = or_(
        EventAction.action_name.ilike("%Прибыт%"),
        EventAction.action_name.ilike("%Прибыл%"),
        EventAction.action_name.ilike("%На объект%"),
        EventAction.action_name.ilike("%Доех%"),
    )
    cancelled_match_strict = or_(
        EventAction.action_name.ilike("%Отмен%груп%"),
        EventAction.action_name.ilike("%Отмен%реаг%"),
        EventAction.action_name.ilike("%Отмен%ГБР%"),
        EventAction.action_name.ilike("%Отбой%груп%"),
        EventAction.action_name.ilike("%Отбой%реаг%"),
        EventAction.action_name.ilike("%Отбой%ГБР%"),
    )
    cancelled_match_loose = or_(
        EventAction.action_name.ilike("%Отмен%"),
        EventAction.action_name.ilike("%Отбой%"),
        EventAction.action_name.ilike("%Ложн%тревог%"),
        EventAction.action_name.ilike("%Ложный%"),
    )

    called_ts = func.coalesce(
        func.min(case((called_match_strict, EventAction.action_time), else_=None)),
        func.min(case((called_match_loose, EventAction.action_time), else_=None)),
    ).label("called_ts")
    arrived_ts = func.coalesce(
        func.min(case((arrived_match_strict, EventAction.action_time), else_=None)),
        func.min(case((arrived_match_loose, EventAction.action_time), else_=None)),
    ).label("arrived_ts")
    cancelled_ts = func.coalesce(
        func.min(case((cancelled_match_strict, EventAction.action_time), else_=None)),
        func.min(case((cancelled_match_loose, EventAction.action_time), else_=None)),
    ).label("cancelled_ts")
    accepted_ts = func.min(case((_accept_action_predicate(), EventAction.action_time), else_=None)).label("accepted_ts")
    called_operator = func.coalesce(
        func.min(case((called_match_strict, EventAction.operator_name), else_=None)),
        func.min(case((called_match_loose, EventAction.operator_name), else_=None)),
        func.min(case((_accept_action_predicate(), EventAction.operator_name), else_=None)),
    ).label("called_operator")
    gbr_name = func.coalesce(
        func.min(case((called_match_strict, EventAction.gbr_name), else_=None)),
        func.min(case((called_match_loose, EventAction.gbr_name), else_=None)),
        func.min(EventAction.gbr_name),
    ).label("gbr_name")

    actions_by_event: dict[str, dict[str, Any]] = {}
    if event_ids:
        actions_q = (
            select(
                EventAction.event_id,
                gbr_name,
                called_ts,
                arrived_ts,
                cancelled_ts,
                accepted_ts,
                called_operator,
            )
            .where(EventAction.event_id.in_(event_ids))
            .group_by(EventAction.event_id)
        )
        for (eid, gbr, called, arrived, cancelled, accepted, op) in (await session.execute(actions_q)).all():
            actions_by_event[str(eid)] = {
                "gbr": gbr,
                "called": called,
                "arrived": arrived,
                "cancelled": cancelled,
                "accepted": accepted,
                "operator": op,
            }

    objects_by_id: dict[str, Object] = {}
    if object_ids:
        objects = (await session.execute(select(Object).where(Object.id.in_(object_ids)))).scalars().all()
        objects_by_id = {str(obj.id): obj for obj in objects}

    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Тревожные сообщения"

    def clean_excel_text(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, str):
            return re.sub(ILLEGAL_CHARACTERS_RE, "", value)
        return value

    columns = [
        "Дата тревоги",
        "№ Объекта",
        "Адрес объекта",
        "ФИО",
        "Принят",
        "Дата принятия",
        "Система",
        "Шлейф",
        "Инженер",
        "ГБР",
        "Оператор",
        "Время вызова",
        "Время прибытия",
        "Результат",
        "Результат осмотра",
        "Заметки",
        "Заявка",
        "Выполнена",
        "Результат заявки",
        "Штраф",
        "№ квитанции",
        "Пропажа",
        "Устранена",
        "Договорной отдел",
        "Время в пути",
        "Пропажи",
        "Ложная сработка",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value=clean_excel_text("Тревожные сообщения")).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    dt_from = _parse_dt(dateFrom) if dateFrom else None
    dt_to = _parse_dt(dateTo) if dateTo else None
    period_text = ""
    if dt_from and dt_to:
        period_text = f"За период: {dt_from.strftime('%d.%m.%Y %H:%M')} — {dt_to.strftime('%d.%m.%Y %H:%M')}"
    elif dt_from:
        period_text = f"С: {dt_from.strftime('%d.%m.%Y %H:%M')}"
    elif dt_to:
        period_text = f"До: {dt_to.strftime('%d.%m.%Y %H:%M')}"

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws.cell(row=2, column=1, value=clean_excel_text(period_text)).alignment = Alignment(horizontal="center")

    header_row = 4
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = [
        20, 14, 38, 26, 10, 20, 16, 10, 16, 12, 20, 12, 12, 52, 32, 24, 10, 12, 22, 10, 14, 10, 12, 16, 14, 12, 14,
    ]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col_idx, title in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=clean_excel_text(title))
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    def _fmt_time_ru(dt: datetime | None) -> str:
        if dt is None:
            return ""
        return dt.strftime("%H:%M:%S")

    def _flag_text(value: Any, default: str = "") -> str:
        if value is None:
            return default
        if isinstance(value, bool):
            return "1" if value else "0"
        text = str(value).strip()
        if not text:
            return default
        low = text.lower()
        if low in {"1", "да", "true", "yes", "+"}:
            return "1"
        if low in {"0", "нет", "false", "no", "-"}:
            return "0"
        return text

    start_row = header_row + 1
    for i, e in enumerate(events_rows, start=0):
        row_idx = start_row + i

        a = actions_by_event.get(str(e.id)) or {}
        called = a.get("called")
        arrived = a.get("arrived")
        cancelled = a.get("cancelled")
        accepted = a.get("accepted")
        obj = objects_by_id.get(str(getattr(e, "object_id", "") or ""))

        travel_seconds: float | None = None
        if isinstance(called, datetime) and isinstance(arrived, datetime):
            try:
                travel_seconds = float((arrived - called).total_seconds())
            except Exception:
                travel_seconds = None

        shleif = _extract_desc_value(e.description, "Шлейф")
        engineer = _extract_desc_value(e.description, "Инженер")
        system_name = _extract_desc_value(e.description, "Система")
        osmotr = _extract_desc_value(e.description, "Осмотр")
        result_osmotr = _extract_desc_value(e.description, "Результат осмотра") or _extract_desc_value(e.description, "Результат")
        notes = _extract_desc_value(e.description, "Заметки")
        zayavka = _extract_desc_value(e.description, "Заявка")
        result_zayavka = _extract_desc_value(e.description, "Результат заявки")
        shtraf = _extract_desc_value(e.description, "Штраф")
        receipt_no = _extract_desc_value(e.description, "№ квитанции") or _extract_desc_value(e.description, "Квитанция")
        contract_department = _extract_desc_value(e.description, "Договорной отдел")
        missing_count = _extract_desc_value(e.description, "Пропажи")
        missing_flag = _extract_desc_value(e.description, "Пропажа")
        eliminated_flag = _extract_desc_value(e.description, "Устранена")
        completed_flag = _extract_desc_value(e.description, "Выполнена")
        false_alarm = _extract_desc_value(e.description, "Ложная сработка")

        operator = ((str(a.get("operator") or "").strip()) or _extract_desc_value(e.description, "Оператор") or "")
        gbr = (str(a.get("gbr") or "").strip()) or _extract_desc_value(e.description, "ГБР") or ""
        result_main = (getattr(e, "result_text", None) or "").strip()
        if not result_main:
            result_main = notes

        combined_text = " ".join(
            [
                str(getattr(e, "result_text", None) or ""),
                str(getattr(e, "description", None) or ""),
                str(getattr(e, "code_text", None) or ""),
            ]
        ).lower()

        if not false_alarm and "ложн" in combined_text:
            false_alarm = "1"
        if not missing_flag and "пропаж" in combined_text:
            missing_flag = "1"
        if not completed_flag and getattr(e, "status", None) == "resolved":
            completed_flag = "1"
        if not eliminated_flag and getattr(e, "status", None) == "resolved":
            eliminated_flag = "1"

        values = [
            _fmt_dt_ru(getattr(e, "timestamp", None)),
            getattr(e, "object_id", None) or "",
            getattr(e, "location", None) or getattr(e, "object_name", None) or "",
            getattr(e, "client_name", None) or "",
            "1" if accepted or operator else "0",
            _fmt_dt_ru(accepted if isinstance(accepted, datetime) else getattr(obj, "created_at", None)),
            system_name,
            shleif,
            engineer,
            gbr,
            operator,
            _fmt_time_ru(called if isinstance(called, datetime) else getattr(e, "timestamp", None)),
            _fmt_time_ru(arrived) if isinstance(arrived, datetime) else ("Отмена" if isinstance(cancelled, datetime) else ""),
            result_main,
            result_osmotr or osmotr,
            notes,
            _flag_text(zayavka, "0"),
            _flag_text(completed_flag, "0"),
            result_zayavka,
            _flag_text(shtraf, "0"),
            receipt_no,
            _flag_text(missing_flag, "0"),
            _flag_text(eliminated_flag, "0"),
            _flag_text(contract_department, "0"),
            _fmt_travel(travel_seconds),
            missing_count,
            _flag_text(false_alarm, "0"),
        ]

        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=clean_excel_text(v))
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = ws["A5"]
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    out = BytesIO()
    wb.save(out)
    return (out.getvalue(), len(events_rows))


def _event_to_out(e: Event) -> dict[str, Any]:
    return {
        "id": e.id,
        "parentEventId": getattr(e, "parent_event_id", None),
        "timestamp": e.timestamp.isoformat(),
        "type": e.type,
        "objectId": e.object_id,
        "objectName": e.object_name,
        "clientName": e.client_name,
        "severity": e.severity,
        "status": e.status,
        "code": getattr(e, "code", None),
        "codeText": getattr(e, "code_text", None),
        "stateName": getattr(e, "state_name", None),
        "resultText": getattr(e, "result_text", None),
        "meterCount": getattr(e, "meter_count", None),
        "timeMeterCount": (
            getattr(e, "time_meter_count", None).isoformat()
            if getattr(e, "time_meter_count", None) is not None
            else None
        ),
        "description": e.description,
        "location": e.location,
        "operatorId": e.operator_id,
    }


def _action_to_out(a: EventAction) -> dict[str, Any]:
    return {
        "actionName": a.action_name,
        "actionTime": a.action_time.isoformat(),
        "operatorName": a.operator_name,
        "computer": a.computer,
        "gbrName": a.gbr_name,
        "dateKey": a.date_key,
        "rawEventId": a.raw_event_id,
        "sourceTable": a.source_table,
        "sourcePk": a.source_pk,
    }


def _parse_agency_event_key(event_id: str) -> tuple[int, int] | None:
    parts = str(event_id).split(":")
    if len(parts) < 3:
        return None
    try:
        date_key = int(parts[-2])
        raw_event_id = int(parts[-1])
    except Exception:
        return None
    return (date_key, raw_event_id)


def _eventservice_source_table(date_key: int) -> str:
    s = str(int(date_key))
    if len(s) != 8:
        suffix = s[:6] + "01"
    else:
        suffix = s[:6] + "01"
    return f"eventservice{suffix}"


def _coerce_dt(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value)
        except Exception:
            return None
    return None


def _extract_desc_value(description: str | None, label: str) -> str:
    if not description:
        return ""
    want = (label or "").strip().lower()
    if not want:
        return ""
    for raw in str(description).splitlines():
        line = raw.strip()
        if not line:
            continue
        low = line.lower()
        if low.startswith(want) and ":" in line:
            head, tail = line.split(":", 1)
            if head.strip().lower() == want:
                return tail.strip()
    return ""


def _fmt_dt_ru(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%d.%m.%Y %H:%M:%S")


def _fmt_date_ru(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%d.%m.%Y")


def _fmt_travel(seconds: float | int | None) -> str:
    try:
        if seconds is None:
            return ""
        total = int(round(float(seconds)))
        if total < 0:
            total = abs(total)
        h = total // 3600
        m = (total % 3600) // 60
        s = total % 60
        return f"{h:d}:{m:02d}:{s:02d}" if h > 0 else f"{m:d}:{s:02d}"
    except Exception:
        return ""


def _action_row_to_out(row: dict[str, Any]) -> dict[str, Any] | None:
    action_name = str(row.get("NameState") or "").strip()
    if not action_name:
        return None

    action_time = _coerce_dt(row.get("OperationTime"))
    if action_time is None:
        return None

    try:
        date_key = int(row.get("Date_Key"))
        raw_event_id = int(row.get("Event_id"))
        source_pk = int(row.get("Service_id"))
    except Exception:
        return None

    return {
        "actionName": action_name,
        "actionTime": action_time.isoformat(),
        "operatorName": str(row.get("PersonName") or "").strip() or None,
        "computer": str(row.get("Computer") or "").strip() or None,
        "gbrName": str(row.get("GrResponseName") or "").strip() or None,
        "dateKey": date_key,
        "rawEventId": raw_event_id,
        "sourceTable": _eventservice_source_table(date_key),
        "sourcePk": source_pk,
    }


@router.get("/details")
async def get_event_details_by_query(
    eventId: str = Query(..., min_length=1),
    actionsLimit: int = Query(500, ge=0, le=5000),
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    # Use query param to avoid putting IDs like "mssql:YYYYMMDD:NNN" into the URL path.
    return await get_event_details(event_id=eventId, actionsLimit=actionsLimit, session=session)


@router.get("/{event_id}")
async def get_event_details(
    event_id: str,
    actionsLimit: int = Query(500, ge=0, le=5000),
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    e = await session.get(Event, event_id)
    if not e:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "Event not found"})

    actions: list[EventAction] = []
    actions_meta: dict[str, Any] = {
        "eventId": event_id,
        "eventTimestamp": e.timestamp.isoformat(),
        "foundInDb": False,
        "attemptedKey": None,
        "attemptedKeys": [],
        "fallback": None,
        "fallbackError": None,
    }
    if actionsLimit > 0:
        stmt: Select[tuple[EventAction]] = (
            select(EventAction)
            .where(EventAction.event_id == event_id)
            .order_by(EventAction.action_time.asc())
            .limit(actionsLimit)
        )
        actions = (await session.execute(stmt)).scalars().all()

        # Fallback for numeric event IDs: link by (date_key, raw_event_id).
        if not actions:
            try:
                raw_event_id = int(str(event_id))
            except Exception:
                raw_event_id = None

            if raw_event_id is not None:
                from datetime import timedelta

                base = e.timestamp.date()
                date_keys = [
                    int(base.strftime("%Y%m%d")),
                    int((base - timedelta(days=1)).strftime("%Y%m%d")),
                    int((base + timedelta(days=1)).strftime("%Y%m%d")),
                ]
                actions_meta["attemptedKeys"] = [(dk, raw_event_id) for dk in date_keys]

                for dk in date_keys:
                    stmt2: Select[tuple[EventAction]] = (
                        select(EventAction)
                        .where(EventAction.raw_event_id == raw_event_id)
                        .where(EventAction.date_key == dk)
                        .order_by(EventAction.action_time.asc())
                        .limit(actionsLimit)
                    )
                    actions = (await session.execute(stmt2)).scalars().all()
                    if actions:
                        actions_meta["attemptedKey"] = (dk, raw_event_id)
                        actions_meta["foundInDb"] = True
                        break

        if actions:
            return {
                "event": _event_to_out(e),
                "actions": [_action_to_out(a) for a in actions],
                "actionsMeta": actions_meta,
            }

    # Best-effort fallback: if actions are not synced into SVOD DB yet,
    # try to fetch them directly from the agency DB by (Date_Key, Event_id).
    if actionsLimit > 0 and not actions and settings.agency_database_url:
        key = _parse_agency_event_key(event_id)
        if key is None:
            # Numeric event IDs: use timestamp day as Date_Key.
            try:
                raw_event_id = int(str(event_id))
                from datetime import timedelta

                base = e.timestamp.date()
                date_keys = [
                    int(base.strftime("%Y%m%d")),
                    int((base - timedelta(days=1)).strftime("%Y%m%d")),
                    int((base + timedelta(days=1)).strftime("%Y%m%d")),
                ]
                keys = [(dk, raw_event_id) for dk in date_keys]
            except Exception:
                keys = []

            if key is not None:
                keys = [key]

            actions_meta["attemptedKeys"] = keys

            url = settings.agency_database_url
            scheme = (url.split(":", 1)[0] or "").lower()
            actions_meta["fallback"] = {
                "scheme": scheme,
                "archivesDb": settings.agency_archives_db_name,
                "agencyUrlHost": (url.split("@", 1)[1].split("/", 1)[0] if "@" in url else None),
            }

            rows: list[dict[str, Any]] = []
            last_error: str | None = None
            for k in keys:
                try:
                    if scheme.startswith("sqlite"):
                        rows = await asyncio.to_thread(
                            fetch_eventservice_actions_for_event_pairs_sqlite,
                            url,
                            event_pairs=[k],
                        )
                    elif scheme.startswith("mssql"):
                        rows = await asyncio.to_thread(
                            fetch_eventservice_actions_for_event_pairs_mssql,
                            url,
                            archives_db_name=settings.agency_archives_db_name,
                            event_pairs=[k],
                        )
                    else:
                        rows = []
                except Exception as ex:
                    last_error = str(ex)
                    rows = []

                if rows:
                    actions_meta["attemptedKey"] = k
                    break

            if last_error:
                actions_meta["fallbackError"] = last_error
                logger.warning(
                    "event_actions fallback error: event_id=%s keys=%s scheme=%s archives_db=%s err=%s",
                    event_id,
                    keys,
                    scheme,
                    settings.agency_archives_db_name,
                    last_error,
                )
            elif not rows:
                logger.info(
                    "event_actions fallback empty: event_id=%s keys=%s scheme=%s archives_db=%s",
                    event_id,
                    keys,
                    scheme,
                    settings.agency_archives_db_name,
                )

            out_rows: list[dict[str, Any]] = []
            for r in rows[:actionsLimit]:
                mapped = _action_row_to_out(r)
                if mapped is not None:
                    out_rows.append(mapped)

            if out_rows:
                return {"event": _event_to_out(e), "actions": out_rows, "actionsMeta": actions_meta}

    return {"event": _event_to_out(e), "actions": [], "actionsMeta": actions_meta}


@router.patch("/operator-comment")
async def upsert_operator_comment(
    payload: OperatorCommentIn,
    session: AsyncSession = Depends(get_session),
    current: dict = Depends(get_current_user),
) -> dict[str, Any]:
    e = await session.get(Event, payload.eventId)
    if not e:
        raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "Event not found"})

    text = (payload.resultText or "").strip()
    e.result_text = text if text else None

    # Best-effort: if comment is added and operator_id is empty, tag it
    # with current username (used by some UI heuristics).
    if text and not (str(getattr(e, "operator_id", "") or "").strip()):
        who = str(current.get("username") or "").strip() or str(current.get("id") or "").strip()
        if who:
            e.operator_id = who

    await session.commit()
    await session.refresh(e)
    return {"event": _event_to_out(e)}


@router.get("")
async def list_events(
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=500),
    dateFrom: str | None = None,
    dateTo: str | None = None,
    type: str | None = None,  # noqa: A002
    objectId: str | None = None,
    severity: str | None = None,
    status: str | None = None,
    search: str | None = None,
    includeNoise: bool = Query(False, description="Include access/noise events (arming/disarming, etc.)"),
    includeSystem: bool = Query(False, description="Include system-handled events (no operator)"),
    includeCancelled: bool = Query(False, description="Include cancelled events"),
    onlyWithOperatorComment: bool = Query(
        False,
        description="Show only events that have an operator comment (Result_Text)",
    ),
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    search_active = bool(search and search.strip())
    filters = await _build_event_filters(
        session=session,
        dateFrom=dateFrom,
        dateTo=dateTo,
        type=type,
        objectId=objectId,
        severity=severity,
        status=status,
        search=search,
        includeNoise=includeNoise,
        includeSystem=includeSystem,
        includeCancelled=includeCancelled,
        onlyWithOperatorComment=onlyWithOperatorComment,
        use_default_lookback=True,
    )

    where = and_(*filters) if filters else None

    stmt: Select[tuple[Event]] = select(Event).order_by(Event.timestamp.desc())
    if where is not None:
        stmt = stmt.where(where)
    offset = (page - 1) * pageSize

    if search_active:
        stmt = stmt.offset(offset).limit(pageSize + 1)
        raw_rows = (await session.execute(stmt)).scalars().all()
        has_more = len(raw_rows) > pageSize
        rows = raw_rows[:pageSize]
        total = offset + len(rows) + (1 if has_more else 0)
        total_pages = page + (1 if has_more else 0)
    else:
        count_stmt = select(func.count()).select_from(Event)
        if where is not None:
            count_stmt = count_stmt.where(where)
        total = (await session.execute(count_stmt)).scalar_one()

        stmt = stmt.offset(offset).limit(pageSize)
        rows = (await session.execute(stmt)).scalars().all()
        total_pages = (total + pageSize - 1) // pageSize if pageSize else 1

    page_items = [_event_to_out(e) for e in rows]

    return {
        "data": page_items,
        "total": total,
        "page": page,
        "pageSize": pageSize,
        "totalPages": total_pages,
        "totalIsEstimate": search_active,
    }


@router.get("/export")
async def export_events_export(
    dateFrom: str | None = None,
    dateTo: str | None = None,
    type: str | None = None,  # noqa: A002
    objectId: str | None = None,
    severity: str | None = None,
    status: str | None = None,
    search: str | None = None,
    includeNoise: bool = Query(False, description="Include access/noise events (arming/disarming, etc.)"),
    includeSystem: bool = Query(False, description="Include system-handled events (no operator)"),
    includeCancelled: bool = Query(False, description="Include cancelled events"),
    onlyWithOperatorComment: bool = Query(
        False,
        description="Show only events that have an operator comment (Result_Text)",
    ),
    limit: int = Query(50000, ge=1, le=200000),
    session: AsyncSession = Depends(get_session),
) -> Response:
    filters = await _build_event_filters(
        session=session,
        dateFrom=dateFrom,
        dateTo=dateTo,
        type=type,
        objectId=objectId,
        severity=severity,
        status=status,
        search=search,
        includeNoise=includeNoise,
        includeSystem=includeSystem,
        includeCancelled=includeCancelled,
        onlyWithOperatorComment=onlyWithOperatorComment,
        use_default_lookback=False,
    )

    where = and_(*filters) if filters else None

    stmt: Select[tuple[Event]] = select(Event).order_by(Event.timestamp.desc()).limit(limit)
    if where is not None:
        stmt = stmt.where(where)
    rows = (await session.execute(stmt)).scalars().all()

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "События"

    def _agency_event_id(event_id: str | None) -> str:
        if not event_id:
            return ""
        parts = str(event_id).split(":")
        if len(parts) >= 3:
            return parts[-1] or ""
        return ""

    headers = [
        "ID (SVOD)",
        "ID события (аг.)",
        "Дата/время",
        "Тип",
        "Номер объекта",
        "Название объекта",
        "Контрагент",
        "Код",
        "Расшифровка кода",
        "Статус (агентство)",
        "Важность",
        "Статус",
        "Адрес",
        "Параметр (MeterCount)",
        "Время параметра (TimeMeterCount)",
        "Пометка (Result_Text)",
        "Описание",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for e in rows:
        ws.append(
            [
                e.id,
                _agency_event_id(e.id),
                e.timestamp.isoformat(),
                e.type,
                e.object_id or "",
                e.object_name,
                e.client_name,
                getattr(e, "code", "") or "",
                getattr(e, "code_text", "") or "",
                getattr(e, "state_name", "") or "",
                e.severity,
                e.status,
                e.location or "",
                getattr(e, "meter_count", "") or "",
                (
                    getattr(e, "time_meter_count", None).isoformat()
                    if getattr(e, "time_meter_count", None) is not None
                    else ""
                ),
                getattr(e, "result_text", "") or "",
                (e.description or "").replace("\r\n", "\n"),
            ]
        )

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 40
    ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 40
    ws.column_dimensions["N"].width = 40
    ws.column_dimensions["O"].width = 22
    ws.column_dimensions["P"].width = 22
    ws.column_dimensions["Q"].width = 40
    ws.column_dimensions["R"].width = 80

    xlsx = io.BytesIO()
    wb.save(xlsx)

    filename = f"events-export-{datetime.utcnow().date().isoformat()}.xlsx"
    return Response(
        content=xlsx.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/xlsx")
async def export_events_xlsx(
    dateFrom: str | None = None,
    dateTo: str | None = None,
    type: str | None = None,  # noqa: A002
    objectId: str | None = None,
    severity: str | None = None,
    status: str | None = None,
    search: str | None = None,
    includeNoise: bool = Query(False, description="Include access/noise events (arming/disarming, etc.)"),
    includeSystem: bool = Query(False, description="Include system-handled events (no operator)"),
    includeCancelled: bool = Query(False, description="Include cancelled events"),
    onlyWithOperatorComment: bool = Query(
        False,
        description="Show only events that have an operator comment (Result_Text)",
    ),
    limit: int = Query(50000, ge=1, le=200000),
    session: AsyncSession = Depends(get_session),
) -> Response:
    return await export_events_export(
        dateFrom=dateFrom,
        dateTo=dateTo,
        type=type,
        objectId=objectId,
        severity=severity,
        status=status,
        search=search,
        includeNoise=includeNoise,
        includeSystem=includeSystem,
        includeCancelled=includeCancelled,
        onlyWithOperatorComment=onlyWithOperatorComment,
        limit=limit,
        session=session,
    )


@router.get("/export/raport/xlsx")
async def export_events_raport_xlsx(
    dateFrom: str | None = None,
    dateTo: str | None = None,
    type: str | None = None,  # noqa: A002
    objectId: str | None = None,
    severity: str | None = None,
    status: str | None = None,
    search: str | None = None,
    includeNoise: bool = Query(False, description="Include access/noise events (arming/disarming, etc.)"),
    includeSystem: bool = Query(False, description="Include system-handled events (no operator)"),
    includeCancelled: bool = Query(False, description="Include cancelled events"),
    onlyWithOperatorComment: bool = Query(
        False,
        description="Show only events that have an operator comment (Result_Text)",
    ),
    limit: int = Query(50000, ge=1, le=200000),
    session: AsyncSession = Depends(get_session),
) -> Response:
    """XLSX: «Рапорт» по текущему списку событий (как в интерфейсе).

    Источники колонок:
    - events: № объекта, адрес, timestamp, result_text
    - event_actions: вызов/прибыл/время в пути, ГБР, оператор
    - events.description: шлейф/инженер/заявка/штраф/результат осмотра
    - events (агрегация): «сработок за полгода» по объекту
    """

    xlsx, _count = await build_events_raport_xlsx_bytes(
        dateFrom=dateFrom,
        dateTo=dateTo,
        type=type,
        objectId=objectId,
        severity=severity,
        status=status,
        search=search,
        includeNoise=includeNoise,
        includeSystem=includeSystem,
        includeCancelled=includeCancelled,
        onlyWithOperatorComment=onlyWithOperatorComment,
        limit=limit,
        session=session,
    )

    filename = f"raport-{datetime.utcnow().date().isoformat()}.xlsx"
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/stream")
async def stream_events(
    since: str | None = Query(None, description="ISO timestamp; stream events newer than this"),
    pollSeconds: float = Query(1.0, ge=0.2, le=10.0),
    includeNoise: bool = Query(False, description="Include access/noise events (arming/disarming, etc.)"),
    includeSystem: bool = Query(False, description="Include system-handled events (no operator)"),
    includeCancelled: bool = Query(False, description="Include cancelled events"),
    onlyWithOperatorComment: bool = Query(
        False,
        description="Stream only events that have an operator comment (Result_Text)",
    ),
    session: AsyncSession = Depends(get_session),
    _current: dict = Depends(get_current_user),
):
    """Server-Sent Events stream for new events.

    This is intentionally simple: it polls SQLite/Postgres for new rows.
    Works well for local/SQLite and small-to-medium throughput.
    """

    last_ts = _parse_dt(since) if since else datetime.utcnow()
    keepalive_every = 15.0

    actions_linked = True
    if not includeSystem:
        try:
            actions_linked = await _actions_linked_present(session)
        except Exception:
            actions_linked = True

    try:
        dialect_name = getattr(getattr(session.get_bind(), "dialect", None), "name", None)
    except Exception:
        dialect_name = None

    async def gen():
        nonlocal last_ts
        last_keepalive = asyncio.get_event_loop().time()

        # Initial hello event helps clients validate connection quickly.
        yield b"event: hello\n"
        yield f"data: {json.dumps({'serverTime': datetime.utcnow().isoformat()})}\n\n".encode("utf-8")

        while True:
            # Query a small batch; client can keep connection open.
            stmt: Select[tuple[Event]] = select(Event).where(Event.timestamp > last_ts)
            if not includeNoise:
                stmt = stmt.where(Event.type != "access")
            if not includeCancelled:
                stmt = stmt.where(Event.status != "cancelled")
            if onlyWithOperatorComment:
                stmt = stmt.where(_has_operator_comment_predicate())
            if not includeSystem:
                if actions_linked:
                    handled_alarm = or_(
                        _is_operator_handled_predicate(),
                        _operator_action_exists(dialect_name=dialect_name),
                    )
                    stmt = stmt.where(or_(Event.type != "alarm", handled_alarm))
            stmt = stmt.order_by(Event.timestamp.asc()).limit(500)
            rows = (await session.execute(stmt)).scalars().all()
            for e in rows:
                payload = _event_to_out(e)
                last_ts = max(last_ts, e.timestamp)
                yield b"event: event\n"
                yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n".encode("utf-8")

            now = asyncio.get_event_loop().time()
            if now - last_keepalive >= keepalive_every:
                last_keepalive = now
                yield b": keep-alive\n\n"

            await asyncio.sleep(pollSeconds)

    return StreamingResponse(
        gen(),
        media_type="text/event-stream; charset=utf-8",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.get("/plain/{event_id}")
async def get_event_plain(
    event_id: str,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    e = await session.get(Event, event_id)
    if e:
        return _event_to_out(e)
    raise HTTPException(status_code=404, detail={"code": "NOT_FOUND", "message": "Event not found"})
